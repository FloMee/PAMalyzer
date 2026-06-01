# Version 3.2-BirdNET 21/03/2024
# Authors: Stephen Marsland, Nirosha Priyadarshani, Julius Juodakis, Virginia Listanti, Florian Meerheim

# This is the main class for the AviaNZ interface.

#    AviaNZ bioacoustic analysis program
#    Copyright (C) 2017--2024

#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.

# ? click, shutil
import copy
import datetime
import fnmatch
import json
import math
import os
import pathlib
import platform
import re
import shutil
import time
import webbrowser

import numpy as np
import openpyxl
import pyqtgraph as pg
import pyqtgraph.exporters as pge
import pyqtgraph.functions as fn
import superqt
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import (
    QDir,
    QLocale,
    QModelIndex,
    QPoint,
    QPointF,
    QRectF,
    Qt,
    QTimer,
)
from PyQt5.QtGui import QIcon, QKeySequence, QPixmap, QStandardItem, QStandardItemModel
from PyQt5.QtMultimedia import QAudio
from PyQt5.QtWidgets import (
    QAbstractSpinBox,
    QActionGroup,
    QApplication,
    QCheckBox,
    QComboBox,
    QDoubleSpinBox,
    QFileDialog,
    QFrame,
    QGraphicsProxyWidget,
    QGroupBox,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMessageBox,
    QPushButton,
    QScrollBar,
    QShortcut,
    QSizePolicy,
    QSpinBox,
    QToolButton,
    QTreeView,
    QVBoxLayout,
    QWidget,
    QWidgetAction,
)
from pyqtgraph.dockarea import Dock, DockArea
from pyqtgraph.parametertree import Parameter
from scipy.ndimage.filters import median_filter

import BirdNET
import colourMaps
import Database
import Dialogs
import Segment
import SignalProc
import SupportClasses
import SupportClasses_GUI
import WaveletFunctions
import wavio

pg.setConfigOption("background", "w")
pg.setConfigOption("foreground", "k")
pg.setConfigOption("antialias", True)
print("Package import complete.")

# import pdb
# from PyQt5.QtCore import pyqtRemoveInputHook
# from pdb import set_trace
#
# def debug_trace():
#     pyqtRemoveInputHook()
#     set_trace()


class AviaNZ(QMainWindow):
    """Main class for the user interface.
    Contains most of the user interface and plotting code"""

    def __init__(self, configdir=None):
        """Initialisation of the class. Load main config and bird lists from configdir.
        Also initialises the data structures and loads an initial file (specified explicitly)
        and sets up the window.
        One interesting configuration point is the DOC setting, which hides the more 'research' functions.
        """
        print("Starting PAMalyzer...")

        super(AviaNZ, self).__init__()

        # configdir passes the standard user app dir based on OS.
        # At this point, the main config file should already be ensured to exist.
        self.configdir = configdir
        self.configfile = os.path.join(configdir, "AviaNZconfig.txt")
        self.ConfigLoader = SupportClasses.ConfigLoader()
        self.config = self.ConfigLoader.config(self.configfile)
        self.saveConfig = True
        self.settingsChanged = False
        self.updateFileIcons = False

        # Load call types
        self.calltypesDir = os.path.join(configdir, "Calltypes")
        self.CalltypeDicts = self.ConfigLoader.calltypes(self.calltypesDir)

        # Load the birdlists - both are now necessary:
        self.shortBirdList = self.ConfigLoader.shortbl(
            self.config["BirdListShort"], configdir
        )
        if self.shortBirdList is None:
            raise OSError("Short bird list missing, cannot continue")
        self.longBirdList = self.ConfigLoader.longbl(
            self.config["BirdListLong"], configdir
        )
        if self.longBirdList is None:
            raise OSError("Long bird list missing, cannot continue")
        self.batList = self.ConfigLoader.batl(self.config["BatList"], configdir)
        if self.batList is None:
            raise OSError("Bat list missing, cannot continue")

        # avoid comma/point problem in number parsing
        # QLocale.setDefault(QLocale(QLocale.English, QLocale.NewZealand))
        print("Locale is set to " + QLocale().name())

        # The data structures for the segments
        self.listLabels = []
        self.listRectanglesa1 = []
        self.listRectanglesa2 = []
        self.SegmentRects = []
        self.segmentPlots = []
        self.shapePlots = []
        self.box1id = -1

        self.started = False
        self.startedInAmpl = False
        self.startTime = 0
        self.segmentsToSave = False
        self.viewCallType = False
        self.viewCertainty = True

        if os.path.exists(self.config["Database"]):
            self.dbPath = self.config["Database"]
        else:
            self.dbPath = os.path.join(configdir, "default.db")

        self.db = Database.DatabaseHandler(self.dbPath)

        # Spectrogram default settings
        # TODO: put in config?
        self.sgOneSided = True
        self.sgMeanNormalise = True
        self.sgEqualLoudness = False
        self.sgType = "Standard"
        self.sgNormMode = "Log"
        self.sgScale = "Linear"
        self.nfilters = 128
        self.windowType = "Hann"

        self.lastSpecies = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]
        self.DOC = self.config["DOC"]
        self.slowSpeed = 2

        # placeholder for BirdNET-Lite and BirdNET-Analyzer
        self.BirdNET = None

        # Whether or not the context menu allows multiple birds.
        self.multipleBirds = self.config["MultipleSpecies"]

        if len(self.config["RecentFiles"]) > 0:
            firstFile = self.config["RecentFiles"][-1]
            self.SoundFileDir = os.path.dirname(firstFile)
            if not os.path.isdir(self.SoundFileDir) or not os.path.isfile(firstFile):
                self.SoundFileDir = os.path.abspath(self.config["SoundFileDir"])
                firstFile = os.path.join(self.SoundFileDir, "kiwi_1min.wav")
        else:
            self.SoundFileDir = os.path.abspath(self.config["SoundFileDir"])
            firstFile = os.path.join(self.SoundFileDir, "kiwi_1min.wav")
        self.SoundFileDirOld = ""
        self.filename = None
        self.focusRegion = None

        self.operator = self.config["operator"]
        self.reviewer = self.config["reviewer"]

        # For preventing callbacks involving overview panel
        self.updateRequestedByOverview = False

        # parse firstFile to dir and file parts
        print("Working dir set to %s" % self.SoundFileDir)
        print("Opening file %s" % firstFile)

        # to keep code simpler, graphic options are created even in CLI mode
        # they're just not shown because QMainWindow.__init__ is skipped
        QMainWindow.__init__(self)

        # parse mouse settings
        if self.config["drawingRightBtn"]:
            self.MouseDrawingButton = Qt.RightButton
        else:
            self.MouseDrawingButton = Qt.LeftButton

        # Boxes with area smaller than this will be ignored -
        # to avoid accidentally creating little boxes
        self.minboxsize = 0.1

        self.createMenu()
        self.createFrame()

        self.resetStorageArrays()

        # Make the window and associated widgets
        self.setWindowTitle("PAMalyzer")
        self.setWindowIcon(QIcon("img/PAMalyzer.ico"))
        # Show the window
        if self.config["StartMaximized"]:
            self.showMaximized()
            # extra toggle because otherwise Windows starts at a non-maximized size
            self.setWindowState(self.windowState() ^ Qt.WindowMaximized)
            self.setWindowState(self.windowState() | Qt.WindowMaximized)
        else:
            self.show()

        # Save the segments every minute
        self.timer = QTimer()
        self.timer.timeout.connect(self.saveSegments)
        self.timer.start(self.config["secsSave"] * 1000)

        self.listLoadFile(os.path.basename(firstFile))

        if self.operator == "" or self.reviewer == "":
            self.setOperatorReviewerDialog()

    def createMenu(self):
        """Create the menu entries at the top of the screen and link them as appropriate.
        Some of them are initialised according to the data in the configuration file."""

        fileMenu = self.menuBar().addMenu("&File")
        openIcon = self.style().standardIcon(QtWidgets.QStyle.SP_DialogOpenButton)
        fileMenu.addAction(openIcon, "&Open sound file", self.openFile, "Ctrl+O")
        # fileMenu.addAction("&Change Directory", self.chDir)
        fileMenu.addAction(
            "Set Operator/Reviewer (Current File)", self.setOperatorReviewerDialog
        )
        fileMenu.addSeparator()
        for recentfile in self.config["RecentFiles"]:
            fileMenu.addAction(recentfile, lambda arg=recentfile: self.openFile(arg))
        fileMenu.addSeparator()
        fileMenu.addAction(
            QIcon(QPixmap("img/exit.png")), "&Quit", QApplication.quit, "Ctrl+Q"
        )

        # This is a very bad way to do this, but I haven't worked anything else out (setMenuRole() didn't work)
        # Add it a second time, then it appears!
        if platform.system() == "Darwin":
            fileMenu.addAction("&Quit", QApplication.quit, "Ctrl+Q")

        viewMenu = self.menuBar().addMenu("&View")

        self.useAmplitudeTick = viewMenu.addAction(
            "Show amplitude plot", self.useAmplitudeCheck
        )
        self.useAmplitudeTick.setCheckable(True)
        self.useAmplitudeTick.setChecked(self.config["showAmplitudePlot"])
        self.useAmplitude = True

        self.useFilesTick = viewMenu.addAction("Show file list", self.useFilesCheck)
        self.useFilesTick.setCheckable(True)
        self.useFilesTick.setChecked(self.config["showListofFiles"])

        # this can go under "Change interface settings"
        self.showOverviewSegsTick = viewMenu.addAction(
            "Show annotation overview", self.showOverviewSegsCheck
        )
        self.showOverviewSegsTick.setCheckable(True)
        self.showOverviewSegsTick.setChecked(self.config["showAnnotationOverview"])

        self.showPointerDetails = viewMenu.addAction(
            "Show pointer details in spectrogram", self.showPointerDetailsCheck
        )
        self.showPointerDetails.setCheckable(True)
        self.showPointerDetails.setChecked(self.config["showPointerDetails"])

        viewMenu.addSeparator()

        colMenu = viewMenu.addMenu("Choose colour map")
        colGroup = QActionGroup(self)
        for colour in self.config["ColourList"]:
            cm = colMenu.addAction(colour)
            cm.setCheckable(True)
            if colour == self.config["cmap"]:
                cm.setChecked(True)
            receiver = lambda checked, cmap=colour: self.setColourMap(cmap)
            cm.triggered.connect(receiver)
            colGroup.addAction(cm)
        self.invertcm = viewMenu.addAction("Invert colour map", self.invertColourMap)
        self.invertcm.setCheckable(True)
        self.invertcm.setChecked(self.config["invertColourMap"])

        # viewMenu.addSeparator()
        viewMenu.addAction(
            "&Change spectrogram parameters", self.showSpectrogramDialog, "Ctrl+C"
        )

        viewMenu.addSeparator()
        markMenu = viewMenu.addMenu("Mark on spectrogram")
        self.showFundamental = markMenu.addAction(
            "Fundamental frequency", self.showFundamentalFreq, "Ctrl+F"
        )
        self.showFundamental.setCheckable(True)
        self.showFundamental.setChecked(True)
        # self.showSpectral = markMenu.addAction(
        #     "Spectral derivative", self.showSpectralDeriv
        # )
        # self.showSpectral.setCheckable(True)
        # self.showSpectral.setChecked(False)
        # if not self.DOC:
        #     self.showFormant = markMenu.addAction("Formants", self.showFormants)
        #     self.showFormant.setCheckable(True)
        #     self.showFormant.setChecked(False)
        self.showEnergies = markMenu.addAction("Maximum energies", self.showMaxEnergy)
        self.showEnergies.setCheckable(True)
        self.showEnergies.setChecked(False)

        # if not self.DOC:
        #     cqt = viewMenu.addAction("Show CQT", self.showCQT)

        viewMenu.addSeparator()

        self.readonly = viewMenu.addAction("Make read only", self.makeReadOnly)
        self.readonly.setCheckable(True)
        self.readonly.setChecked(self.config["readOnly"])

        self.sortRank = viewMenu.addAction(
            "Sort files by maximum confidence", self.toggleRankSort
        )
        self.sortRank.setCheckable(True)

        exportViewMenu = viewMenu.addMenu("Export view")
        exportViewMenu.addAction("Overview spectrogram as image", self.saveImageRaw)
        exportViewMenu.addAction(
            "Current spectrogram as image", self.saveImage, "Ctrl+I"
        )

        viewMenu.addSeparator()
        viewMenu.addAction("Restore default layout", self.dockReplace)
        viewMenu.addAction("Interface settings", self.changeSettings)

        analysisMenu = self.menuBar().addMenu("&Analysis")
        analysisMenu.addAction("Classify recordings with BirdNET", self.classifyBirdNET)
        analysisMenu.addSeparator()
        self.denoiseAction = analysisMenu.addAction("Denoise", self.showDenoiseDialog)
        analysisMenu.addAction("Add metadata about noise", self.addNoiseData, "Ctrl+N")

        # analysisMenu.addAction("Filter spectrogram", self.medianFilterSpec)

        # analysisMenu.addSeparator()
        #
        # analysisMenu.addAction("Calculate segment statistics", self.calculateStats)

        annotationMenu = self.menuBar().addMenu("&Annotations")

        importMenu = annotationMenu.addMenu("&Import Annotations")
        importMenu.addAction("From AviaNZ (.data files)", self.importAvianzData)
        importMenu.addAction("From Raven (selection tables)", self.importRavenData)
        # importMenu.addAction("From Excel", self.excel2Annotation)

        exportMenu = annotationMenu.addMenu("&Export Annotations")
        exportMenu.addAction("To Excel (File)", self.exportSeg)
        exportMenu.addAction("To Excel (Directory)", self.exportExcel)
        exportMenu.addAction("To .data files (Directory)", self.exportAvianzData)
        exportMenu.addAction("To database", self.exportDatabase)
        exportFilesMenu = annotationMenu.addMenu("&Export Audio")
        exportFilesMenu.addAction("Files with selected species", self.exportFiles)
        exportFilesMenu.addAction(
            "File segments with selected species", self.exportAudioSegments
        )
        annotationMenu.addSeparator()

        deleteMenu = annotationMenu.addMenu("&Delete annotations")
        deleteMenu.addAction("File annotations", self.deleteAll, "Ctrl+D")
        deleteMenu.addAction("Directory annotations", self.deleteDirAnnotations)

        self.addRegularAction = annotationMenu.addAction(
            "Mark regular segments", self.addRegularSegments, "Ctrl+M"
        )

        # annotationMenu.addAction("Backup annotations", self.backupAnnotations)

        fileMenu.addSeparator()

        databaseMenu = annotationMenu.addMenu("Database")
        databaseMenu.addAction("Choose database", self.chooseDatabase)
        databaseMenu.addAction(
            "Update file directory",
            self.updateDirectory,
        )

        helpMenu = self.menuBar().addMenu("&Help")
        helpMenu.addAction("&Manual", self.showHelp, "Ctrl+H")
        helpMenu.addAction("&Cheat Sheet", self.showCheatSheet)
        helpMenu.addSeparator()
        helpMenu.addAction("&About", self.showAbout, "Ctrl+A")
        if platform.system() == "Darwin":
            helpMenu.addAction("About", self.showAbout, "Ctrl+A")

    def showAbout(self):
        """Create the About Message Box"""
        msg = SupportClasses_GUI.MessagePopup("a", "About", ".")
        msg.exec_()
        return

    def showHelp(self):
        """Show the user manual (a pdf file), make it offline for easy access"""
        webbrowser.open_new(r"file://" + os.path.realpath("./Docs/AviaNZManual.pdf"))
        # webbrowser.open_new(r'http://avianz.net/docs/AviaNZManual.pdf')

    def showCheatSheet(self):
        """Show the cheatsheet of sample spectrograms"""
        # webbrowser.open_new(r'http://www.avianz.net/index.php/resources/cheat-sheet/about-cheat-sheet')
        webbrowser.open_new(
            r"file://" + os.path.realpath("./Docs/AviaNZCheatSheet.pdf")
        )

    def createFrame(self):
        """Creates the main window.
        This consists of a set of pyqtgraph docks with widgets in.
         d_ for docks, w_ for widgets, p_ for plots"""

        # Make the window and set its size
        self.area = DockArea()
        self.setCentralWidget(self.area)
        self.resize(1240, 600)
        self.move(100, 50)

        # Make the colours that are used in the interface
        # The dark ones are to draw lines instead of boxes
        self.ColourSelected = QtGui.QColor(
            self.config["ColourSelected"][0],
            self.config["ColourSelected"][1],
            self.config["ColourSelected"][2],
            self.config["ColourSelected"][3],
        )
        self.ColourNamed = QtGui.QColor(
            self.config["ColourNamed"][0],
            self.config["ColourNamed"][1],
            self.config["ColourNamed"][2],
            self.config["ColourNamed"][3],
        )
        self.ColourNone = QtGui.QColor(
            self.config["ColourNone"][0],
            self.config["ColourNone"][1],
            self.config["ColourNone"][2],
            self.config["ColourNone"][3],
        )
        self.ColourPossible = QtGui.QColor(
            self.config["ColourPossible"][0],
            self.config["ColourPossible"][1],
            self.config["ColourPossible"][2],
            self.config["ColourPossible"][3],
        )

        self.ColourSelectedDark = QtGui.QColor(
            self.config["ColourSelected"][0],
            self.config["ColourSelected"][1],
            self.config["ColourSelected"][2],
            255,
        )
        self.ColourNamedDark = QtGui.QColor(
            self.config["ColourNamed"][0],
            self.config["ColourNamed"][1],
            self.config["ColourNamed"][2],
            255,
        )
        self.ColourNoneDark = QtGui.QColor(
            self.config["ColourNone"][0],
            self.config["ColourNone"][1],
            self.config["ColourNone"][2],
            255,
        )
        self.ColourPossibleDark = QtGui.QColor(
            self.config["ColourPossible"][0],
            self.config["ColourPossible"][1],
            self.config["ColourPossible"][2],
            255,
        )

        # Make the docks and lay them out: d_ -> dock
        self.d_overview = Dock("Overview", size=(1200, 150))
        self.d_ampl = Dock("Amplitude", size=(1200, 150))
        self.d_spec = Dock("Spectrogram", size=(1200, 300))
        self.d_controls = Dock("Controls", size=(40, 90))
        self.d_files = Dock("Files", size=(40, 200))
        self.d_controls.setSizePolicy(1, 1)

        self.area.addDock(self.d_files, "left")
        self.area.addDock(self.d_overview, "right", self.d_files)
        self.area.addDock(self.d_ampl, "bottom", self.d_overview)
        self.area.addDock(self.d_spec, "bottom", self.d_ampl)
        self.area.addDock(self.d_controls, "bottom", self.d_files)

        # Store the state of the docks in case the user wants to reset it
        self.state = self.area.saveState()
        containers, docks = self.area.findAll()
        self.state_cont = [cont.sizes() for cont in containers]

        # Put content widgets in the docks: w_ -> widget
        # OVERVIEW dock
        self.w_overview = pg.LayoutWidget()
        self.w_overview.layout.setColumnStretch(1, 10)
        self.w_overview.layout.setColumnStretch(0, 0)
        self.w_overview.layout.setColumnStretch(2, 0)
        self.d_overview.addWidget(self.w_overview)
        # this will hold both overview image and segment boxes
        self.w_overview1 = pg.GraphicsLayoutWidget()
        self.w_overview1.ci.layout.setContentsMargins(0.5, 1, 0.5, 1)
        self.w_overview1.ci.layout.setRowSpacing(0, 0)
        # self.w_overview1.ci.layout.setRowSpacing(1, 0)
        self.w_overview1.ci.layout.setRowStretchFactor(0, 7)
        self.w_overview1.ci.layout.setRowStretchFactor(1, 1)
        # self.w_overview1.ci.layout.setRowMaximumHeight(1, 40)

        fileInfo = QHBoxLayout()
        self.fileInfoSR = QLabel()
        self.fileInfoSR.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoNCh = QLabel()
        self.fileInfoNCh.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoSS = QLabel()
        self.fileInfoSS.setStyleSheet("QLabel {color: #505050}")
        self.fileInfoDur = QLabel()
        self.fileInfoDur.setStyleSheet("QLabel {color: #505050}")
        fileInfo.addWidget(self.fileInfoSR)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoNCh)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoSS)
        fileInfo.addSpacing(20)
        fileInfo.addWidget(self.fileInfoDur)
        fileInfo.addStretch(5)

        # p_overview holds the overview plot image
        self.p_overview = SupportClasses_GUI.DemousedViewBox()
        self.w_overview1.addItem(self.p_overview, row=0, col=0)

        # p_overview2 holds the sexgment boxes in the overview dock
        self.p_overview2 = SupportClasses_GUI.ChildInfoViewBox(
            enableMouse=False, enableMenu=False
        )
        self.w_overview1.addItem(self.p_overview2, row=1, col=0)
        self.p_overview2.setXLink(self.p_overview)
        self.p_overview2.setPreferredHeight(25)

        self.p_overview2.setCursor(Qt.PointingHandCursor)

        # The buttons to move through the overview
        self.leftBtn = QPushButton()
        self.leftBtn.setIcon(QIcon("img/overview-back.png"))
        self.leftBtn.setIconSize(QtCore.QSize(7, 28))
        self.leftBtn.setMinimumWidth(16)
        self.leftBtn.clicked.connect(self.moveLeft)
        self.leftBtn.setToolTip("Move view back")
        self.rightBtn = QPushButton()
        self.rightBtn.setIcon(QIcon("img/overview-next.png"))
        self.rightBtn.setIconSize(QtCore.QSize(7, 28))
        self.rightBtn.setMinimumWidth(16)
        self.rightBtn.clicked.connect(self.moveRight)
        self.rightBtn.setToolTip("Move view forward")
        self.leftBtn.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.MinimumExpanding)
        self.rightBtn.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.MinimumExpanding)

        # Buttons to move to next/previous five minutes
        self.prev5mins = QToolButton()
        self.prev5mins.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaSeekBackward)
        )
        self.prev5mins.setMinimumSize(35, 30)
        self.prev5mins.setToolTip("Previous page")
        self.prev5mins.clicked.connect(self.movePrev5mins)
        self.next5mins = QToolButton()
        self.next5mins.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaSeekForward)
        )
        self.next5mins.setMinimumSize(35, 30)
        self.next5mins.setToolTip("Next page")
        self.next5mins.clicked.connect(self.moveNext5mins)
        self.placeInFileLabel2 = QLabel("Page")
        self.placeInFileLabel = QLabel("")
        self.placeInFileLabel.setSizePolicy(QSizePolicy.Minimum, QSizePolicy.Minimum)
        self.placeInFileSelector = QSpinBox()
        self.placeInFileSelector.setRange(1, 1)
        self.placeInFileSelector.setButtonSymbols(QAbstractSpinBox.NoButtons)
        self.placeInFileSelector.editingFinished.connect(self.moveTo5mins)
        self.placeInFileSelector.setMinimumHeight(25)

        # Find previous annotation buttons
        self.annotPrevLabel = QLabel("Jump to previous mark:")
        self.annotAnyPrevBtn = QToolButton()
        self.annotAnyPrevBtn.setIcon(QIcon("img/findprev-g.png"))
        self.annotAnyPrevBtn.setToolTip("Any label [Ctrl+Left]")
        # self.annotAnyPrevBtn.setAutoRaise(True)
        self.annotAnyPrevBtn.setMinimumSize(35, 30)
        self.annotAnyPrevBtn.setIconSize(QtCore.QSize(20, 17))
        self.annotAnyPrevBtn.clicked.connect(lambda: self.annotJumperPrev(100))

        self.annotAnyPrevKey = QShortcut(QKeySequence("Ctrl+Left"), self)
        self.annotAnyPrevKey.activated.connect(lambda: self.annotJumperPrev(100))

        self.annotUncertPrevBtn = QToolButton()
        self.annotUncertPrevBtn.setIcon(QIcon("img/findprev-y.png"))
        self.annotUncertPrevBtn.setToolTip("Uncertain label")
        # self.annotUncertPrevBtn.setAutoRaise(True)
        self.annotUncertPrevBtn.setMinimumSize(35, 30)
        self.annotUncertPrevBtn.setIconSize(QtCore.QSize(20, 17))
        self.annotUncertPrevBtn.clicked.connect(lambda: self.annotJumperPrev(99))

        # "Find next annotation" buttons
        self.annotNextLabel = QLabel("Jump to next mark:")
        self.annotAnyNextBtn = QToolButton()
        self.annotAnyNextBtn.setIcon(QIcon("img/findnext-g.png"))
        self.annotAnyNextBtn.setToolTip("Any label [Ctrl+Right]")
        # self.annotAnyNextBtn.setAutoRaise(True)
        self.annotAnyNextBtn.setMinimumSize(35, 30)
        self.annotAnyNextBtn.setIconSize(QtCore.QSize(20, 17))
        self.annotAnyNextBtn.clicked.connect(lambda: self.annotJumper(100))
        self.annotAnyNextKey = QShortcut(QKeySequence("Ctrl+Right"), self)
        self.annotAnyNextKey.activated.connect(lambda: self.annotJumper(100))

        self.annotUncertNextBtn = QToolButton()
        self.annotUncertNextBtn.setIcon(QIcon("img/findnext-y.png"))
        self.annotUncertNextBtn.setToolTip("Uncertain label")
        # self.annotUncertNextBtn.setAutoRaise(True)
        self.annotUncertNextBtn.setMinimumSize(35, 30)
        self.annotUncertNextBtn.setIconSize(QtCore.QSize(20, 17))
        self.annotUncertNextBtn.clicked.connect(lambda: self.annotJumper(99))

        # position everything in the dock
        self.w_overview.layout.addLayout(fileInfo, 0, 0, 1, 3)
        # self.w_overview.addWidget(annotInfo, row=1, col=0, colspan=2)
        self.w_overview.addWidget(self.w_overview1, row=2, col=1)
        self.w_overview.addWidget(self.leftBtn, row=2, col=0)
        self.w_overview.addWidget(self.rightBtn, row=2, col=2)
        placeInFileBox = QHBoxLayout()
        placeInFileBox.addStretch(4)
        placeInFileBox.addWidget(self.annotPrevLabel)
        placeInFileBox.addWidget(self.annotAnyPrevBtn)
        placeInFileBox.addWidget(self.annotUncertPrevBtn)
        placeInFileBox.addStretch(4)
        placeInFileBox.addWidget(self.placeInFileLabel2)
        placeInFileBox.addWidget(self.prev5mins)
        placeInFileBox.addWidget(self.placeInFileSelector)
        placeInFileBox.addWidget(self.next5mins)
        placeInFileBox.addWidget(self.placeInFileLabel)
        placeInFileBox.addStretch(4)
        placeInFileBox.addWidget(self.annotNextLabel)
        placeInFileBox.addWidget(self.annotAnyNextBtn)
        placeInFileBox.addWidget(self.annotUncertNextBtn)
        placeInFileBox.addStretch(4)
        self.w_overview.layout.addLayout(placeInFileBox, 3, 1)

        # Corresponding keyboard shortcuts:
        self.moveLeftKey = QShortcut(QKeySequence(Qt.Key_Left), self)
        self.moveLeftKey2 = QShortcut(QKeySequence(Qt.Key_A), self)
        self.moveLeftKey.activated.connect(self.moveLeft)
        self.moveLeftKey2.activated.connect(self.moveLeft)
        self.moveRightKey = QShortcut(QKeySequence(Qt.Key_Right), self)
        self.moveRightKey2 = QShortcut(QKeySequence(Qt.Key_D), self)
        self.moveRightKey.activated.connect(self.moveRight)
        self.moveRightKey2.activated.connect(self.moveRight)
        self.movePrev5MinsKey = QShortcut(QKeySequence("Shift+Left"), self)
        self.movePrev5MinsKey.activated.connect(self.movePrev5mins)
        self.moveNext5MinsKey = QShortcut(QKeySequence("Shift+Right"), self)
        self.moveNext5MinsKey.activated.connect(self.moveNext5mins)

        # AMPLITUDE dock
        self.w_ampl = pg.GraphicsLayoutWidget()
        self.p_ampl = SupportClasses_GUI.DragViewBox(
            self, enableMouse=False, enableMenu=False, enableDrag=False, thisIsAmpl=True
        )
        self.p_ampl.setAutoVisible(False, True)
        self.w_ampl.addItem(self.p_ampl, row=0, col=1)
        self.d_ampl.addWidget(self.w_ampl)

        self.w_spec = pg.GraphicsLayoutWidget()
        self.p_spec = SupportClasses_GUI.DragViewBox(
            self,
            enableMouse=False,
            enableMenu=False,
            enableDrag=self.config["specMouseAction"] == 3,
            thisIsAmpl=False,
        )
        self.w_spec.addItem(self.p_spec, row=0, col=1)
        self.d_spec.addWidget(self.w_spec)

        # The axes
        # Time axis has to go separately in loadFile
        self.ampaxis = pg.AxisItem(orientation="left")
        self.w_ampl.addItem(self.ampaxis, row=0, col=0)
        self.ampaxis.linkToView(self.p_ampl)
        self.ampaxis.setWidth(w=65)
        self.ampaxis.setLabel("")

        self.specaxis = pg.AxisItem(orientation="left")
        self.w_spec.addItem(self.specaxis, row=0, col=0)
        self.specaxis.linkToView(self.p_spec)
        self.specaxis.setWidth(w=65)

        # The slider to show playback position
        # This is hidden, but controls the moving bar
        self.bar = pg.InfiniteLine(
            angle=90, movable=True, pen={"color": "c", "width": 3}
        )
        self.bar.btn = self.MouseDrawingButton
        self.bar.sigPositionChangeFinished.connect(self.barMoved)

        self.guidelines = [0] * len(self.config["guidecol"])
        for gi in range(len(self.config["guidecol"])):
            self.guidelines[gi] = pg.InfiniteLine(
                angle=0,
                movable=False,
                pen={"color": self.config["guidecol"][gi], "width": 2},
            )

        # The print out at the bottom of the spectrogram with data in
        # Note: widgets cannot be directly added to GraphicsLayout, so need to convert
        # them to proxy GraphicsWidgets using the proxy
        self.pointData = QLabel()
        self.pointData.setStyleSheet(
            "QLabel { background-color : white; color : #CC0000; }"
        )
        self.pointDataProxy = QGraphicsProxyWidget()
        self.pointDataProxy.setWidget(self.pointData)
        self.segInfo = QLabel()
        self.segInfo.setStyleSheet(
            "QLabel { background-color : white; color : #CC0000; }"
        )
        self.segInfoProxy = QGraphicsProxyWidget()
        self.segInfoProxy.setWidget(self.segInfo)

        # The various plots
        self.overviewImage = pg.ImageItem(enableMouse=False)
        self.p_overview.addItem(self.overviewImage)
        self.overviewImageRegion = SupportClasses_GUI.LinearRegionItemO(
            pen=pg.mkPen(120, 80, 200, width=2),
            hoverPen=pg.mkPen(60, 40, 230, width=3.5),
        )
        # this is needed for compatibility with other shaded rectangles:
        self.overviewImageRegion.lines[0].btn = Qt.RightButton
        self.overviewImageRegion.lines[1].btn = Qt.RightButton
        self.p_overview.addItem(self.overviewImageRegion, ignoreBounds=True)
        self.amplPlot = pg.PlotDataItem()
        self.p_ampl.addItem(self.amplPlot)
        self.specPlot = pg.ImageItem()
        self.p_spec.addItem(self.specPlot)
        if self.MouseDrawingButton == Qt.RightButton:
            self.p_ampl.unsetCursor()
            self.specPlot.unsetCursor()
            self.bar.setCursor(Qt.OpenHandCursor)
        else:
            self.p_ampl.setCursor(QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0))
            self.specPlot.setCursor(QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0))
            self.bar.unsetCursor()

        # Connect up the listeners
        self.p_ampl.scene().sigMouseClicked.connect(self.mouseClicked_ampl)
        self.p_spec.scene().sigMouseClicked.connect(self.mouseClicked_spec)

        # Connect up so can disconnect if not selected...
        self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
        self.w_spec.addItem(self.segInfoProxy, row=2, col=1)
        self.w_spec.addItem(self.pointDataProxy, row=3, col=1)

        # The content of the other two docks
        self.w_controls = pg.LayoutWidget()
        self.d_controls.addWidget(self.w_controls)
        self.w_files = pg.LayoutWidget()
        self.d_files.addWidget(self.w_files)

        self.skipBackwardKey = QShortcut(QKeySequence("Up"), self)
        self.skipBackwardKey2 = QShortcut(QKeySequence(Qt.Key_W), self)
        self.skipBackwardKey.activated.connect(
            lambda: self.openPreviousFile(skipHidden=True)
        )
        self.skipBackwardKey2.activated.connect(
            lambda: self.openPreviousFile(skipHidden=True)
        )
        self.skipBackwardHiddenKey = QShortcut(QKeySequence("Alt+Up"), self)
        self.skipBackwardHiddenKey2 = QShortcut(
            QKeySequence(Qt.Key_W + Qt.Key_Up), self
        )
        self.skipBackwardHiddenKey.activated.connect(
            lambda: self.openPreviousFile(skipHidden=False)
        )

        self.skipForwardKey = QShortcut(QKeySequence("Down"), self)
        self.skipForwardKey2 = QShortcut(QKeySequence(Qt.Key_S), self)
        self.skipForwardKey.activated.connect(
            lambda: self.openNextFile(skipHidden=True)
        )
        self.skipForwardKey2.activated.connect(
            lambda: self.openNextFile(skipHidden=True)
        )
        self.skipForwardHiddenKey = QShortcut(QKeySequence("Alt+Down"), self)
        self.skipForwardHiddenKey2 = QShortcut(
            QKeySequence(Qt.Key_S + Qt.Key_Down), self
        )
        self.skipForwardHiddenKey.activated.connect(
            lambda: self.openNextFile(skipHidden=False)
        )
        self.skipForwardHiddenKey2.activated.connect(
            lambda: self.openNextFile(skipHidden=False)
        )

        # The buttons inside the controls dock
        self.playButton = QtWidgets.QToolButton()
        self.playButton.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaPlay)
        )
        self.playButton.setIconSize(QtCore.QSize(20, 20))
        self.playButton.setToolTip("Play visible [Space]")
        self.playButton.clicked.connect(self.playVisible)
        self.playKey = QShortcut(QKeySequence("Space"), self)
        self.playKey.activated.connect(self.playVisible)

        self.stopButton = QtWidgets.QToolButton()
        self.stopButton.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
        )
        self.stopButton.setIconSize(QtCore.QSize(20, 20))
        self.stopButton.setToolTip("Stop playback")
        self.stopButton.clicked.connect(self.stopPlayback)

        self.playSegButton = QtWidgets.QToolButton()
        self.playSegButton.setIcon(QIcon("img/playsegment.png"))
        self.playSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playSegButton.setToolTip("Play selected")
        self.playSegButton.clicked.connect(self.playSelectedSegment)
        self.playSegKey = QShortcut(QKeySequence("Ctrl+Space"), self)
        self.playSegKey.activated.connect(self.playSelectedSegment)

        self.playSlowButton = QtWidgets.QToolButton()
        self.playSlowButton.setIcon(QIcon("img/playSlow-w.png"))
        self.playSlowButton.setIconSize(QtCore.QSize(35, 20))
        self.playSlowButton.setToolTip("Play slowly")
        self.playSlowButton.clicked.connect(self.playSlowSegment)

        speedMenu = QMenu()
        extraGroup = QActionGroup(self)
        for ename in ["2", "\u00bd", "\u00bc"]:
            em = speedMenu.addAction(ename)
            em.setCheckable(True)
            if ename == "0.5":
                em.setChecked(True)
            receiver = lambda checked, ename=ename: self.setSpeed(ename)
            em.triggered.connect(receiver)
            extraGroup.addAction(em)
        self.playSlowButton.setMenu(speedMenu)

        self.quickDenButton = QtWidgets.QToolButton()
        self.quickDenButton.setIcon(QIcon("img/denoisesegment.png"))
        self.quickDenButton.setIconSize(QtCore.QSize(20, 20))
        self.quickDenButton.setToolTip("Denoise segment")
        self.quickDenButton.clicked.connect(self.denoiseSeg)

        self.toggleLabelTypeBtn = QtWidgets.QToolButton()
        self.toggleLabelTypeBtn.setIcon(QIcon("img/splarge-ct.png"))
        self.toggleLabelTypeBtn.setIconSize(QtCore.QSize(35, 20))
        self.toggleLabelTypeBtn.setToolTip(
            "Toggle between species/calltype views [Tab]"
        )
        self.toggleLabelTypeBtn.clicked.connect(self.toggleLabelType)

        self.toggleLabelTypeKey = QShortcut(QKeySequence("Tab"), self)
        self.toggleLabelTypeKey.activated.connect(self.toggleLabelType)

        self.playBandLimitedSegButton = QtWidgets.QToolButton()
        self.playBandLimitedSegButton.setIcon(QtGui.QIcon("img/playBandLimited.png"))
        self.playBandLimitedSegButton.setIconSize(QtCore.QSize(20, 20))
        self.playBandLimitedSegButton.setToolTip("Play selected-band limited")
        self.playBandLimitedSegButton.clicked.connect(self.playBandLimitedSegment)

        # Volume, brightness and contrast sliders.
        # Need to pass true (config) values to set up correct initial positions
        self.specControls = SupportClasses_GUI.BrightContrVol(
            self.config["brightness"],
            self.config["contrast"],
            self.config["invertColourMap"],
            horizontal=False,
        )
        self.specControls.colChanged.connect(self.setColourLevels)
        self.specControls.volChanged.connect(self.volSliderMoved)

        # Confirm button - auto ups the certainty to 100
        self.confirmButton = QPushButton("   Confirm labels")
        self.confirmButton.clicked.connect(self.confirmSegment)
        self.confirmButton.setIcon(QIcon(QPixmap("img/check-mark2.png")))
        self.confirmButton.setStyleSheet("QPushButton {padding: 3px 3px 3px 3px}")
        self.confirmButton.setToolTip(
            "Set all labels in this segment as certain [Return]"
        )

        self.confirmKey = QShortcut(QKeySequence("Return"), self)
        self.confirmKey.activated.connect(self.confirmSegment)

        # Delete segment button. We have to get rid of the extra event args
        self.deleteButton = QPushButton("  Delete segment")
        self.deleteButton.clicked.connect(lambda _: self.deleteSegment())
        self.deleteButton.setIcon(QIcon(QPixmap("img/deleteL.png")))
        self.deleteButton.setStyleSheet("QPushButton {padding: 3px 3px 3px 3px}")

        # export selected sound
        self.exportSoundBtn = QPushButton("  Save sound clip")
        self.exportSoundBtn.clicked.connect(lambda _: self.saveSelectedSound())
        self.exportSoundBtn.setIcon(QIcon(QPixmap("img/storage2.png")))
        self.exportSoundBtn.setToolTip("Export the selected segment to a file")

        # export selected sound
        if not self.DOC:
            self.exportSlowSoundBtn = QPushButton("  Save slow sound clip")
            self.exportSlowSoundBtn.clicked.connect(
                lambda _: self.saveSelectedSound(self.slowSpeed)
            )
            self.exportSlowSoundBtn.setIcon(QIcon(QPixmap("img/storage2.png")))
            self.exportSlowSoundBtn.setToolTip(
                "Export the selected sound to a file at different speed"
            )

        # flips buttons to Disabled state
        self.refreshSegmentControls()

        # The spinbox for changing the width shown in the controls dock
        windowLabel = QLabel("Visible window (seconds)")
        windowLabel.setAlignment(Qt.AlignBottom)
        self.widthWindow = QDoubleSpinBox()
        self.widthWindow.setSingleStep(1.0)
        self.widthWindow.setDecimals(2)
        self.widthWindow.setValue(self.config["windowWidth"])
        self.widthWindow.valueChanged[float].connect(self.changeWidth)

        # Place all these widgets in the Controls dock
        self.w_controls.addWidget(self.playButton, row=0, col=0)
        self.w_controls.addWidget(self.playSegButton, row=0, col=1)
        self.w_controls.addWidget(self.playBandLimitedSegButton, row=0, col=2)
        self.w_controls.addWidget(self.playSlowButton, row=0, col=3)
        self.w_controls.addWidget(self.stopButton, row=1, col=0)
        # self.w_controls.addWidget(self.speedButton,row=1,col=1)
        if not self.DOC:
            self.w_controls.addWidget(self.quickDenButton, row=1, col=2)
            # self.w_controls.addWidget(self.quickDenNButton,row=1,col=1)

        self.w_controls.addWidget(self.toggleLabelTypeBtn, row=1, col=3)

        self.w_controls.addWidget(self.specControls, row=2, col=0, rowspan=2, colspan=4)

        self.w_controls.addWidget(QLabel("Visible window"), row=8, col=0, colspan=4)
        self.w_controls.addWidget(self.widthWindow, row=9, col=0, colspan=2)
        self.w_controls.addWidget(QLabel("seconds"), row=9, col=2, colspan=2)
        # spacer b/c pyqtgraph can't add spacer items
        spacer = QWidget()
        self.w_controls.addWidget(spacer, row=10, col=0, colspan=4)
        self.w_controls.layout.setRowMinimumHeight(10, 3)

        # empty widget to add in the gridlayout
        segContrs = QGroupBox("Selected segment")
        segContrs.setStyleSheet("QGroupBox:title{color: #505050; font-weight: 50}")
        segContrsBox = QVBoxLayout()
        segContrs.setLayout(segContrsBox)
        segContrsBox.addWidget(self.confirmButton)
        segContrsBox.addWidget(self.deleteButton)
        segContrsBox.addWidget(self.exportSoundBtn)
        if not self.DOC:
            segContrsBox.addWidget(self.exportSlowSoundBtn)
        self.w_controls.addWidget(segContrs, row=12, col=0, colspan=4)

        # A slider to move through the file easily
        self.scrollSlider = QScrollBar(Qt.Horizontal)
        self.scrollSlider.valueChanged.connect(self.scroll)
        self.d_spec.addWidget(self.scrollSlider)

        # List to hold the list of files
        self.listFiles = SupportClasses_GUI.LightedFileList(
            self.ColourNone, self.ColourPossibleDark, self.ColourNamed, self.db
        )
        self.listFiles.itemDoubleClicked.connect(self.listLoadFile)
        self.listSpecies = QComboBox()
        self.listSpecies.setToolTip(
            "Select species and tick the Box to reduce filelist."
        )
        self.listSpecies.currentIndexChanged.connect(self.updateListSpeciesIndexChanged)
        self.currentSpecies = "Species"
        # self.tickSpecies = QCheckBox("Only files with selected species?")
        self.tickSpecies = QCheckBox()
        self.tickSpecies.setToolTip("Tick to reduce filelist to current species.")
        self.tickSpecies.stateChanged.connect(self.updateTickSpeciesStateChanged)
        self.tickSpecies.setSizePolicy(QSizePolicy.Maximum, QSizePolicy.Fixed)

        self.confRangeSlider = superqt.QLabeledRangeSlider(Qt.Horizontal)
        self.confRangeSlider.setRange(0, 100)
        self.confRangeSlider.setHandleLabelPosition(0)
        self.confRangeSlider.setEdgeLabelMode(2)
        self.confRangeSlider.setValue((0, 100))
        self.confRangeSlider.setTracking(False)
        self.confRangeSlider.setToolTip(
            "Set confidence range to filter detections in files"
        )
        self.confRangeSlider.valuesChanged.connect(
            self.updateConfRangeSliderValueChanged
        )
        self.confRangeSlider.slidersMoved.connect(self.updateConfRangeSliderMoved)
        self.confRangeSlider.sliderReleased.connect(self.updateConfRangeSliderReleased)

        self.confidenceRange = (0, 100)

        self.timeRangeStartLabel = QLabel("00:00")
        self.timeRangeStartLabel.setToolTip(
            "Set time range to filter files. Does not filter directories."
        )

        self.timeRangeSlider = superqt.QRangeSlider(Qt.Horizontal)
        self.timeRangeSlider.setRange(0, 24 * 6)
        self.timeRangeSlider.setValue((0, 24 * 6))
        self.timeRangeSlider.setTickInterval(1)
        self.timeRangeSlider.setPageStep(6)
        self.timeRangeSlider.setToolTip(
            "Set time range to filter files. Does not filter directories."
        )
        self.timeRangeSlider.valuesChanged.connect(
            self.updateTimeRangeSliderValuesChanged
        )
        self.timeRangeSlider.slidersMoved.connect(self.updateTimeRangeSlidersMoved)
        self.timeRangeSlider.sliderReleased.connect(self.updateTimeRangeSliderReleased)

        self.timeRangeEndLabel = QLabel("24:00")
        self.timeRangeEndLabel.setToolTip(
            "Set time range to filter files. Does not filter directories."
        )

        self.timeRange = (0, 24 * 3600)
        self.w_files.addWidget(self.listSpecies, row=2, col=0, colspan=2)
        self.w_files.addWidget(self.tickSpecies, row=2, col=2)
        self.w_files.addWidget(self.confRangeSlider, row=4, col=0, colspan=3)
        self.w_files.addWidget(self.timeRangeStartLabel, row=5, col=0)
        self.w_files.addWidget(self.timeRangeSlider, row=5, col=1)
        self.w_files.addWidget(self.timeRangeEndLabel, row=5, col=2)
        self.w_files.addWidget(self.listFiles, row=8, colspan=3)

        # The context menu (drops down on mouse click) to select birds
        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.menuBirdList = QMenu()
        self.menuBird2 = QMenu("Other")
        # New line to allow multiple selections
        self.menuBirdList.installEventFilter(self)
        self.menuBird2.installEventFilter(self)
        self.fillBirdList()
        self.menuBirdList.triggered.connect(self.birdSelectedMenu)
        self.menuBirdList.aboutToHide.connect(self.refreshFileColor)
        self.menuBird2.triggered.connect(self.birdSelectedMenu)

        # Hack to get the type of an ROI
        p_spec_r = SupportClasses_GUI.ShadedRectROI(0, 0)
        self.ROItype = type(p_spec_r)

        # Listener for key presses
        self.w_ampl.installEventFilter(self)
        self.w_spec.installEventFilter(self)

        # add statusbar
        self.statusLeft = QLabel("Left")
        # Not sure what's the difference between Sunken and Panel?
        self.statusLeft.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.statusDB = QLabel("")
        self.statusDB.setAlignment(Qt.AlignCenter)
        self.statusDB.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.statusRO = QLabel("")
        self.statusRO.setAlignment(Qt.AlignCenter)
        self.statusRO.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.statusRight = QLabel("")
        self.statusRight.setAlignment(Qt.AlignRight)
        self.statusRight.setFrameStyle(QFrame.Panel | QFrame.Sunken)
        self.statusBar().addPermanentWidget(self.statusLeft, 3)
        self.statusBar().addPermanentWidget(self.statusDB, 1)
        self.statusBar().addPermanentWidget(self.statusRO, 1)
        self.statusBar().addPermanentWidget(self.statusRight, 2)

        # Set the message in the status bar
        self.statusLeft.setText("Ready")
        self.statusDB.setText(os.path.basename(self.dbPath))
        self.statusRO.setText("Read-only mode" if self.config["readOnly"] else "")

        # Function calls to check if should show various parts of the interface, whether dragging boxes or not
        self.makeReadOnly()
        self.useAmplitudeCheck()
        self.useFilesCheck()
        self.showOverviewSegsCheck()
        self.dragRectsTransparent()
        self.showPointerDetailsCheck()
        self.w_spec.setFocus()

    def refreshSegmentControls(self):
        """Just toggles all the segment controls on/off when a segment
        is (de)selected. Call this after changing self.box1id.
        Remember to update this when segment controls change!
        """
        # basic buttons which toggle on any segment selection
        if self.DOC:
            btns = [
                self.deleteButton,
                self.playSegButton,
                self.playSlowButton,
                self.quickDenButton,
                self.exportSoundBtn,
            ]
        else:
            btns = [
                self.deleteButton,
                self.playSegButton,
                self.playSlowButton,
                self.quickDenButton,
                self.exportSoundBtn,
                self.exportSlowSoundBtn,
            ]

        # if self.box1id is not -1, flip on, otherwise off
        if self.box1id < 0:
            for btn in btns:
                btn.setEnabled(False)
            self.playBandLimitedSegButton.setEnabled(False)
            self.confirmButton.setEnabled(False)
        else:
            for btn in btns:
                btn.setEnabled(True)

            # special case for BandLimitedButton b/c it requires set freq bands
            if type(self.listRectanglesa2[self.box1id]) is self.ROItype:
                # it's a rectangle box:
                self.playBandLimitedSegButton.setEnabled(True)
            else:
                # it's a 0 to inf segment:
                self.playBandLimitedSegButton.setEnabled(False)

            # special case for Confirm button b/c it requires yellow segment
            self.confirmButton.setEnabled(False)
            for sp in self.segments[self.box1id][4]:
                if sp["certainty"] < 100 and sp["species"] != "Don't Know":
                    self.confirmButton.setEnabled(True)
                    break

    def makeFullBirdList(self, unsure=False):
        """Makes a combo box holding the complete list of birds.
        Some work is needed to keep track of the indices since it's a two column
        list: species and subspecies in most cases.
        Also parses the DOC files, which use > to mark the subspecies."""
        fullbirdlist = QComboBox()
        fullbirdlist.setView(QTreeView())
        fullbirdlist.setRootModelIndex(QModelIndex())

        fullbirdlist.view().setHeaderHidden(True)
        fullbirdlist.view().setItemsExpandable(True)
        fullbirdlist.setMouseTracking(True)

        self.model = QStandardItemModel()
        headlist = []
        if self.longBirdList is not None:
            for bird in self.longBirdList:
                # Add ? marks if Ctrl menu is called
                if unsure and bird != "Don't Know" and bird != "Other":
                    bird = bird + "?"

                ind = bird.find(">")
                if ind == -1:
                    ind = len(bird)
                # find or add "genus"
                if bird[:ind] in headlist:
                    item = self.model.findItems(bird[:ind])[0]
                else:
                    headlist.append(bird[:ind])
                    item = QStandardItem(bird[:ind])
                    item.setSelectable(True)
                    self.model.appendRow(item)
                # if there's "species", add that
                if ind < len(bird):
                    subitem = QStandardItem(bird[ind + 1 :])
                    item.setSelectable(False)
                    item.appendRow(subitem)
                    subitem.setSelectable(True)
        item = QStandardItem("Other")
        item.setSelectable(True)
        self.model.appendRow(item)

        fullbirdlist.setModel(self.model)
        return fullbirdlist

    def fillBirdList(self, unsure=False):
        """Sets the contents of the context menu.
        The first 20 items are in the first menu, the next in a second menu.
        Any extras go into the combobox at the end of the second list.
        This is called a lot because the order of birds in the list changes since the last choice
        is moved to the top of the list.
        When calltype-level display is on, fills the list with some possible call types.
        """
        self.menuBirdList.clear()
        self.menuBird2.clear()

        def parse_item(item):
            # Determine certainty
            # Add ? marks if Ctrl menu is called
            if unsure and item != "Don't Know":
                cert = 50
                item = item + "?"
            elif item == "Don't Know":
                cert = 0
            else:
                cert = 100

            # Transform > marks
            pos = item.find(">")
            if pos > -1:
                item = item[:pos] + " (" + item[pos + 1 :] + ")"

            return item, cert

        if self.viewCallType:
            if not hasattr(self, "segments") or self.box1id < 0:
                return

            thisSeg = self.segments[self.box1id]
            for lab in thisSeg[4]:
                if lab["species"] == "Don't Know":
                    continue
                # add the species menu
                spMenu = self.menuBirdList.addMenu(lab["species"])

                # get call types from call type files
                possibleCTs = set()
                for ctfile in self.CalltypeDicts.values():
                    if ctfile["species"] == lab["species"]:
                        possibleCTs.update(ctfile["calltypes"])
                # add standard extras and self
                if "calltype" in lab and lab["calltype"] != "(Other)":
                    possibleCTs.add(lab["calltype"])

                possibleCT_list = []
                possibleCT_list.append("Add calltype")
                possibleCT_list.extend(sorted(possibleCTs))
                possibleCT_list.append("(Other)")
                # put them as actions in the species menu
                for ct in possibleCT_list:
                    ctitem = spMenu.addAction(ct)
                    ctitem.setCheckable(True)

                    # update check marks based on this segment
                    if "calltype" in lab and lab["calltype"] == ct:
                        ctitem.setChecked(True)
        else:
            # otherwise, fill the (correct) species list
            # Put the selected bird name at the top of the list:
            if (
                self.config["ReorderList"]
                and hasattr(self, "segments")
                and self.box1id > -1
            ):
                for key in self.segments[self.box1id].keys:
                    # Either move the label to the top of the list, or delete the last
                    if key[0] in self.shortBirdList:
                        self.shortBirdList.remove(key[0])
                    else:
                        del self.shortBirdList[-1]
                    self.shortBirdList.insert(0, key[0])

            # create menu items and mark them
            for item in self.shortBirdList[:15]:
                item_parsed, cert = parse_item(item)

                bird = self.menuBirdList.addAction(item_parsed)
                bird.setCheckable(True)
                if hasattr(self, "segments") and self.segments[self.box1id].hasLabel(
                    item, cert
                ):
                    bird.setChecked(True)
                self.menuBirdList.addAction(bird)
            self.menuBirdList.addMenu(self.menuBird2)
            for item in self.shortBirdList[15:]:
                item_parsed, cert = parse_item(item)

                bird = self.menuBird2.addAction(item_parsed)
                bird.setCheckable(True)
                if hasattr(self, "segments") and self.segments[self.box1id].hasLabel(
                    item, cert
                ):
                    bird.setChecked(True)
                self.menuBird2.addAction(bird)

            self.fullbirdlist = self.makeFullBirdList(unsure=unsure)  # a QComboBox
            self.showFullbirdlist = QWidgetAction(self.menuBirdList)
            self.showFullbirdlist.setDefaultWidget(self.fullbirdlist)
            self.menuBird2.addAction(self.showFullbirdlist)
            self.fullbirdlist.activated.connect(self.birdSelectedList)

    def fillFileList(self, dir, fileName):
        """Generates the list of files for the file listbox.
        dir - directory to use.
        fileName - currently opened file (marks it in the list).
        """
        if not os.path.isdir(dir):
            print("ERROR: directory %s doesn't exist" % dir)
            return
        self.listFiles.fill(
            dir, fileName, self.confidenceRange, self.timeRange, self.currentSpecies
        )
        self.updateListSpecies()

    def updateListSpecies(self):
        # save current species for later
        currentSpecies = self.listSpecies.currentText().rpartition(" ")[0]

        # disconnect updateListFiles from listSpecies to avoid several function calls
        self.listSpecies.currentIndexChanged.disconnect()

        self.currentSpecies = ""

        # clear species list and insert Items afterwards
        self.listSpecies.clear()
        self.listSpecies.insertItem(0, "Species (All)")
        self.listSpecies.insertItems(
            1,
            sorted(
                [
                    "{} {:.0f}".format(key, value)
                    for key, value in self.listFiles.spListCert.items()
                    if value >= self.confidenceRange[0] or key == currentSpecies
                ]
            ),
        )

        # find index of current species
        idx = self.listSpecies.findText(
            currentSpecies, QtCore.Qt.MatchFlag.MatchStartsWith
        )

        # show current species in the list if it still exists
        if currentSpecies and idx != -1:
            self.listSpecies.setCurrentIndex(idx)
        else:
            self.listSpecies.setCurrentIndex(0)
        self.currentSpecies = self.listSpecies.currentText().rpartition(" ")[0]

        # reconnect updateListFiles to listSpecies
        self.listSpecies.currentIndexChanged.connect(self.updateListSpeciesIndexChanged)

    def resetStorageArrays(self):
        """Called when new files are loaded.
        Resets the variables that hold the data to be saved and/or plotted.
        """

        # Remove the segments
        self.removeSegments()

        # Check if media is playing and stop it if so
        if hasattr(self, "media_obj"):
            if self.media_obj.isPlaying() or self.media_slow.isPlaying():
                self.stopPlayback()

        # This is a flag to say if the next thing that the user clicks on should be a start or a stop for segmentation
        if self.started:
            # This is the second click, so should pay attention and close the segment
            # Stop the mouse motion connection, remove the drawing boxes
            if self.started_window == "a":
                try:
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                self.p_ampl.removeItem(self.vLine_a)
            else:
                try:
                    self.p_spec.scene().sigMouseMoved.disconnect()
                except Exception:
                    pass
                # Add the other mouse move listener back
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                self.p_spec.removeItem(self.vLine_s)
            self.p_ampl.removeItem(self.drawingBox_ampl)
            self.p_spec.removeItem(self.drawingBox_spec)
        self.started = False
        self.startedInAmpl = False
        self.segmentsToSave = False

        # Keep track of start points and selected buttons
        self.windowStart = 0
        self.prevBoxCol = self.config["ColourNone"]
        self.bar.setValue(0)

        # Reset the MultipleSpecies option
        self.multipleBirds = self.config["MultipleSpecies"]

        # reset buttons which require segment selection
        self.refreshSegmentControls()

        # Remove any fundamental frequencies drawn
        self.showFundamental.setChecked(False)
        for r in self.segmentPlots:
            self.p_spec.removeItem(r)
        self.segmentPlots = []

        # Remove any shape marks
        for sh in self.shapePlots:
            try:
                self.p_spec.removeItem(sh)
            except Exception:
                pass

        # Remove spectral derivatives
        # self.showSpectral.setChecked(False)
        try:
            self.p_spec.removeItem(self.derivPlot)
        except Exception:
            pass

        # Remove formants
        # if not self.DOC:
        #     self.showFormant.setChecked(False)
        #     try:
        #         self.p_spec.removeItem(self.formantPlot)
        #     except Exception:
        #         pass

        # remove max energies
        self.showEnergies.setChecked(False)
        try:
            self.p_spec.removeItem(self.energyPlot)
        except Exception:
            pass

    def openFile(self, fileName=None):
        """This handles the menu items for opening a file.
        Pops up a file selection dialog if no fileName provided.
        Splits the directory name and filename out, and then passes the filename to the loader.
        """

        if fileName is None:
            # File -> Open or splash screen:
            fileName, drop = QFileDialog.getOpenFileName(
                self,
                "Choose File",
                self.SoundFileDir,
                "WAV or BMP files (*.wav *.bmp);; Only WAV files (*.wav);; Only BMP files (*.bmp)",
            )
        # (it is provided when this is called by File -> [recent file clicked])
        success = 1
        self.SoundFileDirOld = self.SoundFileDir
        fileNameOld = os.path.basename(self.filename)
        if fileName != "":
            print("Opening file %s" % fileName)
            self.SoundFileDir = os.path.dirname(fileName)
            success = self.listLoadFile(os.path.basename(fileName))
        if success == 1:
            print("Warning: could not load file, reloading current file")
            self.SoundFileDir = self.SoundFileDirOld
            self.filename = os.path.join(self.SoundFileDir, fileNameOld)
            self.listLoadFile(fileNameOld)

    def listLoadFile(self, current):
        """Listener for when the user clicks on a filename (also called by openFile() )
        Does the safety checks for file existence etc.
        Prepares the program for a new file.
        Saves the segments of the current file, resets flags and calls loadFile().
        """

        # Need name of file
        # if type(current) is QListWidgetItem:
        if type(current) is SupportClasses_GUI.SortableListWidgetItem:
            current = current.text()
            current = re.sub("\/.*", "", current)
        fullcurrent = str(pathlib.PurePath(self.SoundFileDir, current))
        if not os.path.isdir(fullcurrent):
            if not os.path.isfile(fullcurrent):
                print("File %s does not exist!" % fullcurrent)
                return 1
            # avoid files with no data (Tier 1 has 0Kb .wavs)
            if os.stat(fullcurrent).st_size == 0:
                print("Cannot open file %s of size 0!" % fullcurrent)
                return 1
            if os.stat(fullcurrent).st_size < 1000:
                print("File %s appears to have only header" % fullcurrent)
                return 1
            if fullcurrent.lower().endswith(".wav"):
                with open(fullcurrent, "br") as f:
                    if f.read(4) != b"RIFF":
                        print("WAV file %s not formatted correctly" % fullcurrent)
                        return 1
            else:
                print("Unrecognized format of file %s " % fullcurrent)
                return 1

            # setting this to True forces initial segment save
            # self.segmentsToSave = True

            # calls the noise data checks, segment saving, recent file updaters
            if self.filename is not None:
                self.closeFile()

        # restrict the file list and update indices
        self.listFiles.restrict(
            self.currentSpecies, self.confidenceRange, self.timeRange
        )
        self.listFiles.updateCurrentIndices()

        # Update the file list to show the right location
        i = 0
        lof = self.listFiles.listOfFiles
        # this is skipped on first load, when len=0
        if len(lof) > 0:
            while i < len(lof) - 1 and lof[i].fileName() != current:
                i += 1
            if lof[i].isDir() or (i == len(lof) - 1 and lof[i].fileName() != current):
                dir = QDir(self.SoundFileDir)
                dir.cd(lof[i].fileName())
                self.SoundFileDir = os.path.abspath(dir.absolutePath())

        # only repopulate the list of files if dir has changed
        if self.SoundFileDir != self.SoundFileDirOld:
            self.fillFileList(self.SoundFileDir, current)
            self.SoundFileDirOld = self.SoundFileDir

        # if a file was clicked, open it
        if not os.path.isdir(fullcurrent):
            self.loadFile(fullcurrent)

        # self.listFiles.setCurrentItem(current)

        return 0

    def toggleRankSort(self):
        self.listFiles.sortRank = self.sortRank.isChecked()
        self.listFiles.setSortingEnabled(True)
        self.listFiles.restrict(
            self.currentSpecies, self.confidenceRange, self.timeRange
        )
        self.listFiles.sortItems()
        self.listFiles.updateCurrentIndices()

    def updateTickSpeciesStateChanged(self, state):
        """This function is called when the species checkbox is clicked.
        It restricts the list of files and updates the segments if needed"""
        self.listFiles.showAll = not state
        self.listFiles.restrict(
            self.currentSpecies, self.confidenceRange, self.timeRange
        )
        self.listFiles.sortItems()
        self.listFiles.updateCurrentIndices()
        self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)
        if not self.currentSpecies == "Species":
            self.removeSegments()
            self.drawfigMain()

    def updateListSpeciesIndexChanged(self, idx):
        """This function is called when the user chooses another species in the list.
        It restricts the list of files and updates the segments if needed"""
        oldSpecies = self.currentSpecies
        self.currentSpecies = self.listSpecies.currentText().rpartition(" ")[0]
        if oldSpecies != self.currentSpecies:
            self.listFiles.currentSpecies = self.currentSpecies
            self.listFiles.restrict(
                self.currentSpecies, self.confidenceRange, self.timeRange
            )
            self.listFiles.sortItems()
            self.listFiles.updateCurrentIndices()
            self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)
            if not self.listFiles.showAll:
                self.removeSegments()
                self.drawfigMain()

    def updateConfRangeSliderValueChanged(self, values):
        """This function is called when the user chooses another confidence range.
        It restricts the list of files accordingly"""
        self.confidenceRange = values
        self.listFiles.restrict(self.currentSpecies, values, self.timeRange)
        self.listFiles.sortItems()
        self.listFiles.updateCurrentIndices()
        self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)
        self.updateListSpecies()

    def updateConfRangeSliderMoved(self, position):
        """This function is called when the user uses the mouse to move the confidence range slider.
        As tracking is turned of the min/max values have to be updated here"""
        self.confRangeSlider._min_label.setValue(position[0])
        self.confRangeSlider._max_label.setValue(position[1])

    def updateConfRangeSliderReleased(self):
        """This function is called after the user moved the confidence range slider with the mouse.
        As tracking of the confidence range slider is turned of this function calls the update
        after value changed function"""
        self.updateConfRangeSliderValueChanged(self.confRangeSlider.sliderPosition())

    def updateTimeRangeSliderValuesChanged(self, values):
        """This function is called when the user chooses another confidence range.
        It restricts the list of files accordingly"""
        self.updateTimeRangeSlidersMoved(values)
        self.listFiles.restrict(
            self.currentSpecies, self.confidenceRange, self.timeRange
        )
        self.listFiles.sortItems()
        self.listFiles.updateCurrentIndices()
        self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)

    def updateTimeRangeSlidersMoved(self, position):
        """This function is called when the user uses the mouse to move the confidence range slider.
        As tracking is turned of the min/max values have to be updated here"""
        s = datetime.timedelta(minutes=position[0] * 10)
        e = datetime.timedelta(minutes=position[1] * 10)
        self.timeRange = (s.seconds, e.days * 24 * 3600 + e.seconds)

        start = (datetime.datetime.min + s).time()
        end = (datetime.datetime.min + e).time()

        self.timeRangeStartLabel.setText(start.strftime("%H:%M"))
        if end == datetime.time(0):
            self.timeRangeEndLabel.setText("24:00")
        else:
            self.timeRangeEndLabel.setText(end.strftime("%H:%M"))

    def updateTimeRangeSliderReleased(self):
        """This function is called after the user moved the confidence range slider with the mouse.
        As tracking of the confidence range slider is turned of this function calls the update
        after value changed function"""
        self.updateTimeRangeSliderValuesChanged(self.timeRangeSlider.sliderPosition())

    def loadFile(self, name=None):
        """This does the work of loading a file.
        We are using wavio to do the reading. We turn the data into a float, but do not normalise it (/2^(15)).
        For 2 channels, just take the first one.
        Normalisation can cause problems for some segmentations, e.g. Harma.

        This method also gets the spectrogram to plot it, loads the segments from a *.data file, and
        passes the new data to any of the other classes that need it.
        Then it sets up the audio player and fills in the appropriate time data in the window, and makes
        the scroll bar and overview the appropriate lengths.

        name: full path to the file to be loaded. If None, loads the next section of the current file
        """

        self.resetStorageArrays()

        if self.listFiles.currentItem().isHidden():
            self.listFiles.currentItem().setHidden(False)
        # self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)

        with pg.ProgressDialog("Loading..", 0, 6) as dlg:
            dlg.setCancelButton(None)
            dlg.setWindowIcon(QIcon("img/PAMalyzer.ico"))
            dlg.setWindowTitle("PAMalyzer")
            dlg.show()
            dlg.update()
            if name is not None:
                if not os.path.exists(name):
                    print("ERROR: tried to open non-existing file %s", name)
                    return
                self.filename = name

                # Create an instance of the Signal Processing class
                if not hasattr(self, "sp"):
                    self.sp = SignalProc.SignalProc(
                        self.config["window_width"],
                        self.config["incr"],
                        self.config["minFreq"],
                        self.config["maxFreq"],
                    )

                self.currentFileSection = 0
                self.setTimeAxis()

            dlg += 1
            dlg.update()

            # Read in the file and make the spectrogram
            # Determine where to start and how much to read for this page (in seconds):
            self.startRead = self.currentFileSection * (
                self.config["maxFileShow"] - self.config["fileOverlap"]
            )
            # avoid files with no data (Tier 1 has 0Kb .wavs)
            if os.stat(self.filename).st_size == 0:
                self.statusLeft.setText("File appears empty")
                return

            self.sp.minFreqShow = self.config["minFreq"]
            self.sp.maxFreqShow = self.config["maxFreq"]

            self.sp.readWav(
                self.filename, len=self.config["maxFileShow"], off=self.startRead
            )
            self.datalength = np.shape(self.sp.data)[0]

            # Parse wav format details based on file header:
            self.sampleRate = self.sp.sampleRate
            self.audiodata = self.sp.data
            # self.sp.audioFormat will be set
            # self.sp.fileLength will be determined from wav header
            # self.sp.minFreq and maxFreq will be set based on sample rate
            # self.sp.*Show will be set based on SignalProc settings

            dlg += 1
            dlg.update()

            self.datalengthSec = self.datalength / self.sampleRate
            print(
                "Length of file is ",
                self.datalengthSec,
                " seconds (",
                self.datalength,
                " samples) loaded from ",
                self.sp.fileLength / self.sampleRate,
                "seconds (",
                self.sp.fileLength,
                " samples) with sample rate ",
                self.sampleRate,
                " Hz.",
            )

            if name is not None:  # i.e. starting a new file, not next section
                if self.datalength != self.sp.fileLength:
                    self.nFileSections = int(
                        np.ceil(
                            (
                                self.sp.fileLength / self.sampleRate
                                - self.config["fileOverlap"]
                            )
                            / (self.config["maxFileShow"] - self.config["fileOverlap"])
                        )
                    )
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(True)
                    self.movePrev5MinsKey.setEnabled(False)
                    self.moveNext5MinsKey.setEnabled(True)
                else:
                    self.nFileSections = 1
                    self.prev5mins.setEnabled(False)
                    self.next5mins.setEnabled(False)
                    self.movePrev5MinsKey.setEnabled(False)
                    self.moveNext5MinsKey.setEnabled(False)
                print("number of pages: ", self.nFileSections)

            # Update overview info
            if self.nFileSections == 1:
                self.placeInFileLabel.setText("(%d s in 1 page)" % self.datalengthSec)
                self.placeInFileSelector.setVisible(False)
                self.placeInFileLabel2.setVisible(False)
            else:
                self.placeInFileLabel2.setVisible(True)
                self.placeInFileSelector.setVisible(True)
                self.placeInFileSelector.setValue(self.currentFileSection + 1)
                self.placeInFileSelector.setMaximum(self.nFileSections)
                self.placeInFileLabel.setText(
                    "of %d (%d s in page)" % (self.nFileSections, self.datalengthSec)
                )
            self.fileInfoSR.setText("<b>Sampling rate:</b> %d Hz" % self.sampleRate)
            self.fileInfoNCh.setText(
                "<b>Channels:</b> %d" % self.sp.audioFormat.channelCount()
            )
            self.fileInfoSS.setText(
                "<b>Bit depth:</b> %d" % self.sp.audioFormat.sampleSize()
            )
            self.fileInfoDur.setText(
                "<b>Duration:</b> %d min %d s"
                % divmod(self.sp.fileLength // self.sampleRate, 60)
            )

            # Create the main spectrogram
            self.sp.spectrogram(
                window_width=self.config["window_width"],
                incr=self.config["incr"],
                window=str(self.windowType),
                sgType=str(self.sgType),
                sgScale=str(self.sgScale),
                nfilters=int(str(self.nfilters)),
                mean_normalise=self.sgMeanNormalise,
                equal_loudness=self.sgEqualLoudness,
                onesided=self.sgOneSided,
            )
            # Normalize the spectrogram, appropriately for the current mode and user settings
            self.setSpectrogram()

            self.drawProtocolMarks()

            self.statusRight.setText(
                "Operator: " + str(self.operator) + ", Reviewer: " + str(self.reviewer)
            )

            if hasattr(self, "seg"):
                self.seg.setNewData(self.sp)
            else:
                self.seg = Segment.Segmenter(self.sp, self.sampleRate)

            # Update the Dialogs
            # Also close any ones that could get buggy when moving between bird-bat modes
            if hasattr(self, "spectrogramDialog"):
                self.spectrogramDialog.setValues(
                    self.sp.minFreq,
                    self.sp.maxFreq,
                    self.sp.minFreqShow,
                    self.sp.maxFreqShow,
                )
            if hasattr(self, "denoiseDialog"):
                self.denoiseDialog.setValues(self.sp.minFreq, self.sp.maxFreq)

            # Delete any denoising backups from the previous file
            if hasattr(self, "audiodata_backup"):
                self.audiodata_backup = None

            self.timeaxis.setOffset(self.startRead + self.startTime)

            # Set the window size
            self.windowSize = self.config["windowWidth"]
            self.timeaxis.setRange(0, self.windowSize)
            self.widthWindow.setRange(0.01, self.datalengthSec)

            # Reset it if the file is shorter than the window
            if self.datalengthSec < self.windowSize:
                self.windowSize = self.datalengthSec
            self.widthWindow.setValue(self.windowSize)
            if self.windowSize < 3:
                self.timeaxis.setShowMS(True)

            self.totalTime = self.convertMillisecs(1000 * self.datalengthSec)

            # Load the file for playback
            self.media_obj = SupportClasses_GUI.ControllableAudio(self.sp.audioFormat)
            # this responds to audio output timer
            self.media_obj.notify.connect(self.movePlaySlider)
            # Not needed for DOC mode, but easier if it exists
            # Enable the snail button to play at other speeds
            oldSR = self.sp.audioFormat.sampleRate()
            self.sp.audioFormat.setSampleRate(
                self.sp.audioFormat.sampleRate() // self.slowSpeed
            )
            self.media_slow = SupportClasses_GUI.ControllableAudio(self.sp.audioFormat)
            self.sp.audioFormat.setSampleRate(oldSR)
            self.media_slow.notify.connect(self.movePlaySlowSlider)

            # Reset the media player
            self.stopPlayback()
            self.volSliderMoved(0)
            self.segmentStop = 50
            # self.media_obj.filterSeg(0, 50, self.audiodata)
            self.volSliderMoved(self.specControls.volSlider.value())

            # Set the length of the scrollbar.
            self.scrollSlider.setRange(
                0,
                int(
                    np.shape(self.sg)[0]
                    - self.convertAmpltoSpec(self.widthWindow.value())
                ),
            )
            self.scrollSlider.setValue(0)

            self.drawOverview()
            dlg += 1
            dlg.update()
            self.drawfigMain()
            self.setWindowTitle("PAMalyzer - Manual Processing " + self.filename)
            dlg += 1
            dlg.update()
            self.w_spec.setFocus()
            self.statusLeft.setText("Ready")
            self.listFiles.scrollToItem(self.listFiles.currentItem(), 3)

    def openPreviousFile(self, skipHidden=True):
        """Listener for previous file << button.
        Get the previous file in the list and call the loader."""

        # If the user has navigated away from the dir with currently open file, return:
        if self.listFiles.soundDir != os.path.dirname(self.filename):
            self.SoundFileDir = os.path.dirname(self.filename)
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

        i = self.listFiles.currentRow()
        if not skipHidden and i > 1:
            self.listFiles.setCurrentRow(i - 1)
            self.listLoadFile(self.listFiles.currentItem())
            return

        currentIndex = self.listFiles.currentIndices.index(i)
        if skipHidden and currentIndex > 0:
            self.listFiles.setCurrentRow(
                self.listFiles.currentIndices[currentIndex - 1]
            )

        else:
            # Tell the user they've finished
            msg = SupportClasses_GUI.MessagePopup(
                "d", "First file", "This is already the first file"
            )
            msg.exec_()
            return
        self.listLoadFile(self.listFiles.currentItem())

    def openNextFile(self, skipHidden=True):
        """Listener for next file >> button.
        Get the next file in the list and call the loader."""

        # If the user has navigated away from the dir with currently open file, return:
        if self.listFiles.soundDir != os.path.dirname(self.filename):
            self.SoundFileDir = os.path.dirname(self.filename)
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

        i = self.listFiles.currentRow()
        if not skipHidden and i + 1 < len(self.listFiles):
            self.listFiles.setCurrentRow(i + 1)
            self.listLoadFile(self.listFiles.currentItem())
            return

        currentIndex = self.listFiles.currentIndices.index(i)
        if skipHidden and currentIndex + 1 < len(self.listFiles.currentIndices):
            self.listFiles.setCurrentRow(
                self.listFiles.currentIndices[currentIndex + 1]
            )

        else:
            # Tell the user they've finished
            msg = SupportClasses_GUI.MessagePopup(
                "d", "Last file", "You've finished processing the folder"
            )
            msg.exec_()
            return

        self.listLoadFile(self.listFiles.currentItem())

    def showPointerDetailsCheck(self):
        """Listener for the menuitem that sets if detailed info should be shown when hovering over spectrogram.
        Turning this off saves lots of CPU performance."""
        self.config["showPointerDetails"] = self.showPointerDetails.isChecked()
        if self.showPointerDetails.isChecked():
            self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
            self.w_spec.addItem(self.pointDataProxy, row=3, col=1)
        else:
            self.p_spec.scene().sigMouseMoved.disconnect()
            self.w_spec.removeItem(self.pointDataProxy)

    def dragRectsTransparent(self):
        """Listener for the check menu item that decides if the user wants the dragged rectangles to have colour or not.
        It's a switch from Brush to Pen or vice versa.
        """
        if self.config["transparentBoxes"]:
            for box in self.listRectanglesa2:
                if type(box) is self.ROItype:
                    col = box.brush.color()
                    col.setAlpha(255)
                    box.transparent = True
                    box.setBrush(pg.mkBrush(None))
                    box.setPen(pg.mkPen(col, width=1))
                    box.update()
                    col.setAlpha(100)
        else:
            for box in self.listRectanglesa2:
                if type(box) is self.ROItype:
                    col = box.pen.color()
                    col.setAlpha(self.ColourNamed.alpha())
                    box.transparent = False
                    box.setBrush(pg.mkBrush(col))
                    box.setPen(pg.mkPen(None))
                    box.update()
                    col.setAlpha(100)

    def useAmplitudeCheck(self):
        """Listener for the check menu item saying if the user wants to see the waveform.
        Does not remove the dock, just hides it. It's therefore easy to replace, but could have some performance overhead.
        """
        if self.useAmplitudeTick.isChecked():
            self.useAmplitude = True
            self.d_ampl.show()
        else:
            self.useAmplitude = False
            self.d_ampl.hide()
        self.config["showAmplitudePlot"] = self.useAmplitudeTick.isChecked()

    def useFilesCheck(self):
        """Listener to process if the user swaps the check menu item to see the file list."""
        if self.useFilesTick.isChecked():
            self.d_files.show()
        else:
            self.d_files.hide()
        self.config["showListofFiles"] = self.useFilesTick.isChecked()

    def showOverviewSegsCheck(self):
        """Listener to process if the user swaps the check menu item to see the overview segment boxes."""
        if self.showOverviewSegsTick.isChecked():
            self.p_overview2.show()
        else:
            self.p_overview2.hide()
        self.config["showAnnotationOverview"] = self.showOverviewSegsTick.isChecked()

    def makeReadOnly(self):
        """Listener to process the check menu item to make the plots read only.
        Turns off the listeners for the amplitude and spectrogram plots.
        Also has to go through all of the segments, turn off the listeners, and make them unmovable.
        """
        self.config["readOnly"] = self.readonly.isChecked()
        self.statusRO.setText("Read-only mode" if self.config["readOnly"] else "")
        if self.readonly.isChecked():
            # this is for accepting drag boxes or not
            self.p_spec.enableDrag = False

            # when clicking is used to draw segments/boxes,
            # read-only changes are implemented in the button signals.
            # Because connecting-disconnecting slots is very dirty.

            # this will re-make segment boxes with correct moving abilities:
            if hasattr(self, "sp"):
                self.removeSegments(delete=False)
                self.drawfigMain(reusing=True)
        else:
            self.p_spec.enableDrag = self.config["specMouseAction"] == 3
            if hasattr(self, "sp"):
                self.removeSegments(delete=False)
                self.drawfigMain(reusing=True)

    def dockReplace(self):
        """Listener for if the docks should be replaced menu item.
        A rewrite of pyqtgraph.dockarea.restoreState.
        """
        containers, docks = self.area.findAll()
        # main recursion of restoreState:
        self.area.buildFromState(self.state["main"], docks, self.area, missing="error")
        # restoreState doesn't restore non-floating window sizes smh
        containers, docks = self.area.findAll()
        # basically say that left panel and controls should be as small as possible:
        self.d_controls.setSizePolicy(1, 1)
        containers[1].setSizePolicy(1, 1)
        # self.useAmplitudeTick.setChecked(True)
        # self.useAmplitude = True
        # self.config['showAmplitudePlot'] = True
        self.useFilesTick.setChecked(True)
        self.config["showListofFiles"] = True
        self.showOverviewSegsTick.setChecked(True)
        self.config["showAnnotationOverview"] = True
        self.useAmplitudeCheck()
        # for cont in range(len(containers)):
        #     containers[cont].setSizes(self.state_cont[cont])

    def showFundamentalFreq(self):
        """Calls the SignalProc class to compute, and then draws the fundamental frequency.
        Uses the yin algorithm."""

        with pg.BusyCursor():
            if self.showFundamental.isChecked():
                self.statusLeft.setText("Drawing fundamental frequency...")
                segs = self.sp.drawFundFreq(self.seg)

                # Get the individual pieces
                self.segmentPlots = []
                # draw each contiguous "segment" of fund freq
                for s in segs:
                    self.segmentPlots.append(pg.PlotDataItem())
                    self.segmentPlots[-1].setData(
                        s[0], s[1], pen=pg.mkPen("r", width=3)
                    )
                    self.p_spec.addItem(self.segmentPlots[-1])
            else:
                self.statusLeft.setText("Removing fundamental frequency...")
                for r in self.segmentPlots:
                    self.p_spec.removeItem(r)
                self.segmentPlots = []
            self.statusLeft.setText("Ready")

    def setTimeAxis(self):
        if hasattr(self, "timeaxis"):
            self.w_spec.removeItem(self.timeaxis)

        # Check if the filename is in standard DOC format
        # Which is xxxxxx_xxxxxx.wav or ccxx_cccc_xxxxxx_xxxxxx.wav (c=char, x=0-9), could have _ afterward
        # So this checks for the 6 ints _ 6 ints part anywhere in string
        doc, startTime = self.listFiles.currentItem().getStartTimeDOC()
        self.startTime = startTime

        if doc:
            self.timeaxis = SupportClasses_GUI.TimeAxisHour(
                orientation="bottom", linkView=self.p_ampl
            )
        else:
            self.timeaxis = SupportClasses_GUI.TimeAxisMin(
                orientation="bottom", linkView=self.p_ampl
            )

        self.w_spec.addItem(self.timeaxis, row=1, col=1)

    def showMaxEnergy(self):
        with pg.BusyCursor():
            if self.showEnergies.isChecked():
                self.statusLeft.setText("Drawing max energies...")
                x, y = self.sp.max_energy(self.sg)

                self.energyPlot = pg.ScatterPlotItem()
                self.energyPlot.setBrush(None)
                self.energyPlot.setData(
                    x, y, brush=pg.mkBrush((0, 255, 0, 130)), pen=pg.mkPen(None), size=5
                )

                self.p_spec.addItem(self.energyPlot)
            else:
                self.statusLeft.setText("Removing max energies...")
                self.p_spec.removeItem(self.energyPlot)
            self.statusLeft.setText("Ready")

    # def showSpectralDeriv(self):
    #     with pg.BusyCursor():
    #         if self.showSpectral.isChecked():
    #             self.statusLeft.setText("Drawing spectral derivative...")
    #             x, y = self.sp.drawSpectralDeriv()
    #
    #             self.derivPlot = pg.ScatterPlotItem()
    #             self.derivPlot.setData(x, y, pen=pg.mkPen("b", width=5))
    #
    #             self.p_spec.addItem(self.derivPlot)
    #         else:
    #             self.statusLeft.setText("Removing spectral derivative...")
    #             self.p_spec.removeItem(self.derivPlot)
    #         self.statusLeft.setText("Ready")

    # def showFormants(self):
    #     with pg.BusyCursor():
    #         if self.showFormant.isChecked():
    #             self.statusLeft.setText("Drawing formants...")
    #             x, y = self.sp.drawFormants()
    #             self.formantPlot = pg.ScatterPlotItem()
    #             # step = self.config['window_width'] // self.config['incr']
    #             # starts = np.arange(0,np.shape(self.sg)[0],step)
    #             # for i in range(len(starts)):
    #             # for j in range(len(formants[i])):
    #             # self.formantPlot.addPoints(starts[i], formants[i][j], pen=pg.mkPen('b', width=5))
    #             self.formantPlot.setData(x, y, pen=pg.mkPen("b", width=0.05))
    #             self.p_spec.addItem(self.formantPlot)
    #         else:
    #             self.statusLeft.setText("Removing formants...")
    #             self.p_spec.removeItem(self.formantPlot)
    #         self.statusLeft.setText("Ready")

    def drawGuidelines(self):
        # Frequency guides for bat mode
        print("Updating guidelines...")
        if self.config["guidelinesOn"]:
            for gi in range(len(self.guidelines)):
                self.guidelines[gi].setValue(
                    self.convertFreqtoY(self.config["guidepos"][gi])
                )
                self.guidelines[gi].setPen(color=self.config["guidecol"][gi], width=2)
                self.p_spec.addItem(self.guidelines[gi], ignoreBounds=True)
        else:
            # easy way to hide
            for g in self.guidelines:
                g.setValue(-1000)

    # def showCQT(self):
    #     cqt = self.sp.comp_cqt()
    #     print(np.shape(cqt),np.shape(self.sg))
    #     self.specPlot.setImage(10*np.log10(np.real(cqt*np.conj(cqt))).T)
    #     self.p_spec.setXRange(0, np.shape(cqt)[1], update=True, padding=0)

    def medianFilterSpec(self):
        """Median filter the spectrogram. To be used in conjunction with spectrogram inversion."""
        # TODO: Play with this
        with pg.BusyCursor():
            self.statusLeft.setText("Filtering...")
            median_filter(self.sg, size=(100, 20))
            self.specPlot.setImage(self.sg)
            self.statusLeft.setText("Ready")

    def convertAmpltoSpec(self, x):
        """Unit conversion"""
        return x * self.sampleRate / self.config["incr"]

    def convertSpectoAmpl(self, x):
        """Unit conversion"""
        return x * self.config["incr"] / self.sampleRate

    def convertMillisecs(self, millisecs):
        """Unit conversion"""
        seconds = (millisecs / 1000) % 60
        minutes = (millisecs / (1000 * 60)) % 60
        return "%02d" % minutes + ":" + "%02d" % seconds

    def convertYtoFreq(self, y, sgy=None):
        """Unit conversion"""
        if sgy is None:
            sgy = np.shape(self.sg)[1]
        return y * self.sampleRate // 2 / sgy + self.sp.minFreqShow

    def convertFreqtoY(self, f):
        """Unit conversion"""
        sgy = np.shape(self.sg)[1]
        return (f - self.sp.minFreqShow) * sgy / (self.sampleRate // 2)

    def drawOverview(self):
        """On loading a new file, update the overview figure to show where you are up to in the file.
        Also, compute the new segments for the overview, make sure that the listeners are connected
        for clicks on them, and disconnect old listeners."""
        self.overviewImage.setImage(self.sg)
        self.overviewImageRegion.setBounds([0, len(self.sg)])
        self.overviewImageRegion.setRegion(
            [0, self.convertAmpltoSpec(self.widthWindow.value())]
        )
        try:
            self.overviewImageRegion.sigRegionChangeFinished.disconnect()
        except Exception:
            pass
        self.overviewImageRegion.sigRegionChangeFinished.connect(self.updateOverview)

        # Three y values are No. not known, No. known, No. possible
        # widthOverviewSegment is in seconds
        numSegments = int(
            np.ceil(
                np.shape(self.sg)[0]
                / self.convertAmpltoSpec(self.config["widthOverviewSegment"])
            )
        )
        self.widthOverviewSegment = np.shape(self.sg)[0] // numSegments

        # init self.overviewSegments: holds number of segments for each label category (Don't know, Possible, Named) per overview segment
        self.overviewSegments = np.zeros((numSegments, 3))

        # Delete the overview segments
        for r in self.SegmentRects:
            self.p_overview2.removeItem(r)
        self.SegmentRects = []

        # Add new overview segments
        for i in range(numSegments):
            r = SupportClasses_GUI.ClickableRectItem(
                i * self.widthOverviewSegment, 0, self.widthOverviewSegment, 1
            )
            r.setPen(pg.mkPen(100, 100, 100))
            r.setBrush(pg.mkBrush("w"))
            self.SegmentRects.append(r)
            self.p_overview2.addItem(r)
        try:
            self.p_overview2.sigChildMessage.disconnect()
        except Exception:
            pass
        self.p_overview2.sigChildMessage.connect(self.overviewSegmentClicked)
        self.p_overview2.setYRange(-0.2, 1, padding=0.02)

    def overviewSegmentClicked(self, x):
        """Listener for an overview segment being clicked on.
        Work out which one, and move the region appropriately. Calls updateOverview to do the work.
        """
        minX, maxX = self.overviewImageRegion.getRegion()
        halfwin = (maxX - minX) / 2
        self.overviewImageRegion.setRegion([x - halfwin, x + halfwin])

    def updateOverview(self):
        """Listener for when the overview box is changed. Other functions call it indirectly by setRegion.
        Does the work of keeping all the plots in the right place as the overview moves.
        It sometimes updates a bit slowly."""
        if hasattr(self, "media_obj"):
            if (
                self.media_obj.state() == QAudio.ActiveState
                or self.media_obj.state() == QAudio.SuspendedState
                or self.media_slow.state() == QAudio.ActiveState
            ):
                self.stopPlayback()

        minX, maxX = self.overviewImageRegion.getRegion()

        # (the region bounds are checked against spec size in our subclass)

        # Temporarily block callback, and update window size (b/c setRegion may have changed it to fit bounds)
        self.updateRequestedByOverview = True
        self.widthWindow.setValue(self.convertSpectoAmpl(maxX - minX))
        self.p_ampl.setXRange(
            self.convertSpectoAmpl(minX),
            self.convertSpectoAmpl(maxX),
            update=True,
            padding=0,
        )
        self.p_spec.setXRange(minX, maxX, update=True, padding=0)

        # # # I know the next two lines SHOULD be unnecessary. But they aren't!
        self.p_ampl.setXRange(
            self.convertSpectoAmpl(minX), self.convertSpectoAmpl(maxX), padding=0
        )
        self.p_spec.setXRange(minX, maxX, padding=0)

        self.scrollSlider.setValue(int(minX))
        self.config["windowWidth"] = self.convertSpectoAmpl(maxX - minX)
        # self.saveConfig = True
        self.timeaxis.update()
        QApplication.processEvents()
        self.updateRequestedByOverview = False

    def drawfigMain(self, reusing=False):
        """Draws the main amplitude and spectrogram plots and any segments on them.
        Has to do some work to get the axis labels correct.
        """
        if len(self.audiodata) > 0:
            self.amplPlot.setData(
                np.linspace(
                    0.0, self.datalengthSec, num=self.datalength, endpoint=True
                ),
                self.audiodata,
            )

        self.timeaxis.setLabel("")

        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        pixelstart = int(self.sp.minFreqShow / height)
        pixelend = int(self.sp.maxFreqShow / height)

        self.overviewImage.setImage(self.sg[:, pixelstart:pixelend])
        self.overviewImageRegion.setBounds([0, len(self.sg)])
        self.specPlot.setImage(self.sg[:, pixelstart:pixelend])

        self.setColourMap(self.config["cmap"])
        self.setColourLevels()

        # Sort out the spectrogram frequency axis
        # The constants here are divided by 1000 to get kHz, and then remember the top is sampleRate/2

        # There are two options for logarithmic axis (Mel/Bark): keep the numbers equally spaced, but correct the labels, or keep the numbers but space the labels correctly.
        # I'm doing the first for now.

        FreqRange = self.sp.maxFreqShow - self.sp.minFreqShow
        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        SpecRange = FreqRange / height
        self.drawGuidelines()

        labels = [
            self.sp.minFreqShow,
            self.sp.minFreqShow + FreqRange / 4,
            self.sp.minFreqShow + FreqRange / 2,
            self.sp.minFreqShow + 3 * FreqRange / 4,
            self.sp.minFreqShow + FreqRange,
        ]

        if self.sgScale == "Mel Frequency":
            for i in range(len(labels)):
                labels[i] = self.sp.convertHztoMel(labels[i])
            self.specaxis.setLabel("Mels")
        elif self.sgScale == "Bark Frequency":
            for i in range(len(labels)):
                labels[i] = self.sp.convertHztoBark(labels[i])
            self.specaxis.setLabel("Barks")
        else:
            self.specaxis.setLabel("kHz")

        self.specaxis.setTicks(
            [
                [
                    (0, str(round(labels[0] / 1000, 2))),
                    (SpecRange / 4, str(round(labels[1] / 1000, 2))),
                    (SpecRange / 2, str(round(labels[2] / 1000, 2))),
                    (3 * SpecRange / 4, str(round(labels[3] / 1000, 2))),
                    (SpecRange, str(round(labels[4] / 1000, 2))),
                ]
            ]
        )
        # self.specaxis.setTicks([[(0,round(self.sp.minFreqShow/1000, 2)),
        # (SpecRange/4,round(self.sp.minFreqShow/1000+FreqRange/4000, 2)),
        # (SpecRange/2,round(self.sp.minFreqShow/1000+FreqRange/2000, 2)),
        # (3*SpecRange/4,round(self.sp.minFreqShow/1000+3*FreqRange/4000, 2)),
        # (SpecRange,round(self.sp.minFreqShow/1000+FreqRange/1000, 2))]])

        self.updateOverview()
        self.textpos = int(
            (self.sp.maxFreqShow - self.sp.minFreqShow) / height
        )  # + self.config['textoffset']

        # This is the moving bar for the playback
        self.p_spec.addItem(self.bar, ignoreBounds=True)
        if not reusing:
            # ANNOTATIONS: init empty list
            self.segments = Segment.SegmentList()
            # Load any previous segments stored
            self.segments.getData(self, self.filename)
            # If there are segments, show them
        for count in range(len(self.segments)):
            self.addSegment(
                startpoint=self.segments[count][0],
                endpoint=self.segments[count][1],
                y1=self.segments[count][2],
                y2=self.segments[count][3],
                species=self.segments[count][4],
                saveSeg=False,
                index=count,
                reusing=reusing,
                coordsAbsolute=True,
            )
        # redraw file icon if currently marked item in the list is open file
        if self.filename == os.path.join(
            self.listFiles.soundDir, self.listFiles.currentItem().text()
        ):
            self.refreshFileColor()

        QApplication.processEvents()

    def setSpeed(self, speed):
        # self.speedButton.setText(speed)
        if type(speed) is str:
            # convert Unicode fractions to floats
            speedchar = ord(speed)
            if speedchar == 188:
                speed = 0.25
            elif speedchar == 189:
                speed = 0.5
        self.slowSpeed = 1 / float(speed)
        oldSR = self.sp.audioFormat.sampleRate()
        self.sp.audioFormat.setSampleRate(
            self.sp.audioFormat.sampleRate() // self.slowSpeed
        )
        self.media_slow = SupportClasses_GUI.ControllableAudio(self.sp.audioFormat)
        print("modified playback speed set to Fs =", self.sp.audioFormat.sampleRate())
        self.sp.audioFormat.setSampleRate(oldSR)
        try:
            self.media_slow.notify.disconnect()
        except Exception:
            pass
        self.media_slow.notify.connect(self.movePlaySlowSlider)

    def updateRegion_spec(self):
        """This is the listener for when a segment box is changed in the spectrogram.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa2[i] != sender and i < len(self.listRectanglesa2):
            i = i + 1
        if i == len(self.listRectanglesa2):
            print("ERROR: segment not found!")
            return

        # update the overview boxes, step 1
        self.refreshOverviewWith(self.segments[i], delete=True)

        # save affected segment for future reference in update
        old_segment = copy.copy(self.segments[i])
        # fix the position of the text label
        if type(sender) is self.ROItype:
            # using box coordinates
            x1 = self.convertSpectoAmpl(sender.pos()[0])
            x2 = self.convertSpectoAmpl(sender.pos()[0] + sender.size()[0])
            self.segments[i][2] = int(self.convertYtoFreq(sender.pos()[1]))
            self.segments[i][3] = int(
                self.convertYtoFreq(sender.pos()[1] + sender.size()[1])
            )
            self.listLabels[i].setPos(sender.pos()[0], self.textpos)
        else:
            # using segment coordinates
            x1 = self.convertSpectoAmpl(sender.getRegion()[0])
            x2 = self.convertSpectoAmpl(sender.getRegion()[1])
            self.listLabels[i].setPos(sender.getRegion()[0], self.textpos)

        # update the amplitude segment
        self.listRectanglesa1[i].blockSignals(True)
        self.listRectanglesa1[i].setRegion([x1, x2])
        self.listRectanglesa1[i].blockSignals(False)

        self.segments[i][0] = x1 + self.startRead
        self.segments[i][1] = x2 + self.startRead

        # update segment in database
        self.db.update_segment(
            old_segment,
            self.segments[i],
            self.SoundFileDir,
            self.listFiles.currentItem().text(),
        )
        # update the overview boxes, step 2
        self.refreshOverviewWith(self.segments[i])

    def updateRegion_ampl(self):
        """This is the listener for when a segment box is changed in the waveform plot.
        It updates the position of the matching box, and also the text within it.
        """
        sender = self.sender()
        i = 0
        while self.listRectanglesa1[i] != sender and i < len(self.listRectanglesa1):
            i = i + 1
        if i == len(self.listRectanglesa1):
            print("Segment not found!")

        else:
            old_segment = copy.copy(self.segments[i])
            # update the overview boxes, step 1
            self.refreshOverviewWith(self.segments[i], delete=True)

            # fix the position of the text label
            x1 = self.convertAmpltoSpec(sender.getRegion()[0])
            x2 = self.convertAmpltoSpec(sender.getRegion()[1])
            self.listLabels[i].setPos(x1, self.textpos)

            # update the corresponding spectrogram segment
            self.listRectanglesa2[i].blockSignals(True)
            if type(self.listRectanglesa2[i]) is self.ROItype:
                # update the box
                y1 = self.listRectanglesa2[i].pos().y()
                y2 = self.listRectanglesa2[i].size().y()
                self.listRectanglesa2[i].setPos(pg.Point(x1, y1))
                self.listRectanglesa2[i].setSize(pg.Point(x2 - x1, y2))
            else:
                # update the segment
                self.listRectanglesa2[i].setRegion([x1, x2])
            self.listRectanglesa2[i].blockSignals(False)

            self.segments[i][0] = sender.getRegion()[0] + self.startRead
            self.segments[i][1] = sender.getRegion()[1] + self.startRead

            self.db.update_segment(
                old_segment,
                self.segments[i],
                self.SoundFileDir,
                self.listFiles.currentItem().text(),
            )
            self.db.commit()
            # update the overview boxes, step 2
            self.refreshOverviewWith(self.segments[i])

    def addRegularSegments(self):
        """Perform the Hartley bodge: add 10s segments every minute."""
        if self.box1id > -1:
            self.deselectSegment(self.box1id)
        segtimes = [(seg[0], seg[1]) for seg in self.segments]
        i = 0
        print(
            "Adding segments (%d s every %d s)"
            % (self.config["protocolSize"], self.config["protocolInterval"])
        )
        while i < self.segments.metadata["Duration"]:
            # check for segment presence in case of double click or other issues
            if len(segtimes) > 0 and (i, i + self.config["protocolSize"]) in segtimes:
                print("segment already exists, skipping")
            else:
                self.addSegment(i, i + self.config["protocolSize"], coordsAbsolute=True)
            i += self.config["protocolInterval"]
        self.segmentsToSave = True

    def drawProtocolMarks(self):
        # if check-ignore protocol is used, mark check-ignore limits.
        # Also called when the relevant parameters are changed in interface settings.

        # Clean old marks, if any
        if hasattr(self, "protocolMarks"):
            for m in self.protocolMarks:
                self.p_spec.removeItem(m)
        self.protocolMarks = []

        if self.config["protocolOn"]:
            linePen = pg.mkPen((148, 0, 211), width=5)
            linestart = 0

            # pages >1 start with an overlap zone, so need to offset marks:
            if self.currentFileSection > 0:
                linestart += self.config["fileOverlap"]
            while linestart < self.datalength / self.sampleRate:
                lineend = min(
                    self.datalength / self.sampleRate,
                    linestart + self.config["protocolSize"],
                )
                # print("Adding to", linestart, lineend)
                line = pg.ROI(
                    pos=(self.convertAmpltoSpec(linestart), 0),
                    size=(self.convertAmpltoSpec(lineend - linestart), 0),
                    movable=False,
                    pen=linePen,
                )
                self.protocolMarks.append(line)
                self.p_spec.addItem(line)
                linestart += self.config["protocolInterval"]

    def refreshOverviewWith(self, segment, delete=False):
        """Recalculates the overview box colours and refreshes their display.
        To be used when segments are added, deleted or moved.
        Takes Segments as an input and either removes or adds to the box counters."""

        # Work out which overview segment this segment is in (could be more than one)
        # max/min deal with segments continuing past the edge of current page
        inds = max(
            0,
            int(
                self.convertAmpltoSpec(segment[0] - self.startRead)
                / self.widthOverviewSegment
            ),
        )
        inde = min(
            int(
                self.convertAmpltoSpec(segment[1] - self.startRead)
                / self.widthOverviewSegment
            ),
            len(self.overviewSegments) - 1,
        )

        for label in segment[4]:
            if label["certainty"] == 0:
                # "red" label counter
                if delete:
                    self.overviewSegments[inds : inde + 1, 0] -= 1
                else:
                    self.overviewSegments[inds : inde + 1, 0] += 1
            elif label["certainty"] == 100:
                # "green" label counter
                if delete:
                    self.overviewSegments[inds : inde + 1, 1] -= 1
                else:
                    self.overviewSegments[inds : inde + 1, 1] += 1
            else:
                # "yellow" label counter
                if delete:
                    self.overviewSegments[inds : inde + 1, 2] -= 1
                else:
                    self.overviewSegments[inds : inde + 1, 2] += 1

        if np.any(self.overviewSegments < 0):
            print("Warning: something went wrong with overview colors!")
            print(self.overviewSegments)

        # set the colour of these boxes in the overview
        for box in range(inds, inde + 1):
            if self.overviewSegments[box, 0] > 0:
                self.SegmentRects[box].setBrush(self.ColourNone)
            elif self.overviewSegments[box, 2] > 0:
                self.SegmentRects[box].setBrush(self.ColourPossible)
            elif self.overviewSegments[box, 1] > 0:
                self.SegmentRects[box].setBrush(self.ColourNamed)
            else:
                # boxes w/o segments
                self.SegmentRects[box].setBrush(pg.mkBrush("w"))
            self.SegmentRects[box].update()

    def addSegment(
        self,
        startpoint,
        endpoint,
        y1=0,
        y2=0,
        species=[],
        saveSeg=True,
        index=-1,
        reusing=False,
        coordsAbsolute=False,
    ):
        """When a new segment is created, does the work of creating it and connecting its
        listeners. Also updates the relevant overview segment.
        If a segment is too long for the current section, truncates it.
        Args:
        startpoint, endpoint - in secs, either from page start, or absolute (then set coordsAbsolute=True)
        y1, y2 should be the frequencies (between 0 and Fs//2)
        species - list of labels (including certainties, .data format)
        saveSeg - store the created segment on self.segments. Set to False when drawing the saved ones.
        reusing - can be turned to True to reuse existing graphics objects
        coordsAbsolute - set to True to accept start,end in absolute coords (from file start)
        """
        print("Segment added at %d-%d, %d-%d" % (startpoint, endpoint, y1, y2))
        if self.box1id > -1:
            self.deselectSegment(self.box1id)

        # Make sure startpoint and endpoint are in the right order
        if startpoint > endpoint:
            startpoint, endpoint = endpoint, startpoint
        # same for freqs
        if y1 > y2:
            y1, y2 = y2, y1
        # since we allow passing empty list here:
        if len(species) == 0:
            species = [{"species": "Don't Know", "certainty": 0, "filter": "M"}]

        if coordsAbsolute:
            # convert from absolute times to relative-to-page times
            startpoint = startpoint - self.startRead
            endpoint = endpoint - self.startRead
        species_list = [entry["species"] for entry in species]
        conf_list = [entry["certainty"] for entry in species]

        if not saveSeg:
            # check if this segment fits in the current spectrogram page
            if endpoint < 0 or startpoint > self.datalengthSec:
                print("Warning: a segment was not shown")
                show = False
            elif (
                y1 != 0
                and y2 != 0
                and (y1 > self.sp.maxFreqShow or y2 < self.sp.minFreqShow)
            ):
                print("Warning: a segment was not shown")
                show = False
            else:
                show = True
            if (
                not self.listFiles.showAll
                and self.currentSpecies != "Species"
                and 100 not in conf_list
                and self.currentSpecies not in species_list
            ):
                show = False

        else:
            self.segmentsToSave = True
            show = True

        if saveSeg or show:
            # create a Segment. this will check for errors and standardize the labels
            # Note - we convert time from _relative to page_ to _relative to file start_
            newSegment = Segment.Segment(
                [
                    startpoint + self.startRead,
                    endpoint + self.startRead,
                    y1,
                    y2,
                    species,
                ]
            )

            # Add the segment to the data
            if saveSeg:
                self.segments.append(newSegment)

        if not show:
            # Add a None element into the array so that the correct boxids work
            if reusing:
                self.listRectanglesa1[index] = None
                self.listRectanglesa2[index] = None
                self.listLabels[index] = None
            else:
                self.listRectanglesa1.append(None)
                self.listRectanglesa2.append(None)
                self.listLabels.append(None)
            return
        # otherwise, this is a visible segment.

        # --- rest of this function only does the graphics ---
        cert = min([lab["certainty"] for lab in species])
        if cert == 0:
            self.prevBoxCol = self.ColourNone
        elif cert == 100:
            self.prevBoxCol = self.ColourNamed
        else:
            self.prevBoxCol = self.ColourPossible

        self.refreshOverviewWith(newSegment)

        segsMovable = not (self.config["readOnly"])
        scenerect = QRectF(0, 0, np.shape(self.sg)[0], np.shape(self.sg)[1])

        # Add the segment in both plots and connect up the listeners
        p_ampl_r = SupportClasses_GUI.LinearRegionItem2(
            self,
            brush=self.prevBoxCol,
            movable=segsMovable,
            bounds=[0, self.datalengthSec],
        )
        self.p_ampl.addItem(p_ampl_r, ignoreBounds=True)
        p_ampl_r.setRegion([startpoint, endpoint])
        p_ampl_r.sigRegionChangeFinished.connect(self.updateRegion_ampl)

        # full-height segments:
        if y1 == 0 and y2 == 0:
            # filled-in segments normally, transparent ones for bats:
            p_spec_r = None
            p_spec_r = SupportClasses_GUI.LinearRegionItem2(
                self,
                brush=self.prevBoxCol,
                movable=segsMovable,
                bounds=[0, np.shape(self.sg)[0]],
            )
            p_spec_r.setRegion(
                [self.convertAmpltoSpec(startpoint), self.convertAmpltoSpec(endpoint)]
            )
        # rectangle boxes:
        else:
            specy1 = self.convertFreqtoY(max(y1, self.sp.minFreqShow))
            specy2 = self.convertFreqtoY(min(y2, self.sp.maxFreqShow))
            startpointS = QPointF(self.convertAmpltoSpec(startpoint), specy1)
            endpointS = QPointF(self.convertAmpltoSpec(endpoint), specy2)
            p_spec_r = SupportClasses_GUI.ShadedRectROI(
                startpointS,
                endpointS - startpointS,
                movable=segsMovable,
                maxBounds=scenerect,
                parent=self,
            )
            if self.config["transparentBoxes"]:
                col = self.prevBoxCol.rgb()
                col = QtGui.QColor(col)
                col.setAlpha(255)
                p_spec_r.transparent = True
                p_spec_r.setBrush(None)
                p_spec_r.setHoverBrush(None)
                p_spec_r.setPen(pg.mkPen(col, width=1))
                col.setAlpha(100)
            else:
                p_spec_r.setBrush(pg.mkBrush(self.prevBoxCol))
                col = self.prevBoxCol
                col.setAlpha(180)
                p_spec_r.transparent = False
                p_spec_r.setHoverBrush(pg.mkBrush(col))
                p_spec_r.setPen(pg.mkPen(None))
                col.setAlpha(100)
        self.p_spec.addItem(p_spec_r, ignoreBounds=True)
        p_spec_r.sigRegionChangeFinished.connect(self.updateRegion_spec)

        # Put the text into the box
        label = pg.TextItem(text="new", color="k", anchor=(0, 1))
        # label = pg.TextItem(text=species, color='k')
        self.p_spec.addItem(label)
        label.setPos(self.convertAmpltoSpec(startpoint), self.textpos)

        # Add the segments to the relevent lists
        if reusing:
            self.listRectanglesa1[index] = p_ampl_r
            self.listRectanglesa2[index] = p_spec_r
            self.listLabels[index] = label
        else:
            self.listRectanglesa1.append(p_ampl_r)
            self.listRectanglesa2.append(p_spec_r)
            self.listLabels.append(label)

        # mark this as the current segment
        if index > -1:
            box1id = index
        else:
            box1id = len(self.listLabels) - 1

        # update its displayed label
        self.updateText(box1id)

    def selectSegment(self, boxid):
        """Changes the segment colors and enables playback buttons."""
        # print("selected %d" % boxid)
        self.box1id = boxid
        self.refreshSegmentControls()

        # helps dealing with edge effects for various review functions
        if boxid > len(self.listRectanglesa1) or self.listRectanglesa1[boxid] is None:
            return

        self.prevBoxCol = self.listRectanglesa1[boxid].brush.color()
        brush = fn.mkBrush(self.ColourSelected)
        if (
            self.listRectanglesa1[boxid] is not None
            and self.listRectanglesa2[boxid] is not None
        ):
            self.listRectanglesa1[boxid].setBrush(brush)
            self.listRectanglesa2[boxid].setBrush(brush)
            self.listRectanglesa1[boxid].setHoverBrush(brush)
            self.listRectanglesa2[boxid].setHoverBrush(brush)

            self.listRectanglesa1[boxid].update()
            self.listRectanglesa2[boxid].update()

        # show details of selection
        self.segInfo.setText(self.segments[boxid].infoString())

    def deselectSegment(self, boxid):
        """Restores the segment colors and disables playback buttons."""
        # print("deselected %d" % boxid)
        self.box1id = -1
        self.refreshSegmentControls()
        # hide details of selection
        self.segInfo.setText("")

        # helps dealing with edge effects for various review functions
        if boxid > len(self.listRectanglesa1) or self.listRectanglesa1[boxid] is None:
            return

        # filled-in segments normally, transparent ones for bats:
        # (This is somewhat convoluted to keep amplitude segments updated even in bat mode,
        # as they are used for tracking prevBoxCol)
        col = self.prevBoxCol
        # col.setAlpha(100)
        self.listRectanglesa1[boxid].setBrush(fn.mkBrush(col))
        self.listRectanglesa2[boxid].setBrush(fn.mkBrush(col))
        self.listRectanglesa1[boxid].setHoverBrush(fn.mkBrush(col))
        self.listRectanglesa2[boxid].setHoverBrush(fn.mkBrush(col))
        # col.setAlpha(100)

        if (
            self.config["transparentBoxes"]
            and type(self.listRectanglesa2[boxid]) is self.ROItype
        ):
            col = self.prevBoxCol.rgb()
            col = QtGui.QColor(col)
            # col.setAlpha(255)
            self.listRectanglesa2[boxid].setBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setHoverBrush(pg.mkBrush(None))
            self.listRectanglesa2[boxid].setPen(col, width=1)
            # col.setAlpha(100)

        self.listRectanglesa1[boxid].update()
        self.listRectanglesa2[boxid].update()

    ### mouse management

    def mouseMoved(self, evt):
        """Listener for mouse moves.
        If the user moves the mouse in the spectrogram, print the time, frequency, power for the mouse location.
        """
        if not self.showPointerDetails.isChecked():
            return
        elif self.p_spec.sceneBoundingRect().contains(evt):
            mousePoint = self.p_spec.mapSceneToView(evt)
            indexx = int(mousePoint.x())
            indexy = int(mousePoint.y())
            if (
                indexx > 0
                and indexx < np.shape(self.sg)[0]
                and indexy > 0
                and indexy < np.shape(self.sg)[1]
            ):
                time = (
                    self.convertSpectoAmpl(mousePoint.x())
                    + self.currentFileSection * self.config["maxFileShow"]
                    - (self.currentFileSection > 0) * self.config["fileOverlap"]
                    + self.startTime
                )
                seconds = time % 60
                minutes = (time // 60) % 60
                hours = (time // 3600) % 24
                if hours > 0:
                    self.pointData.setText(
                        "time=%.2d:%.2d:%05.3f (hh:mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)"
                        % (
                            hours,
                            minutes,
                            seconds,
                            mousePoint.y() * self.sampleRate // 2 / np.shape(self.sg)[1]
                            + self.sp.minFreqShow,
                            self.sg[indexx, indexy],
                        )
                    )
                else:
                    self.pointData.setText(
                        "time=%.2d:%05.2f (mm:ss.ms), freq=%0.1f (Hz), power=%0.1f (dB)"
                        % (
                            minutes,
                            seconds,
                            mousePoint.y() * self.sampleRate // 2 / np.shape(self.sg)[1]
                            + self.sp.minFreqShow,
                            self.sg[indexx, indexy],
                        )
                    )

    def mouseClicked_ampl(self, evt):
        """Listener for if the user clicks on the amplitude plot.
        If there is a box selected, get its colour.
        If the user has clicked inside the scene, they could be
        (1) clicking in an already existing box -> select it
        (2) clicking anywhere else -> start a box
        (3) clicking a second time to finish a box -> create the segment
        """
        pos = evt.scenePos()

        # if any box is selected, deselect (wherever clicked)
        wasSelected = self.box1id
        if self.box1id > -1:
            self.deselectSegment(self.box1id)

        # if clicked inside scene:
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)

            # if this is the second click and not a box, close the segment
            if self.started:
                # can't finish boxes in ampl plot
                if self.config["specMouseAction"] > 1:
                    if self.startedInAmpl:
                        # started in ampl and finish in ampl,
                        # so continue as usual to draw a segment
                        pass
                    else:
                        # started in spec so ignore this bullshit
                        return

                # remove the drawing box:
                self.p_spec.removeItem(self.vLine_s)
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # disconnect GrowBox listeners, leave the position listener
                self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)

                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.ShiftModifier:
                    self.addSegment(
                        self.start_ampl_loc,
                        max(mousePoint.x(), 0.0),
                        species=copy.deepcopy(self.lastSpecies),
                    )
                    self.refreshFileColor()
                elif modifiers == Qt.ControlModifier:
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(), 0.0))
                    # Context menu
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(
                        QPoint(int(evt.screenPos().x()), int(evt.screenPos().y()))
                    )
                else:
                    self.addSegment(self.start_ampl_loc, max(mousePoint.x(), 0.0))
                    # Context menu
                    self.fillBirdList()
                    self.menuBirdList.popup(
                        QPoint(int(evt.screenPos().x()), int(evt.screenPos().y()))
                    )
                self.p_ampl.setFocus()

                # the new segment is now selected and can be played
                self.selectSegment(len(self.segments) - 1)
                self.started = not (self.started)
                self.startedInAmpl = False

                # reset cursor to not drawing (or leave as drawing if LMB draws)
                if self.MouseDrawingButton == Qt.RightButton:
                    self.p_ampl.unsetCursor()
                    self.specPlot.unsetCursor()
            # if this is the first click:
            else:
                # if this is right click (drawing mode):
                # (or whatever you want)
                if evt.button() == self.MouseDrawingButton:
                    if self.config["readOnly"]:
                        return
                    # this would prevent starting boxes in ampl plot
                    # if self.config['specMouseAction']>1:
                    #    return

                    nonebrush = self.ColourNone
                    self.start_ampl_loc = mousePoint.x()

                    # spectrogram plot bar and mouse followers:
                    self.vLine_s = pg.InfiniteLine(
                        angle=90, movable=False, pen={"color": "r", "width": 3}
                    )
                    self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                    self.vLine_s.setPos(self.convertAmpltoSpec(self.start_ampl_loc))

                    self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                    self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                    self.drawingBox_spec.setRegion(
                        [
                            self.convertAmpltoSpec(self.start_ampl_loc),
                            self.convertAmpltoSpec(self.start_ampl_loc),
                        ]
                    )
                    self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)

                    # amplitude plot bar and mouse followers:
                    self.vLine_a = pg.InfiniteLine(
                        angle=90, movable=False, pen={"color": "r", "width": 3}
                    )
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion(
                        [self.start_ampl_loc, self.start_ampl_loc]
                    )
                    self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    self.started = not (self.started)
                    self.startedInAmpl = True

                    # Force cursor to drawing
                    self.p_ampl.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                    self.specPlot.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # Note: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    min_distance = float("inf")
                    for count in range(len(self.listRectanglesa1)):
                        rect = self.listRectanglesa1[count]
                        if rect is not None:
                            x1, x2 = rect.getRegion()
                            if x1 <= mousePoint.x() <= x2:
                                center_x = (x1 + x2) / 2
                                distance = abs(mousePoint.x() - center_x)
                                if distance < min_distance:
                                    min_distance = distance
                                    box1id = count

                    # User clicked in a segment:
                    if box1id > -1:
                        # select the segment:
                        self.selectSegment(box1id)
                        # is it the first click on this segment?
                        if wasSelected == box1id:
                            # popup dialog
                            modifiers = QApplication.keyboardModifiers()
                            if modifiers == Qt.ControlModifier:
                                self.fillBirdList(unsure=True)
                            else:
                                self.fillBirdList()
                            self.menuBirdList.popup(
                                QPoint(
                                    int(evt.screenPos().x()), int(evt.screenPos().y())
                                )
                            )

    def mouseClicked_spec(self, evt):
        """Listener for if the user clicks on the spectrogram plot.
        See the amplitude version (mouseClicked_ampl()) for details. Although much of the code is a repeat,
        it is separated for clarity.
        """
        pos = evt.scenePos()

        # if any box is selected, deselect (wherever clicked)
        wasSelected = self.box1id
        if self.box1id > -1:
            self.deselectSegment(self.box1id)

        # when drawing boxes near scene borders, it's easy to release mouse outside scene,
        # and all the dragging gets messed up then. We map such cases to closest
        # scene positions here:
        if self.started and self.config["specMouseAction"] == 3:
            bounds = self.p_spec.sceneBoundingRect()
            if not bounds.contains(pos):
                newX = min(bounds.right(), max(bounds.left(), pos.x()))
                newY = min(bounds.bottom(), max(bounds.top(), pos.y()))
                pos.setX(newX)
                pos.setY(newY)

        # if clicked inside scene:
        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)

            # if this is the second click, close the segment/box
            # note: can finish segment with either left or right click
            if self.started:
                if self.config["specMouseAction"] > 1 and self.startedInAmpl:
                    # started in ampl, and spec is used for boxes, so can't continue here
                    return

                # remove the drawing box:
                if not self.config["specMouseAction"] > 1:
                    self.p_spec.removeItem(self.vLine_s)
                    self.p_ampl.scene().sigMouseMoved.disconnect()
                self.p_ampl.removeItem(self.vLine_a)
                self.p_ampl.removeItem(self.drawingBox_ampl)
                self.p_spec.removeItem(self.drawingBox_spec)
                # disconnect GrowBox listeners, leave the position listener
                self.p_spec.scene().sigMouseMoved.disconnect()
                if self.showPointerDetails.isChecked():
                    self.p_spec.scene().sigMouseMoved.connect(self.mouseMoved)
                # reset the trackers
                self.started = not (self.started)
                self.startedInAmpl = False

                # reset cursor to not drawing (or leave as drawing if LMB draws)
                if self.MouseDrawingButton == Qt.RightButton:
                    self.p_ampl.unsetCursor()
                    self.specPlot.unsetCursor()

                # Pass either default y coords or box limits:
                x1 = self.start_ampl_loc
                x2 = self.convertSpectoAmpl(max(mousePoint.x(), 0.0))
                # Could add this check if right edge seems dangerous:
                # endx = min(x2, np.shape(self.sg)[0]+1)
                if self.config["specMouseAction"] > 1:
                    y1 = self.start_spec_y
                    y2 = mousePoint.y()
                    miny = self.convertFreqtoY(self.sp.minFreqShow)
                    maxy = self.convertFreqtoY(self.sp.maxFreqShow)
                    y1 = min(max(miny, y1), maxy)
                    y2 = min(max(miny, y2), maxy)

                    # When dragging, can sometimes make boxes by mistake, which is annoying.
                    # To avoid, check that the box isn't too small
                    # Instead of creating a box, move the playback position bar.
                    if np.abs((x2 - x1) * (y2 - y1)) < self.minboxsize:
                        self.bar.setValue(self.convertAmpltoSpec(x1))
                        self.barMoved()
                        return

                    y1 = int(self.convertYtoFreq(y1))
                    y2 = int(self.convertYtoFreq(y2))
                else:
                    y1 = 0
                    y2 = 0

                # If the user has pressed shift, copy the last species and don't use the context menu
                # If they pressed Control, add ? to the names
                # note: Ctrl+Shift combo doesn't have a Qt modifier and is ignored.
                modifiers = QApplication.keyboardModifiers()
                if modifiers == Qt.ShiftModifier:
                    self.addSegment(
                        x1, x2, y1, y2, species=copy.deepcopy(self.lastSpecies)
                    )
                    self.refreshFileColor()
                elif modifiers == Qt.ControlModifier:
                    self.addSegment(x1, x2, y1, y2)
                    # Context menu
                    self.fillBirdList(unsure=True)
                    self.menuBirdList.popup(
                        QPoint(int(evt.screenPos().x()), int(evt.screenPos().y()))
                    )
                else:
                    self.addSegment(x1, x2, y1, y2)
                    # Context menu
                    self.fillBirdList()
                    self.menuBirdList.popup(
                        QPoint(int(evt.screenPos().x()), int(evt.screenPos().y()))
                    )
                self.p_spec.setFocus()

                # select the new segment/box
                self.selectSegment(len(self.segments) - 1)

            # if this is the first click:
            else:
                # if this is right click (drawing mode):
                if evt.button() == self.MouseDrawingButton:
                    if self.config["readOnly"]:
                        return
                    nonebrush = self.ColourNone
                    self.start_ampl_loc = self.convertSpectoAmpl(mousePoint.x())
                    self.start_spec_y = mousePoint.y()

                    # start a new box:
                    if self.config["specMouseAction"] > 1:
                        # spectrogram mouse follower box:
                        startpointS = QPointF(mousePoint.x(), mousePoint.y())
                        endpointS = QPointF(mousePoint.x(), mousePoint.y())

                        self.drawingBox_spec = SupportClasses_GUI.ShadedRectROI(
                            startpointS, endpointS - startpointS, invertible=True
                        )
                        self.drawingBox_spec.setBrush(nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                    # start a new segment:
                    else:
                        # spectrogram bar and mouse follower:
                        self.vLine_s = pg.InfiniteLine(
                            angle=90, movable=False, pen={"color": "r", "width": 3}
                        )
                        self.p_spec.addItem(self.vLine_s, ignoreBounds=True)
                        self.vLine_s.setPos(mousePoint.x())

                        self.drawingBox_spec = pg.LinearRegionItem(brush=nonebrush)
                        self.p_spec.addItem(self.drawingBox_spec, ignoreBounds=True)
                        self.drawingBox_spec.setRegion([mousePoint.x(), mousePoint.x()])
                        self.p_spec.scene().sigMouseMoved.connect(self.GrowBox_spec)
                        # note - only in segment mode react to movement over ampl plot:
                        self.p_ampl.scene().sigMouseMoved.connect(self.GrowBox_ampl)

                    # for box and segment - amplitude plot bar:
                    self.vLine_a = pg.InfiniteLine(
                        angle=90, movable=False, pen={"color": "r", "width": 3}
                    )
                    self.p_ampl.addItem(self.vLine_a, ignoreBounds=True)
                    self.vLine_a.setPos(self.start_ampl_loc)

                    self.drawingBox_ampl = pg.LinearRegionItem(brush=nonebrush)
                    self.p_ampl.addItem(self.drawingBox_ampl, ignoreBounds=True)
                    self.drawingBox_ampl.setRegion(
                        [self.start_ampl_loc, self.start_ampl_loc]
                    )

                    self.started = not (self.started)
                    self.startedInAmpl = False

                    # Force cursor to drawing
                    self.p_ampl.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                    self.specPlot.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                # if this is left click (selection mode):
                else:
                    # Check if the user has clicked in a box
                    # Note: Returns the first one it finds, i.e. the newest
                    box1id = -1
                    min_distance = float("inf")
                    for count in range(len(self.listRectanglesa2)):
                        rect = self.listRectanglesa2[count]
                        if rect is None:
                            continue

                        if type(rect) is self.ROItype:
                            x1 = rect.pos().x()
                            y1 = rect.pos().y()
                            x2 = x1 + rect.size().x()
                            y2 = y1 + rect.size().y()

                            if (
                                x1 <= mousePoint.x() <= x2
                                and y1 <= mousePoint.y() <= y2
                            ):
                                center_x = (x1 + x2) / 2
                                center_y = (y1 + y2) / 2
                                distance = math.hypot(
                                    mousePoint.x() - center_x, mousePoint.y() - center_y
                                )
                                if distance < min_distance:
                                    min_distance = distance
                                    box1id = count
                        else:
                            x1, x2 = rect.getRegion()
                            if x1 <= mousePoint.x() <= x2:
                                center_x = (x1 + x2) / 2
                                distance = abs(mousePoint.x() - center_x)
                                if distance < min_distance:
                                    min_distance = distance
                                    box1id = count

                    # User clicked in a segment:
                    if box1id > -1:
                        # select the segment:
                        self.selectSegment(box1id)
                        # if this segment is clicked again, pop up bird menu:
                        if wasSelected == box1id:
                            modifiers = QApplication.keyboardModifiers()
                            if modifiers == Qt.ControlModifier:
                                self.fillBirdList(unsure=True)
                            else:
                                self.fillBirdList()
                            self.menuBirdList.popup(
                                QPoint(
                                    int(evt.screenPos().x()), int(evt.screenPos().y())
                                )
                            )

    def GrowBox_ampl(self, pos):
        """Listener for when a segment is being made in the amplitude plot.
        Makes the blue box that follows the mouse change size."""
        if self.p_ampl.sceneBoundingRect().contains(pos):
            mousePoint = self.p_ampl.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion([self.start_ampl_loc, mousePoint.x()])
            self.drawingBox_spec.setRegion(
                [
                    self.convertAmpltoSpec(self.start_ampl_loc),
                    self.convertAmpltoSpec(mousePoint.x()),
                ]
            )

    def GrowBox_spec(self, pos):
        """Listener for when a segment is being made in the spectrogram plot.
        Makes the blue box that follows the mouse change size."""
        # When dragging spectrogram boxes near scene edges, we have special rules
        # to keep tracking the potential box
        if self.config["specMouseAction"] == 3:
            bounds = self.p_spec.sceneBoundingRect()
            if not bounds.contains(pos):
                newX = min(bounds.right(), max(bounds.left(), pos.x()))
                newY = min(bounds.bottom(), max(bounds.top(), pos.y()))
                pos.setX(newX)
                pos.setY(newY)

        if self.p_spec.sceneBoundingRect().contains(pos):
            mousePoint = self.p_spec.mapSceneToView(pos)
            self.drawingBox_ampl.setRegion(
                [self.start_ampl_loc, self.convertSpectoAmpl(mousePoint.x())]
            )
            if self.config["specMouseAction"] > 1 and not self.startedInAmpl:
                # making a box
                posY = mousePoint.y() - self.start_spec_y
                self.drawingBox_spec.setSize(
                    [mousePoint.x() - self.convertAmpltoSpec(self.start_ampl_loc), posY]
                )
            else:
                # making a segment
                self.drawingBox_spec.setRegion(
                    [self.convertAmpltoSpec(self.start_ampl_loc), mousePoint.x()]
                )

    def toggleLabelType(self):
        """Toggles between species-calltype level displays.
        Needs to swap the context menu, and the segment label text.
        """
        self.menuBirdList.triggered.disconnect()
        if self.viewCallType:
            self.viewCallType = False
            self.menuBirdList.triggered.connect(self.birdSelectedMenu)
            self.toggleLabelTypeBtn.setIcon(QIcon("img/splarge-ct.png"))
            # not sure if we need to re-set the size after every icon change
            # self.toggleLabelTypeBtn.setIconSize(QtCore.QSize(35, 20))
        else:
            self.viewCallType = True
            self.menuBirdList.triggered.connect(self.callSelectedMenu)
            self.toggleLabelTypeBtn.setIcon(QIcon("img/sp-ctlarge.png"))

        for seg in range(len(self.listLabels)):
            if self.listLabels[seg] is not None:
                self.updateText(seg)

    def birdSelectedList(self, index):
        """If the user clicks in the full bird list, update the text, and copy the species into the short list"""
        birdname = self.fullbirdlist.view().currentIndex().parent().data(0)
        if birdname is None:
            birdname = self.fullbirdlist.currentText()
        else:
            # two-level name
            if self.fullbirdlist.currentText().endswith("?"):
                birdname = birdname + " (" + self.fullbirdlist.currentText()[:-1] + ")?"
            else:
                birdname = birdname + " (" + self.fullbirdlist.currentText() + ")"
        self.birdSelectedMenu(birdname)
        if not self.multipleBirds:
            self.menuBirdList.hide()

    def birdSelectedMenu(self, birditem):
        """Collects the label for a bird from the context menu and processes it.
        Has to update the overview segments in case their colour should change.
        Also handles getting the name through a message box if necessary.
        """
        if type(birditem) is not str:
            birdname = birditem.text()
        else:
            birdname = birditem
        if birdname is None or birdname == "":
            return

        # special dialog for manual name entry
        if birdname == "Other":
            # Ask the user for the new name, and save it
            birdname, ok = QInputDialog.getText(
                self, "Bird name", "Enter the bird name as genus (species)"
            )
            if not ok:
                return

            birdname = str(birdname).title()
            # splits "A (B)", with B optional, into groups A and B
            match = re.fullmatch(r"(.*?)(?: \((.*)\))?", birdname)
            if not match:
                print(
                    "ERROR: provided name %s does not match format requirements"
                    % birdname
                )
                return

            if birdname.lower() == "don't know" or birdname.lower() == "other":
                print("ERROR: provided name %s is reserved, cannot create" % birdname)
                return

            if "?" in birdname:
                print("ERROR: provided name %s contains reserved symbol '?'" % birdname)
                return

            if len(birdname) == 0 or len(birdname) > 150:
                print("ERROR: provided name appears to be too short or too long")
                return

            twolevelname = ">".join(match.groups(default=""))
            if birdname in self.longBirdList or twolevelname in self.longBirdList:
                # bird is already listed
                print(
                    "Warning: not adding species %s as it is already present" % birdname
                )
                return

            # maybe the genus is already listed?
            index = self.model.findItems(match.group(1), Qt.MatchFixedString)
            if len(index) == 0:
                # Genus isn't in list
                item = QStandardItem(match.group(1))
                item.setSelectable(True)
                self.model.appendRow(item)
                # store as typed
                nametostore = birdname
            else:
                # Get the species item
                item = index[0]
                if match.group(2) is None:
                    print(
                        "ERROR: genus %s already exists, please provide species as well"
                        % match.group(1)
                    )
                    return
                # store in two-level format
                nametostore = twolevelname
                subitem = QStandardItem(match.group(2))
                item.setSelectable(False)
                item.appendRow(subitem)
                subitem.setSelectable(True)

            # update the main list:
            self.longBirdList.append(nametostore)
            self.longBirdList.remove("Unidentifiable")
            self.longBirdList = sorted(self.longBirdList, key=str.lower)
            self.longBirdList.append("Unidentifiable")
            self.ConfigLoader.blwrite(
                self.longBirdList, self.config["BirdListLong"], self.configdir
            )

        # parse birdname to certainty
        if birdname == "Don't Know":
            species = birdname
            certainty = 0
            self.prevBoxCol = self.ColourNone
        elif birdname[-1] == "?":
            species = birdname[:-1]
            certainty = 50
            self.prevBoxCol = self.ColourPossible
        else:
            species = birdname
            certainty = 100
            self.prevBoxCol = self.ColourNamed

        workingSeg = self.segments[self.box1id]
        self.refreshOverviewWith(workingSeg, delete=True)

        # toggle the actual label in the segment list
        if workingSeg.hasLabel(species, certainty):
            workingSeg.removeLabel(species, certainty)
        else:
            # in case the only label so far was Don't Know,
            # change it to the new bird (to not waste time unticking it)
            if workingSeg.keys == [("Don't Know", 0)]:
                workingSeg.addLabel(species, certainty, filter="M")
                workingSeg.removeLabel("Don't Know", 0)
                # also need to untick that context menu item manually
                for act in self.menuBirdList.actions() + self.menuBird2.actions():
                    if act.text() == "Don't Know":
                        act.setChecked(False)
            else:
                # in single-bird mode, just remove the current label:
                workingSeg.addLabel(species, certainty, filter="M")
                if not self.multipleBirds:
                    workingSeg.removeLabel(
                        workingSeg[4][0]["species"], workingSeg[4][0]["certainty"]
                    )

        # Put the selected bird name at the top of the list
        if self.config["ReorderList"]:
            # Either move the label to the top of the list, or delete the last
            if species in self.shortBirdList:
                self.shortBirdList.remove(species)
            else:
                del self.shortBirdList[-1]
            self.shortBirdList.insert(0, species)

        # refresh overview boxes after all updates:
        self.refreshOverviewWith(workingSeg)

        self.db.update_segment_species(
            workingSeg,
            self.SoundFileDir,
            self.listFiles.currentItem().text(),
            self.reviewer,
        )
        self.db.commit()
        # Store the species in case the user wants it for the next segment
        self.lastSpecies = [{"species": species, "certainty": 100, "filter": "M"}]
        self.updateText()
        self.updateColour()
        self.segInfo.setText(workingSeg.infoString())

        if not self.multipleBirds:
            # select the bird and close
            self.menuBirdList.hide()
        QApplication.processEvents()

    def callSelectedMenu(self, ctitem):
        """Simplified version of the above for dealing with calltype selection
        from the popup context menu."""
        if ctitem is None or ctitem == "":
            return

        spmenu = ctitem.parentWidget().title()
        if type(ctitem) is not str:
            ctitem = ctitem.text()

        if ctitem == "Add calltype":
            ctitem, ok = QInputDialog.getText(self, "Calltype", "Enter the calltype")
            if not ok:
                return

            ctitem = str(ctitem).title()

            if ctitem.lower() == "don't know" or ctitem.lower() == "other":
                print(
                    "ERROR: provided calltype name %s is reserved, cannot create"
                    % ctitem
                )
                return

            if "?" in ctitem:
                print(
                    "ERROR: provided calltype name %s contains reserved symbol '?'"
                    % ctitem
                )
                return

            if len(ctitem) == 0 or len(ctitem) > 150:
                print(
                    "ERROR: provided calltype name appears to be too short or too long"
                )
                return

            if spmenu in self.CalltypeDicts:
                if ctitem in self.CalltypeDicts[spmenu]["calltypes"]:
                    print("ERROR: provided calltype already present")
                    return
                else:
                    self.CalltypeDicts[spmenu]["calltypes"].append(ctitem)
            else:
                self.CalltypeDicts[spmenu] = {"species": spmenu, "calltypes": [ctitem]}

            self.saveCalltypeDicts()

        workingSeg = self.segments[self.box1id]
        for lab in workingSeg[4]:
            if lab["species"] == spmenu:
                lab["calltype"] = ctitem
        self.lastSpecies = [
            {"species": spmenu, "certainty": 100, "filter": "M", "calltype": ctitem}
        ]
        self.updateText()
        self.segInfo.setText(workingSeg.infoString())
        self.db.update_segment_species(
            workingSeg,
            self.SoundFileDir,
            self.listFiles.currentItem().text(),
            self.reviewer,
        )
        self.db.commit()
        # self.menuBirdList.hide()

    def saveCalltypeDicts(self):
        for species in self.CalltypeDicts:
            with open(
                os.path.join(self.calltypesDir, "{}.txt".format(species)), "w"
            ) as out:
                out.write(json.dumps(self.CalltypeDicts[species]))

    def updateText(self, segID=None):
        """When the user sets or changes the name in a segment, update the text label.
        Only requires the segment ID, or defaults to the selected one, and
        will read the label from it."""
        if segID is None:
            segID = self.box1id
        seg = self.segments[segID]
        if not self.viewCallType:
            # produce text from list of dicts
            text = []
            if self.viewCertainty:
                for lab in seg[4]:
                    if lab["certainty"] == 100:
                        text.append(lab["species"])
                    else:
                        text.append(
                            "{}:{}%".format(lab["species"], int(lab["certainty"]))
                        )
                text = ",".join(text)
            else:
                for lab in seg[4]:
                    if lab["certainty"] == 50:
                        text.append(lab["species"] + "?")
                    else:
                        text.append(lab["species"])
                text = ",".join(text)
        else:
            text = []
            for lab in seg[4]:
                if "calltype" in lab:
                    text.append(lab["calltype"])
                else:
                    text.append("(Other)")
            text = ",".join(text)

        # update the label
        self.listLabels[segID].setText(text, "k")
        self.listLabels[segID].update()
        QApplication.processEvents()

    def updateColour(self, segID=None):
        """Updates the color of a segment (useful for reviewing segments, for example).
        Only requires the segment ID, or defaults to the selected one, and
        will determine the color from it.
        """
        if segID is None:
            segID = self.box1id
        cert = min([lab["certainty"] for lab in self.segments[segID][4]])

        if cert == 0:
            brush = self.ColourNone
        elif cert == 100:
            brush = self.ColourNamed
        else:
            brush = self.ColourPossible

        # if we're updating the currently selected segment,
        # we should just store the new color (it'll be used on deselecting)
        if self.box1id == segID:
            self.prevBoxCol = brush
        # otherwise actually redraw the segment/box:
        else:
            if self.listRectanglesa2[segID] is None:
                return

            col = QtGui.QColor(brush)
            col.setAlpha(100)
            self.listRectanglesa1[segID].setBrush(col)
            self.listRectanglesa2[segID].setBrush(col)
            col.setAlpha(180)
            self.listRectanglesa1[segID].setHoverBrush(fn.mkBrush(col))
            self.listRectanglesa2[segID].setHoverBrush(fn.mkBrush(col))

            if type(self.listRectanglesa2[segID]) == self.ROItype:
                self.listRectanglesa2[segID].transparent = False
                self.listRectanglesa2[segID].setPen(None)
                if self.config["transparentBoxes"]:
                    col.setAlpha(255)
                    self.listRectanglesa2[segID].transparent = True
                    self.listRectanglesa2[segID].setPen(col, width=1)
                    self.listRectanglesa2[segID].setBrush(None)
                    self.listRectanglesa2[segID].setHoverBrush(None)
                    col.setAlpha(100)
            self.listRectanglesa1[segID].update()
            self.listRectanglesa2[segID].update()
        QApplication.processEvents()

    def setColourMap(self, cmap):
        """Listener for the menu item that chooses a colour map.
        Loads them from the file as appropriate and sets the lookup table.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()

        self.config["cmap"] = cmap
        if cmap == "Custom":
            lut = colourMaps.getLookupTable(
                self.config["ccmap"][0], self.config["ccmap"][1]
            )
        else:
            lut = colourMaps.getLookupTableFromColourMap(self.config["cmap"])

        self.specPlot.setLookupTable(lut)
        self.overviewImage.setLookupTable(lut)

    def invertColourMap(self):
        """Listener for the menu item that converts the colour map"""
        # self.config['invertColourMap'] = not self.config['invertColourMap']
        self.config["invertColourMap"] = self.invertcm.isChecked()
        self.setColourLevels()

    def setSpectrogram(self):
        """Normalizes the raw spectrogram on self.sp (ndarray), puts it on self,
        and precalculates some cached properties from it.
        Does NOT update graphics - only internal objects.
        """
        self.sg = self.sp.normalisedSpec(self.sgNormMode)
        self.sgMinimum = np.min(self.sg)
        self.sgMaximum = np.max(self.sg)

    def setColourLevels(self, brightness=None, contrast=None):
        """Listener for the brightness and contrast sliders being changed. Also called when spectrograms are loaded, etc.
        Translates the brightness and contrast values into appropriate image levels.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()
        if brightness is None:
            brightness = self.specControls.brightSlider.value()
        if contrast is None:
            contrast = self.specControls.contrSlider.value()

        if self.config["invertColourMap"]:
            self.config["brightness"] = brightness
        else:
            self.config["brightness"] = 100 - brightness
        self.config["contrast"] = contrast
        self.saveConfig = True

        colRange = colourMaps.getColourRange(
            self.sgMinimum,
            self.sgMaximum,
            self.config["brightness"],
            self.config["contrast"],
            self.config["invertColourMap"],
        )

        self.overviewImage.setLevels(colRange)
        self.specPlot.setLevels(colRange)

    def moveLeft(self):
        """When the left button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap"""
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = max(0, minX - (maxX - minX) * 0.9)
        self.overviewImageRegion.setRegion([newminX, newminX + maxX - minX])

    def moveRight(self):
        """When the right button is pressed (next to the overview plot), move everything along
        Allows a 10% overlap"""
        minX, maxX = self.overviewImageRegion.getRegion()
        newminX = min(np.shape(self.sg)[0] - (maxX - minX), minX + (maxX - minX) * 0.9)
        self.overviewImageRegion.setRegion([newminX, newminX + maxX - minX])

    def prepare5minMove(self):
        self.saveSegments()
        # selr.resetStorageArrays()
        self.loadFile()

    def movePrev5mins(self):
        """When the button to move to the next 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection -= 1
        self.next5mins.setEnabled(True)
        self.moveNext5MinsKey.setEnabled(True)
        if self.currentFileSection <= 0:
            self.prev5mins.setEnabled(False)
            self.movePrev5MinsKey.setEnabled(False)
        self.prepare5minMove()

    def moveNext5mins(self):
        """When the button to move to the previous 5 minutes is pressed, enable that.
        Have to check if the buttons should be disabled or not,
        save the segments and reset the arrays, then call loadFile.
        """
        self.currentFileSection += 1
        self.prev5mins.setEnabled(True)
        self.movePrev5MinsKey.setEnabled(True)
        if self.currentFileSection >= self.nFileSections - 1:
            self.next5mins.setEnabled(False)
            self.moveNext5MinsKey.setEnabled(False)
        self.prepare5minMove()

    def moveTo5mins(self, pagenum=None):
        """Jumps to the requested 5 min page.
        pagenum can be specified if this is called manually
          Otherwise (None) it will be read from the page selector.
        """
        if pagenum is None:
            pagenum = self.placeInFileSelector.value()
        self.placeInFileSelector.findChild(QLineEdit).deselect()
        self.placeInFileSelector.clearFocus()
        if self.currentFileSection == pagenum - 1:
            # no jump needed
            return
        self.currentFileSection = pagenum - 1
        if self.currentFileSection >= self.nFileSections - 1:
            self.next5mins.setEnabled(False)
            self.moveNext5MinsKey.setEnabled(False)
        else:
            self.next5mins.setEnabled(True)
            self.moveNext5MinsKey.setEnabled(True)

        if self.currentFileSection <= 0:
            self.prev5mins.setEnabled(False)
            self.movePrev5MinsKey.setEnabled(False)
        else:
            self.prev5mins.setEnabled(True)
            self.movePrev5MinsKey.setEnabled(True)
        self.prepare5minMove()

    def scroll(self):
        """When the slider at the bottom of the screen is moved, move everything along."""
        newminX = self.scrollSlider.value()
        if not self.updateRequestedByOverview:
            minX, maxX = self.overviewImageRegion.getRegion()
            self.overviewImageRegion.setRegion([newminX, newminX + maxX - minX])

    def changeWidth(self, value):
        """Listener for the spinbox that decides the width of the main window.
        It updates the top figure plots as the window width is changed.
        Slightly annoyingly, it gets called when the value gets reset, hence the first line.
        """
        if not hasattr(self, "overviewImageRegion"):
            return
        self.windowSize = value

        if not self.updateRequestedByOverview:
            # Redraw the highlight in the overview figure appropriately
            minX, maxX = self.overviewImageRegion.getRegion()
            newmaxX = self.convertAmpltoSpec(value) + minX
            self.overviewImageRegion.setRegion([minX, newmaxX])

        self.scrollSlider.setMaximum(
            int(np.shape(self.sg)[0] - self.convertAmpltoSpec(self.widthWindow.value()))
        )

        # Decide whether or not to show milliseconds
        if value > 3:
            self.timeaxis.setShowMS(False)
        else:
            self.timeaxis.setShowMS(True)

    def annotJumper(self, maxcert):
        """Scrolls to next annotation of no more than maxcert certainty."""
        # (this is just a manual pg.BusyCursor)
        QApplication.setOverrideCursor(QtGui.QCursor(Qt.WaitCursor))
        # Identify the "current" annotation: selected or whatever is on screen
        if self.box1id > -1:
            currx = self.segments[self.box1id][0]
            self.deselectSegment(self.box1id)
        else:
            minX, maxX = self.overviewImageRegion.getRegion()
            currx = self.convertSpectoAmpl(minX) + self.startRead
            # set currx to -1 to find segment starting at 0 if no segment is selected yet
            currx = -1 if currx == 0 else currx

        # Find next annotation:
        targetix = None
        for segix in range(len(self.segments)):
            seg = self.segments[segix]
            if seg[0] <= currx:
                continue
            # Note that the segments are not sorted by time,
            # hence some extra mess to find the next one:
            if targetix is not None and seg[0] >= self.segments[targetix][0]:
                continue
            for lab in seg[4]:
                if lab["certainty"] <= maxcert and (
                    lab["species"] == self.currentSpecies
                    or self.currentSpecies == "Species"
                ):
                    targetix = segix
        if targetix is None:
            QApplication.restoreOverrideCursor()
            print("No further annotation to jump to found")
            msg = SupportClasses_GUI.MessagePopup(
                "w", "No more annotations", "No further annotation to jump to found"
            )
            msg.exec_()
            return

        target = self.segments[targetix]
        if target[0] >= self.startRead + self.datalengthSec:
            pagenum = target[0] // (
                self.config["maxFileShow"] - self.config["fileOverlap"]
            )
            pagenum = int(pagenum + 1)
            if pagenum > self.nFileSections:
                print("Warning: annotation outside file bounds")
                QApplication.restoreOverrideCursor()
                msg = SupportClasses_GUI.MessagePopup(
                    "w",
                    "No more annotations",
                    "No further annotation to jump to found in this sound file",
                )
                msg.exec_()
                return
            self.moveTo5mins(pagenum)
        newminT = target[0] - self.startRead - self.windowSize / 2  # in s
        newminX = self.convertAmpltoSpec(newminT)  # in spec pixels
        newmaxX = self.convertAmpltoSpec(newminT + self.windowSize)
        # this will trigger update of the other views
        self.overviewImageRegion.setRegion([newminX, newmaxX])
        self.selectSegment(targetix)
        QApplication.restoreOverrideCursor()

    def annotJumperPrev(self, maxcert):
        """Scrolls to previous annotation of no more than maxcert certainty."""
        # (this is just a manual pg.BusyCursor)
        QApplication.setOverrideCursor(QtGui.QCursor(Qt.WaitCursor))
        # Identify the "current" annotation: selected or whatever is on screen
        if self.box1id > -1:
            currx = self.segments[self.box1id][1]
            self.deselectSegment(self.box1id)
        else:
            minX, maxX = self.overviewImageRegion.getRegion()
            currx = self.convertSpectoAmpl(maxX) + self.startRead

        # Find previous annotation:
        targetix = None
        for segix in range(len(self.segments)):
            seg = self.segments[segix]
            if seg[1] >= currx:
                continue
            # Note that the segments are not sorted by time,
            # hence some extra mess to find the next one:
            if targetix is not None and seg[1] <= self.segments[targetix][1]:
                continue
            for lab in seg[4]:
                if lab["certainty"] <= maxcert and (
                    lab["species"] == self.currentSpecies
                    or self.currentSpecies == "Species"
                ):
                    targetix = segix
        if targetix is None:
            QApplication.restoreOverrideCursor()
            print("No further annotation to jump to found")
            msg = SupportClasses_GUI.MessagePopup(
                "w", "No more annotations", "No further annotation to jump to found"
            )
            msg.exec_()
            return

        target = self.segments[targetix]

        if target[0] < self.startRead:
            pagenum, _ = divmod(target[0], self.config["maxFileShow"])
            pagenum = target[0] // (
                self.config["maxFileShow"] - self.config["fileOverlap"]
            )
            pagenum = int(pagenum + 1)
            if pagenum < 0:
                print("Warning: annotation outside file bounds")
                QApplication.restoreOverrideCursor()
                msg = SupportClasses_GUI.MessagePopup(
                    "w",
                    "No more annotations",
                    "No further annotation to jump to found in this sound file",
                )
                msg.exec_()
                return
            self.moveTo5mins(pagenum)
        newminT = target[0] - self.startRead - self.windowSize / 2  # in s
        newminX = self.convertAmpltoSpec(newminT)  # in spec pixels
        newmaxX = self.convertAmpltoSpec(newminT + self.windowSize)
        # this will trigger update of the other views
        self.overviewImageRegion.setRegion([newminX, newmaxX])
        self.selectSegment(targetix)
        QApplication.restoreOverrideCursor()

    # ===============
    # Generate the various dialogs that match the menu items

    def showSpectrogramDialog(self):
        """Create spectrogram dialog when the button is pressed."""
        if not hasattr(self, "spectrogramDialog"):
            self.spectrogramDialog = Dialogs.Spectrogram(
                self.config["window_width"],
                self.config["incr"],
                self.sp.minFreq,
                self.sp.maxFreq,
                self.sp.minFreqShow,
                self.sp.maxFreqShow,
                self.config["window"],
                self.sgType,
                self.sgNormMode,
                self.sgScale,
                int(str(self.nfilters)),
            )
            self.spectrogramDialog.activate.clicked.connect(self.spectrogram)
        # first save the annotations
        self.saveSegments()
        self.spectrogramDialog.show()
        self.spectrogramDialog.activateWindow()

    def spectrogram(self):
        """Listener for the spectrogram dialog.
        Has to do quite a bit of work to make sure segments are in the correct place, etc.
        """
        [
            self.windowType,
            self.sgType,
            self.sgNormMode,
            self.sgMeanNormalise,
            self.sgEqualLoudness,
            window_width,
            incr,
            minFreq,
            maxFreq,
            sgScale,
            self.nfilters,
        ] = self.spectrogramDialog.getValues()
        if self.sgScale != sgScale:
            self.sgScale = sgScale
            changedY = True
        else:
            changedY = False

        if minFreq >= maxFreq:
            msg = SupportClasses_GUI.MessagePopup(
                "w", "Error", "Incorrect frequency range"
            )
            msg.exec_()
            return
        with pg.BusyCursor():
            self.statusLeft.setText("Updating the spectrogram...")
            self.sp.setWidth(int(str(window_width)), int(str(incr)))
            _ = self.sp.spectrogram(
                window=str(self.windowType),
                sgType=str(self.sgType),
                sgScale=str(self.sgScale),
                nfilters=int(str(self.nfilters)),
                mean_normalise=self.sgMeanNormalise,
                equal_loudness=self.sgEqualLoudness,
                onesided=self.sgOneSided,
            )
            self.setSpectrogram()

            # If the size of the spectrogram has changed, need to update the positions of things
            if (
                int(str(incr)) != self.config["incr"]
                or int(str(window_width)) != self.config["window_width"]
            ):
                self.config["incr"] = int(str(incr))
                self.config["window_width"] = int(str(window_width))
                if hasattr(self, "seg"):
                    self.seg.setNewData(self.sp)

                self.loadFile(self.filename)
                # self.specPlot.setImage(self.sg)   # TODO: interface changes to adapt if window_len and incr changed! overview, main spec ect.

                # these two are usually set by redoFreqAxis, but that is called only later in this case
                self.spectrogramDialog.low.setValue(minFreq)
                self.spectrogramDialog.high.setValue(maxFreq)

        self.redoFreqAxis(minFreq, maxFreq, changedY=changedY)
        self.setColourLevels()
        self.statusLeft.setText("Ready")

    def calculateStats(self):
        """Calculate and export summary statistics for the currently marked segments"""

        import Features

        # these are all segments in file
        print("segs", self.segments)

        csv = open(self.filename[:-4] + "_features.csv", "w")
        csv.write(
            "Start Time (sec),End Time (sec),Avg Power,Delta Power,Energy,Agg Entropy,Avg Entropy,Max Power,Max Freq\n"
        )

        for seg in self.segments:
            # Important because all manual mode functions should operate on the current page only:
            # skip segments that are not visible in this page
            if (
                seg[1] <= self.startRead
                or seg[0] >= self.startRead + self.datalengthSec
            ):
                continue

            # coordinates in seconds from current page start, bounded at page borders:
            starttime = max(0, seg[0] - self.startRead)
            endtime = min(seg[1] - self.startRead, self.datalengthSec)
            print(starttime, endtime)

            # piece of audio/waveform corresponding to this segment
            # (note: coordinates in wav samples)
            data = self.audiodata[
                int(starttime * self.sampleRate) : int(endtime * self.sampleRate)
            ]

            # piece of spectrogram corresponding to this segment
            startInSpecPixels = self.convertAmpltoSpec(starttime)
            endInSpecPixels = self.convertAmpltoSpec(endtime)
            print(startInSpecPixels, endInSpecPixels)
            # self.sg[startInSpecPixels:endInSpecPixels, ]

            # if needed, there's already a SignalProc instance self.sp with the full data on it,
            # so can also do something like:
            # self.sp.calculateMagicStatistic(starttime, endtime)

            # do something with this segment now...
            print("Calculating statistics on this segment...")

            # TODO: Hardcoded for now - add a dialog to read parameters?
            # TODO: Workout the units
            f = Features.Features(
                data=data, sampleRate=self.sampleRate, window_width=256, incr=128
            )
            (
                avgPower,
                deltaPower,
                energy,
                aggEntropy,
                avgEntropy,
                maxPower,
                maxFreq,
            ) = f.get_Raven_spectrogram_measurements(
                f1=int(self.convertFreqtoY(500)), f2=int(self.convertFreqtoY(8000))
            )
            # quartile1, quartile2, quartile3, f5, f95, interquartileRange = f.get_Raven_robust_measurements(f1=int(self.convertFreqtoY(500)), f2=int(self.convertFreqtoY(8000)))
            print(
                avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq
            )
            # print(quartile1, quartile2, quartile3, f5, f95, interquartileRange)
            # csv.write("%s\t%.4f\t%.4f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\t%.2f\n" % (self.filename, starttime, endtime, avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq, quartile1, quartile2, quartile3, f5, f95, interquartileRange))
            # csv.write("%s,%.4f,%.4f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f\n" % (self.filename, starttime, endtime, avgPower, deltaPower, energy, aggEntropy, avgEntropy, maxPower, maxFreq))
            csv.write(
                "%.4f,%.4f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f,%.2f\n"
                % (
                    starttime,
                    endtime,
                    avgPower,
                    deltaPower,
                    energy,
                    aggEntropy,
                    avgEntropy,
                    maxPower,
                    maxFreq,
                )
            )

        csv.close()

    def showDenoiseDialog(self):
        """Create the denoising dialog when the relevant button is pressed."""
        self.denoiseDialog = Dialogs.Denoise(
            DOC=self.DOC, minFreq=self.sp.minFreq, maxFreq=self.sp.maxFreq
        )
        self.denoiseDialog.show()
        self.denoiseDialog.activateWindow()
        self.denoiseDialog.activate.clicked.connect(self.denoise)
        self.denoiseDialog.undo.clicked.connect(self.denoise_undo)
        self.denoiseDialog.save.clicked.connect(self.denoise_save)

    def backup(self):
        """Enables denoising to be undone."""
        if hasattr(self, "audiodata_backup"):
            if self.audiodata_backup is not None:
                audiodata_backup_new = np.empty(
                    (
                        np.shape(self.audiodata_backup)[0],
                        np.shape(self.audiodata_backup)[1] + 1,
                    )
                )
                audiodata_backup_new[:, :-1] = np.copy(self.audiodata_backup)
                audiodata_backup_new[:, -1] = np.copy(self.audiodata)
                self.audiodata_backup = audiodata_backup_new
            else:
                self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
                self.audiodata_backup[:, 0] = np.copy(self.audiodata)
        else:
            self.audiodata_backup = np.empty((np.shape(self.audiodata)[0], 1))
            self.audiodata_backup[:, 0] = np.copy(self.audiodata)
        self.showFreq_backup = [self.sp.minFreqShow, self.sp.maxFreqShow]

    def denoiseSeg(self):
        """Listener for quickDenoise control button.
        Extracts a segment from DATA between START and STOP (in ms),
        denoises that segment, concats with rest of original DATA,
        and updates the original DATA.
        """
        if self.box1id > -1:
            start, stop = self.listRectanglesa1[self.box1id].getRegion()
            start = int(start * self.sampleRate)
            stop = int(stop * self.sampleRate)
        else:
            print("Can't play, no segment selected")
            return

        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()

        # Since there is no dialog menu, settings are preset constants here:
        noiseest = "ols"  # or qr, or const
        thrType = "soft"
        depth = 6  # can also use 0 to autoset
        wavelet = "dmey2"
        aaRec = False  # True if nicer spectrogram is needed - but it's not very clean either way
        aaWP = False
        thr = 2.0  # this one is difficult to set universally...

        self.statusLeft.setText("Denoising...")
        with pg.BusyCursor():
            opstartingtime = time.time()
            print(
                "Denoising requested at "
                + time.strftime("%H:%M:%S", time.gmtime(opstartingtime))
            )

            # extract the piece of audiodata under current segment
            denoised = self.audiodata[start:stop]

            WF = WaveletFunctions.WaveletFunctions(
                data=denoised,
                wavelet=wavelet,
                maxLevel=self.config["maxSearchDepth"],
                samplerate=self.sampleRate,
            )
            denoised = WF.waveletDenoise(
                thrType,
                thr,
                depth,
                aaRec=aaRec,
                aaWP=aaWP,
                noiseest=noiseest,
                costfn="fixed",
            )

            # bandpass to selected zones, if it's a box
            # TODO this could be done faster: pass to waveletDenoise and
            # do not reconstruct from nodes outside the specified band
            if self.segments[self.box1id][3] > 0:
                bottom = max(0.1, self.sp.minFreq, self.segments[self.box1id][2])
                top = min(self.segments[self.box1id][3], self.sp.maxFreq - 0.1)
                print("Extracting samples between %d-%d Hz" % (bottom, top))
                denoised = self.sp.bandpassFilter(
                    denoised, sampleRate=None, start=bottom, end=top
                )

            print(
                "Denoising calculations completed in %.4f seconds"
                % (time.time() - opstartingtime)
            )

            # update full audiodata
            self.sp.data[start:stop] = denoised
            self.audiodata[start:stop] = denoised

            # recalculate spectrogram
            _ = self.sp.spectrogram(
                window=str(self.windowType),
                sgType=str(self.sgType),
                sgScale=str(self.sgScale),
                nfilters=int(str(self.nfilters)),
                mean_normalise=self.sgMeanNormalise,
                equal_loudness=self.sgEqualLoudness,
                onesided=self.sgOneSided,
            )
            self.setSpectrogram()

            # Update the ampl image
            self.amplPlot.setData(
                np.linspace(
                    0.0,
                    self.datalength / self.sampleRate,
                    num=self.datalength,
                    endpoint=True,
                ),
                self.audiodata,
            )

            # Update the spec & overview images.
            # Does not reset to start if the freqs aren't changed
            self.redoFreqAxis(self.sp.minFreqShow, self.sp.maxFreqShow, store=False)

            if hasattr(self, "spectrogramDialog"):
                self.spectrogramDialog.setValues(
                    self.sp.minFreq,
                    self.sp.maxFreq,
                    self.sp.minFreqShow,
                    self.sp.maxFreqShow,
                )

            self.setColourLevels()

            print(
                "Denoising completed in %s seconds"
                % round(time.time() - opstartingtime, 4)
            )
        self.statusLeft.setText("Ready")

    def denoise(self):
        """Listener for the denoising dialog.
        Calls the denoiser and then plots the updated data.
        """
        with pg.BusyCursor():
            opstartingtime = time.time()
            print(
                "Denoising requested at "
                + time.strftime("%H:%M:%S", time.gmtime(opstartingtime))
            )
            self.statusLeft.setText("Denoising...")
            # Note: dialog returns all possible parameters
            if not self.DOC:
                [
                    alg,
                    depth,
                    thrType,
                    thr,
                    wavelet,
                    start,
                    end,
                    width,
                    aaRec,
                    aaWP,
                    noiseest,
                ] = self.denoiseDialog.getValues()
            else:
                wavelet = "dmey2"
                [alg, start, end, width] = self.denoiseDialog.getValues()
            self.backup()

            if str(alg) == "Wavelets":
                # here we override default 0-Fs/2 returns
                start = self.sp.minFreqShow
                end = self.sp.maxFreqShow
                self.waveletDenoiser = WaveletFunctions.WaveletFunctions(
                    data=self.audiodata,
                    wavelet=wavelet,
                    maxLevel=self.config["maxSearchDepth"],
                    samplerate=self.sampleRate,
                )
                if not self.DOC:
                    # pass dialog settings
                    # TODO set costfn determines which leaves will be used, by default 'threshold' (universal threshold).
                    # fixed = use all leaves up to selected level. 'Entropy' is also tested and possible
                    self.sp.data = self.waveletDenoiser.waveletDenoise(
                        thrType,
                        float(str(thr)),
                        depth,
                        aaRec=aaRec,
                        aaWP=aaWP,
                        noiseest=noiseest,
                        costfn="fixed",
                    )
                else:
                    # go with defaults
                    self.sp.data = self.waveletDenoiser.waveletDenoise(
                        "soft",
                        3,
                        aaRec=True,
                        aaWP=False,
                        costfn="fixed",
                        noiseest="ols",
                    )

            else:
                # SignalProc will deal with denoising
                self.sp.denoise(alg, start=start, end=end, width=width)
            self.audiodata = self.sp.data

            print(
                "Denoising calculations completed in %.4f seconds"
                % (time.time() - opstartingtime)
            )

            _ = self.sp.spectrogram(
                window=str(self.windowType),
                sgType=str(self.sgType),
                sgScale=str(self.sgScale),
                nfilters=int(str(self.nfilters)),
                mean_normalise=self.sgMeanNormalise,
                equal_loudness=self.sgEqualLoudness,
                onesided=self.sgOneSided,
            )
            self.setSpectrogram()

            self.amplPlot.setData(
                np.linspace(
                    0.0,
                    self.datalength / self.sampleRate,
                    num=self.datalength,
                    endpoint=True,
                ),
                self.audiodata,
            )

            # Update the frequency axis
            self.redoFreqAxis(start, end, store=False)

            if hasattr(self, "spectrogramDialog"):
                self.spectrogramDialog.setValues(
                    self.sp.minFreq,
                    self.sp.maxFreq,
                    self.sp.minFreqShow,
                    self.sp.maxFreqShow,
                )

            self.setColourLevels()

            print(
                "Denoising completed in %s seconds"
                % round(time.time() - opstartingtime, 4)
            )
            self.statusLeft.setText("Ready")

    def denoise_undo(self):
        """Listener for undo button in denoising dialog."""
        print("Undoing", np.shape(self.audiodata_backup))
        if hasattr(self, "audiodata_backup"):
            if self.audiodata_backup is not None:
                if np.shape(self.audiodata_backup)[1] > 0:
                    self.audiodata = np.copy(self.audiodata_backup[:, -1])
                    self.audiodata_backup = self.audiodata_backup[:, :-1]
                    self.sp.data = self.audiodata

                    _ = self.sp.spectrogram(
                        window=str(self.windowType),
                        sgType=str(self.sgType),
                        sgScale=str(self.sgScale),
                        nfilters=int(str(self.nfilters)),
                        mean_normalise=self.sgMeanNormalise,
                        equal_loudness=self.sgEqualLoudness,
                        onesided=self.sgOneSided,
                    )
                    self.setSpectrogram()

                    self.amplPlot.setData(
                        np.linspace(
                            0.0, self.datalengthSec, num=self.datalength, endpoint=True
                        ),
                        self.audiodata,
                    )
                    if hasattr(self, "seg"):
                        self.seg.setNewData(self.sp)

                    if hasattr(self, "showFreq_backup"):
                        self.redoFreqAxis(
                            self.showFreq_backup[0], self.showFreq_backup[1]
                        )
                    else:
                        self.redoFreqAxis(self.sp.minFreq, self.sp.maxFreq)
                    self.setColourLevels()

    def denoise_save(self):
        """Listener for save button in denoising dialog.
        Adds _d to the filename and saves it as a new sound file.
        """
        filename = self.filename[:-4] + "_d" + self.filename[-4:]
        wavio.write(
            filename,
            self.audiodata.astype("int16"),
            self.sampleRate,
            scale="dtype-limits",
            sampwidth=2,
        )
        self.statusLeft.setText("Saved")
        msg = SupportClasses_GUI.MessagePopup(
            "d", "Saved", "Destination: " + "\n" + filename
        )
        msg.exec_()
        return

    def saveSelectedSound(self, slowdown=1):
        """Listener for 'Save selected (slow) sound' button.
        Chooses destination, file name, and exports.
        slowdown: can pass a factor >1 (<1) to save slower (faster) sound.
        """
        if self.box1id is None or self.box1id < 0:
            print("No box selected")
            msg = SupportClasses_GUI.MessagePopup(
                "w", "No segment", "No sound selected to save"
            )
            msg.exec_()
            return
        else:
            if type(self.listRectanglesa2[self.box1id]) == self.ROItype:
                x1 = self.listRectanglesa2[self.box1id].pos().x()
                x2 = x1 + self.listRectanglesa2[self.box1id].size().x()
                y1 = max(self.sp.minFreq, self.segments[self.box1id][2])
                y2 = min(self.segments[self.box1id][3], self.sp.maxFreq)
            else:
                x1, x2 = self.listRectanglesa2[self.box1id].getRegion()
                y1 = self.sp.minFreq
                y2 = self.sp.maxFreq
            x1 = math.floor(x1 * self.config["incr"])
            x2 = math.floor(x2 * self.config["incr"])
            filename, drop = QFileDialog.getSaveFileName(
                self, "Save File as", "", "*.wav"
            )
            if filename:
                # filedialog doesn't attach extension
                filename = str(filename)
                if not filename.endswith(".wav"):
                    filename = filename + ".wav"
                tosave = self.sp.bandpassFilter(
                    self.audiodata[int(x1) : int(x2)], start=y1, end=y2
                )
                wavio.write(
                    filename,
                    tosave.astype("int16"),
                    int(self.sampleRate / slowdown),
                    scale="dtype-limits",
                    sampwidth=2,
                )
            # update the file list box
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

    def redoFreqAxis(self, start, end, store=True, changedY=False):
        """This is the listener for the menu option to make the frequency axis tight (after bandpass filtering or just spectrogram changes)
        On the same go updates spectrogram and overview plots.
            store: boolean, indicates whether changes should be stored in the config
        """
        changedY = (
            start != self.sp.minFreqShow or end != self.sp.maxFreqShow or changedY
        )
        # Lots of updating can be avoided if the Y freqs aren't changing:
        if changedY:
            self.sp.minFreqShow = max(start, self.sp.minFreq)
            self.sp.maxFreqShow = min(end, self.sp.maxFreq)

            if store:
                self.config["minFreq"] = start
                self.config["maxFreq"] = end

        # draw a spectrogram of proper height:
        height = self.sampleRate // 2 / np.shape(self.sg)[1]
        pixelstart = int(self.sp.minFreqShow / height)
        pixelend = int(self.sp.maxFreqShow / height)

        self.overviewImage.setImage(self.sg[:, pixelstart:pixelend])
        self.overviewImageRegion.setBounds([0, len(self.sg)])
        self.specPlot.setImage(self.sg[:, pixelstart:pixelend])

        # if Y freqs changed, some segments may appear/be dropped:
        if changedY:
            # Remove everything and redraw it
            self.removeSegments(delete=False)
            self.drawOverview()
            self.drawfigMain(reusing=True)

            try:
                for r in self.segmentPlots:
                    self.p_spec.removeItem(r)
                self.segmentPlots = []
            except Exception:
                pass
            else:
                self.showFundamentalFreq()

            try:
                self.p_spec.removeItem(self.derivPlot)
            except Exception:
                pass
            # else:
            #     self.showSpectralDeriv()

            # if not self.DOC:
            #     try:
            #         self.p_spec.removeItem(self.formantPlot)
            #     except Exception:
            #         pass
            #     else:
            #         self.showFormants()

            try:
                self.p_spec.removeItem(self.energyPlot)
            except Exception:
                pass
            else:
                self.showMaxEnergy()

        QApplication.processEvents()

    def excel2Annotation(self):
        """Utility function dialog: Generate AviaNZ style annotations given the start-end of calls in excel format"""
        self.excel2AnnotationDialog = Dialogs.Excel2Annotation()
        self.excel2AnnotationDialog.show()
        self.excel2AnnotationDialog.activateWindow()
        self.excel2AnnotationDialog.btnGenerateAnnot.clicked.connect(
            self.genExcel2Annot
        )

    def backupAnnotations(self):
        """Utility function dialog: backup annotation files"""
        self.backupAnnotationDialog = Dialogs.BackupAnnotation()
        self.backupAnnotationDialog.show()
        self.backupAnnotationDialog.activateWindow()
        self.backupAnnotationDialog.btnCopyAnnot.clicked.connect(self.backupAnnotation)

    def genExcel2Annot(self):
        """Utility function: Generate AviaNZ style annotations given the start-end of calls in excel format"""

        values = self.excel2AnnotationDialog.getValues()
        if values:
            [excelfile, audiofile, species, colstart, colend, collow, colhigh] = values
        else:
            return

        try:
            # Read excel file
            book = openpyxl.load_workbook(excelfile)
            sheet = book.active
            starttime = sheet[colstart + "2" : colstart + str(sheet.max_row)]
            endtime = sheet[colend + "2" : colend + str(sheet.max_row)]
            flow = sheet[collow + "2" : collow + str(sheet.max_row)]
            fhigh = sheet[colhigh + "2" : colhigh + str(sheet.max_row)]

            _, duration, _, _ = wavio.readFmt(audiofile)

            annotation = []
            for i in range(len(starttime)):
                annotation.append(
                    [
                        float(starttime[i][0].value),
                        float(endtime[i][0].value),
                        float(flow[i][0].value),
                        float(fhigh[i][0].value),
                        [
                            {
                                "species": species,
                                "certainty": 100.0,
                                "filter": "M",
                                "calltype": species,
                            }
                        ],
                    ]
                )
            annotation.insert(0, {"Operator": "", "Reviewer": "", "Duration": duration})
            file = open(audiofile + ".data", "w")
            json.dump(annotation, file)
            file.close()
            self.excel2AnnotationDialog.txtSpecies.setText("")
            self.excel2AnnotationDialog.txtAudio.setText("")
            self.excel2AnnotationDialog.txtExcel.setText("")
            msg = SupportClasses_GUI.MessagePopup(
                "d",
                "Generated annotation",
                "Successfully saved the annotation file: " + "\n" + audiofile + ".data",
            )
            msg.exec_()
        except Exception as e:
            print("ERROR: Generating annotation failed with error:")
            print(e)
            return

    def backupAnnotation(self):
        """Utility function: Copy .data and corrections files while preserving directory hierarchy"""
        values = self.backupAnnotationDialog.getValues()
        if values:
            [src, dst] = values
            print(src, dst)
        else:
            return

        l = len(src)
        for root, dirs, files in os.walk(src):
            for d in dirs:
                # print(dst,root,dirs)
                if not os.path.exists(os.path.join(dst, root[l + 1 :], d)):
                    os.mkdir(os.path.join(dst, root[l + 1 :], d))
            for f in files:
                if (
                    f[-5:].lower() == ".data"
                    or "corrections" in f
                    or "BatData" in f
                    or "BatPasses" in f
                ):
                    shutil.copy2(
                        os.path.join(root, f), os.path.join(dst, root[l + 1 :])
                    )
        self.backupAnnotationDialog.close()

        # try:
        # if platform.system() == 'Windows':
        # subprocess.call(['xcopy', src+'\*.data', dst, '/s', '/e'])
        # elif platform.system() == 'Linux' or platform.system() == 'Darwin':     # TODO: zero testing!
        # except Exception as e:
        # print("Warning: Coping failed with error:")
        # print(e)
        # return

    def exportExcel(self):
        """Launched manually by pressing the button.
        Cleans out old excels and creates a single new one.
        Needs set self.species, self.SoundFileDir."""

        # self.species = self.w_spe1.currentText()
        if self.SoundFileDir == "":
            msg = SupportClasses_GUI.MessagePopup(
                "w", "Select Folder", "Please select a folder to process!"
            )
            msg.exec_()
            return

        with pg.BusyCursor():
            # delete old results (xlsx)
            # ! WARNING: any Detection...xlsx files will be DELETED,
            # ! ANYWHERE INSIDE the specified dir, recursively
            self.statusBar().showMessage("Removing old Excel files...")
            self.update()
            self.repaint()
            for root, dirs, files in os.walk(str(self.SoundFileDir)):
                for filename in files:
                    filenamef = os.path.join(root, filename)
                    if fnmatch.fnmatch(filenamef, "*DetectionSummary_*.xlsx"):
                        print("Removing excel file %s" % filenamef)
                        os.remove(filenamef)

        print("Exporting to Excel ...")
        self.statusBar().showMessage("Exporting to Excel ...")
        self.update()
        self.repaint()

        allsegs = []
        # Note: one excel will always be generated for the currently selected species
        # spList = set([self.currentSpecies])

        if self.currentSpecies == "Species":
            segment_data = self.db.get_dir_segments(
                self.SoundFileDir,
                self.confidenceRange,
            )
        else:
            segment_data = self.db.get_dir_species_segments(
                self.SoundFileDir,
                self.currentSpecies,
                self.confidenceRange,
            )

        all_files = []
        for segment in segment_data:
            filepath = os.path.join(segment[0], segment[1])
            if filepath not in all_files:
                all_files.append(filepath)

        with pg.BusyCursor():
            i = 1
            for filename in all_files:
                # print("Reading segments from", filename)
                segments = Segment.SegmentList()
                # segments.parseJSON(filename, silent=True)

                for segment in segment_data:
                    if filename == os.path.join(segment[0], segment[1]):
                        segments.addSegment(
                            [
                                segment[2],
                                segment[3],
                                segment[4],
                                segment[5],
                                [
                                    {
                                        "filter": segment[9],
                                        "certainty": segment[7],
                                        "species": segment[6],
                                        "calltype": segment[8],
                                    }
                                ],
                            ]
                        )

                # sort by time and save
                segments.orderTime()
                # attach filename to be stored in Excel later
                segments.filename = filename

                # Collect all .data contents (as SegmentList objects)
                # for the Excel output (no matter if review dialog exit was clean)
                allsegs.append(segments)
                print("{}/{}".format(i, len(all_files)))
                i += 1

            # Export the actual Excel
            excel = SupportClasses.ExcelIO()
            excsuccess = excel.export(
                allsegs,
                self.SoundFileDir,
                "overwrite",
                simple=True,
                fileperspecies=False,
                species=self.currentSpecies,
                precisionMS=True,
            )

        if excsuccess != 1:
            # if any file wasn't exported well, overwrite the message
            msgtext = (
                "Warning: Excel output at "
                + self.SoundFileDir
                + " was not stored properly"
            )
            print(msgtext)
            msg = SupportClasses_GUI.MessagePopup(
                "w", "Failed to export Excel file", msgtext
            )
        else:
            msgtext = "Excel output is stored in " + os.path.join(
                self.SoundFileDir, "DetectionSummary_*.xlsx"
            )
            msg = SupportClasses_GUI.MessagePopup("d", "Excel output produced", msgtext)
        msg.exec_()

    def exportSeg(self):
        # First, deal with older xls if present:
        foundxls = []
        for f in os.listdir(self.SoundFileDir):
            if f.startswith("DetectionSummary_") and f.endswith(".xlsx"):
                foundxls.append(f)

        if len(foundxls) > 0:
            # check with user
            msg = SupportClasses_GUI.MessagePopup(
                "w",
                "Excel file exists",
                "Detection summaries already present in "
                + self.SoundFileDir
                + ". Overwrite them, append to them, or cancel the operation?",
            )
            msg.setStandardButtons(QMessageBox.Cancel)
            msg.addButton("Overwrite", QMessageBox.YesRole)
            msg.addButton("Append", QMessageBox.YesRole)
            # cancelBtn = msg.addButton(QMessageBox.Cancel)
            reply = msg.exec_()
            # print(reply)
            if reply == 4194304:  # weird const for Cancel
                return
            elif reply == 1:
                action = "append"
            elif reply == 0:
                action = "overwrite"
            else:
                print("ERROR: Unrecognised reply", reply)
                return

            # remove all the old excels:
            if action == "overwrite":
                for f in foundxls:
                    try:
                        os.remove(os.path.join(self.SoundFileDir, f))
                    except Exception as e:
                        print(
                            "Could not remove file", os.path.join(self.SoundFileDir, f)
                        )
                        print(e)
        else:
            # create new workbook, in effect
            action = "overwrite"

        # sort the segments into increasing time order (to make neater output)
        sortOrder = self.segments.orderTime()
        self.listRectanglesa1 = [self.listRectanglesa1[i] for i in sortOrder]
        self.listRectanglesa2 = [self.listRectanglesa2[i] for i in sortOrder]
        self.listLabels = [self.listLabels[i] for i in sortOrder]

        # excel should be split by page size, but for short files just give the file size
        datalen = (
            self.config["maxFileShow"] if self.nFileSections > 1 else self.datalengthSec
        )
        excel = SupportClasses.ExcelIO()
        self.segments.filename = self.filename
        success = excel.export(
            [self.segments],
            self.SoundFileDir,
            action=action,
            pagelenarg=datalen,
            numpages=self.nFileSections,
            startTime=self.startTime,
        )
        # add user notification
        if success == 0:
            print("Warning: Excel output was not saved")
            return
        else:
            msg = SupportClasses_GUI.MessagePopup(
                "d",
                "Segments Exported",
                "Check this directory for the Excel output: "
                + "\n"
                + self.SoundFileDir,
            )
            msg.exec_()
            return

    # ===============
    # Code for playing sounds
    def playVisible(self):
        """Listener for button to play the visible area.
        On PLAY, turns to PAUSE and two other buttons turn to STOPs.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.pausePlayback()
        else:
            if (
                self.media_obj.state() != QAudio.SuspendedState
                and not self.media_obj.keepSlider
            ):
                # restart playback
                start, end = self.p_ampl.viewRange()[0]
                self.setPlaySliderLimits(start * 1000, end * 1000)
                # (else keep play slider range from before)
            else:
                print(
                    "resuming after pause, on segment:",
                    self.segmentStart,
                    self.segmentStop,
                )
            self.bar.setMovable(False)
            self.playButton.setIcon(
                self.style().standardIcon(QtWidgets.QStyle.SP_MediaPause)
            )
            self.playSegButton.setIcon(
                self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
            )
            self.playSlowButton.setIcon(
                self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
            )
            self.playBandLimitedSegButton.setIcon(
                self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
            )

            # OS X doesn't repaint them by default smh
            self.playButton.repaint()
            self.playSegButton.repaint()
            self.playBandLimitedSegButton.repaint()
            self.playSlowButton.repaint()
            QApplication.processEvents()

            # if bar was moved under pause, update the playback
            # start position based on the bar:
            if self.bar.value() > 0:
                start = self.convertSpectoAmpl(self.bar.value()) * 1000  # in ms
                print("found bar at %d ms" % start)
            else:
                start = self.segmentStart
            # (will not be used if resuming without touching the bar)

            self.media_obj.pressedPlay(
                start=start, stop=self.segmentStop, audiodata=self.audiodata
            )

    def playSelectedSegment(self):
        """Listener for PlaySegment button.
        Get selected segment start and end (or return if no segment selected).
        On PLAY, all three buttons turn to STOPs.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()
        else:
            if self.box1id > -1:
                self.stopPlayback()
                # restart playback
                start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
                stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000
                self.setPlaySliderLimits(start, stop)

                self.bar.setMovable(False)
                self.playButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaPause)
                )
                self.playSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playSlowButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playBandLimitedSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )

                # OS X doesn't repaint them by default smh
                self.playButton.repaint()
                self.playSegButton.repaint()
                self.playBandLimitedSegButton.repaint()
                self.playSlowButton.repaint()
                QApplication.processEvents()

                self.media_obj.filterSeg(start, stop, self.audiodata)
            else:
                print("Can't play, no segment selected")

    def playBandLimitedSegment(self):
        """Listener for PlayBandlimitedSegment button.
        Gets the band limits of the segment, bandpass filters, then plays that.
        Currently uses FIR bandpass filter -- Butterworth is commented out.
        On PLAY, all three buttons turn to STOPs.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()
        else:
            if self.box1id > -1:
                self.stopPlayback()
                # check frequency limits, + small buffer bands
                bottom = max(0.1, self.sp.minFreq, self.segments[self.box1id][2])
                top = min(self.segments[self.box1id][3], self.sp.maxFreq - 0.1)

                print("Extracting samples between %d-%d Hz" % (bottom, top))
                start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
                stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000
                self.setPlaySliderLimits(start, stop)
                self.bar.setMovable(False)
                self.playButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaPause)
                )
                self.playSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playSlowButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playBandLimitedSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )

                # OS X doesn't repaint them by default smh
                self.playButton.repaint()
                self.playSegButton.repaint()
                self.playBandLimitedSegButton.repaint()
                self.playSlowButton.repaint()
                QApplication.processEvents()

                # filter the data into a temporary file or buffer
                self.media_obj.filterBand(
                    self.segmentStart,
                    self.segmentStop,
                    bottom,
                    top,
                    self.audiodata,
                    self.sp,
                )
            else:
                print("Can't play, no segment selected")

    def playSlowSegment(self):
        """Listener for PlaySlowSegment button.
        Very similar to play selected.
        """
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            self.stopPlayback()
        else:
            if self.box1id > -1:
                self.stopPlayback()

                # Times in milliseconds
                start = self.listRectanglesa1[self.box1id].getRegion()[0] * 1000
                stop = self.listRectanglesa1[self.box1id].getRegion()[1] * 1000

                self.setPlaySliderLimits(start, stop)
                self.bar.setMovable(False)
                self.playButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaPause)
                )
                self.playSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playSlowButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )
                self.playBandLimitedSegButton.setIcon(
                    self.style().standardIcon(QtWidgets.QStyle.SP_MediaStop)
                )

                # OS X doesn't repaint them by default smh
                self.playButton.repaint()
                self.playSegButton.repaint()
                self.playBandLimitedSegButton.repaint()
                self.playSlowButton.repaint()
                QApplication.processEvents()

                # filter the data into a temporary file or buffer
                # Note the offset
                # print(start,stop,self.slowSpeed,int(start*self.slowSpeed), int(stop*self.slowSpeed))
                self.media_slow.filterSeg(
                    int(start * self.slowSpeed),
                    int(stop * self.slowSpeed),
                    self.audiodata,
                )
            else:
                print("Can't play, no segment selected")

    def pausePlayback(self):
        """Restores the PLAY buttons, calls media_obj to pause playing."""
        self.media_obj.pressedPause()
        self.media_slow.pressedStop()
        self.bar.setMovable(True)

        # Reset all button icons:
        self.playButton.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaPlay)
        )
        self.playSegButton.setIcon(QIcon("img/playsegment.png"))
        self.playSlowButton.setIcon(QIcon("img/playSlow-w.png"))
        self.playBandLimitedSegButton.setIcon(QIcon("img/playBandLimited.png"))

        # OS X doesn't repaint them by default smh
        self.playButton.repaint()
        self.playSegButton.repaint()
        self.playBandLimitedSegButton.repaint()
        self.playSlowButton.repaint()
        QApplication.processEvents()

    def stopPlayback(self):
        """Restores the PLAY buttons, slider, text, calls media_obj to stop playing."""
        self.bar.setMovable(True)
        self.media_obj.pressedStop()
        self.media_slow.pressedStop()
        if not hasattr(self, "segmentStart") or self.segmentStart is None:
            self.segmentStart = 0
        self.bar.setValue(-1000)

        # Reset all button icons:
        self.playButton.setIcon(
            self.style().standardIcon(QtWidgets.QStyle.SP_MediaPlay)
        )
        self.playSegButton.setIcon(QIcon("img/playsegment.png"))
        self.playSlowButton.setIcon(QIcon("img/playSlow-w.png"))
        self.playBandLimitedSegButton.setIcon(QIcon("img/playBandLimited.png"))

        # OS X doesn't repaint them by default smh
        self.playButton.repaint()
        self.playSegButton.repaint()
        self.playBandLimitedSegButton.repaint()
        self.playSlowButton.repaint()
        QApplication.processEvents()

    def movePlaySlider(self):
        """Listener called on sound notify (every 20 ms).
        Controls the slider, text timer, and listens for playback finish.
        """
        eltime = self.media_obj.processedUSecs() // 1000 + self.media_obj.timeoffset
        bufsize = 0.02

        # listener for playback finish. Note small buffer for catching up
        if eltime > (self.segmentStop - 10):
            # TODO: allow the user set looping somehow?
            if self.media_obj.loop:
                self.media_obj.restart()
            else:
                print("Stopped at %d ms" % eltime)
                self.stopPlayback()
        else:
            self.bar.setValue(self.convertAmpltoSpec(eltime / 1000.0 - bufsize))

    def movePlaySlowSlider(self):
        """Listener called on sound notify (every 20 ms).
        Controls the slider, text timer, and listens for playback finish.
        Very similar to previous, but slightly easier just to reproduce the code.
        """
        eltime = (
            self.media_slow.processedUSecs() // 1000 + self.media_slow.timeoffset
        ) // self.slowSpeed
        bufsize = 0.02

        # listener for playback finish. Note small buffer for catching up
        if eltime > (self.segmentStop - 10):
            print("Stopped at %d ms" % eltime)
            self.stopPlayback()
        else:
            self.bar.setValue(self.convertAmpltoSpec(eltime / 1000.0 - bufsize))

    def setPlaySliderLimits(self, start, end):
        """Uses start/end in ms, relative to page start"""
        self.segmentStart = start
        self.segmentStop = end

    def volSliderMoved(self, value):
        self.media_obj.applyVolSlider(value)
        self.media_slow.applyVolSlider(value)

    def barMoved(self, evt=None):
        """Listener for when the bar showing playback position moves.
        Resets both QAudioOutputs so that they don't try to resume
        """
        print("Resetting playback")
        self.media_obj.reset()
        self.media_slow.reset()

    def setOperatorReviewerDialog(self):
        """Listener for Set Operator/Reviewer menu item."""
        if hasattr(self, "operator") and hasattr(self, "reviewer"):
            self.operatorReviewerDialog = Dialogs.OperatorReviewer(
                operator=self.operator, reviewer=self.reviewer
            )
        else:
            self.operatorReviewerDialog = Dialogs.OperatorReviewer(
                operator="", reviewer=""
            )
        self.operatorReviewerDialog.activate.clicked.connect(self.changeOperator)
        self.operatorReviewerDialog.exec()

    def changeOperator(self):
        """Listener for the operator/reviewer dialog."""
        name1, name2 = self.operatorReviewerDialog.getValues()
        self.operator = str(name1)
        self.reviewer = str(name2)
        self.statusRight.setText(
            "Operator: " + self.operator + ", Reviewer: " + self.reviewer
        )
        self.operatorReviewerDialog.close()
        self.config["operator"] = self.operator
        self.config["reviewer"] = self.reviewer
        self.saveConfig = True

    def classifyBirdNET(self):
        self.BirdNETDialog = BirdNET.BirdNETDialog(self)
        self.BirdNETDialog.exec_()

    def addNoiseData(self):
        """Listener for the adding metadata about noise action"""
        # this field isn't required and may not be present at all
        if "noiseLevel" not in self.segments.metadata:
            self.segments.metadata["noiseLevel"] = None
        if "noiseTypes" not in self.segments.metadata:
            self.segments.metadata["noiseTypes"] = []

        self.getNoiseDataDialog = Dialogs.addNoiseData(
            self.segments.metadata["noiseLevel"], self.segments.metadata["noiseTypes"]
        )
        self.getNoiseDataDialog.activate.clicked.connect(self.getNoiseData)
        self.getNoiseDataDialog.exec()

    def getNoiseData(self):
        """Collect data about the noise from the dialog"""
        (
            self.segments.metadata["noiseLevel"],
            self.segments.metadata["noiseTypes"],
        ) = self.getNoiseDataDialog.getNoiseData()
        # print(self.segments.metadata)
        self.getNoiseDataDialog.close()
        self.segmentsToSave = True

    def saveImage(self, imageFile=""):
        exporter = pge.ImageExporter(self.w_spec.scene())

        if imageFile == "":
            imageFile, drop = QFileDialog.getSaveFileName(
                self, "Save Image", "", "Images (*.png *.xpm *.jpg)"
            )
            if not (
                imageFile.endswith(".png")
                or imageFile.endswith(".xpm")
                or imageFile.endswith(".jpg")
            ):
                # exporter won't be able to deduce file type and will quit silently
                imageFile = imageFile + ".png"
        try:
            # works but requires devel (>=0.11) version of pyqtgraph:
            exporter.export(imageFile)
            print("Exporting spectrogram to file %s" % imageFile)
        except Exception as e:
            print("Warning: failed to save image")
            print(e)

    def saveImageRaw(self):
        imageFile = self.filename[:-4] + ".png"
        print("Exporting raw spectrogram to file %s" % imageFile)
        self.specPlot.save(imageFile)

    def exportFiles(self):
        """Copy files with selected species and segments with confidence > selected confidence to specified folder"""
        fileDirList = self.db.get_files_with_species(
            self.currentSpecies,
            self.SoundFileDir,
            self.confidenceRange,
        )
        self.exportFilelist = [os.path.join(dir, file) for file, dir in fileDirList]
        print(self.exportFilelist)
        self.exportDialog = Dialogs.ExportFilesDialog(self)
        self.exportDialog.show()
        self.exportDialog.activateWindow()
        self.exportDialog.btnCopyFiles.clicked.connect(self.copyFiles)

    def prepareExportData(self, segments):
        exportData = {}
        for seg in segments:
            file = os.path.join(seg[0], seg[1])
            seg = [
                seg[2],
                seg[3],
                seg[4],
                seg[5],
                [
                    {
                        "species": seg[6],
                        "certainty": seg[7],
                        "calltype": seg[8],
                        "filter": seg[9],
                    }
                ],
            ]
            if file in exportData.keys():
                exportData[file].append(seg)
            else:
                exportData[file] = [seg]

        return exportData

    def exportAvianzData(self):
        if self.currentSpecies == "Species":
            segments = self.db.get_dir_segments(self.SoundFileDir, self.confidenceRange)
        else:
            segments = self.db.get_dir_species_segments(
                self.SoundFileDir, self.currentSpecies, self.confidenceRange
            )

        to_export = self.prepareExportData(segments)
        for seg in to_export:
            seglist = Segment.SegmentList()
            seglist.metadata["Duration"] = 0
            seglist.metadata["Operator"] = self.operator
            seglist.metadata["Reviewer"] = self.reviewer

            for s in to_export[seg]:
                seglist.addSegment(s)
            seglist.saveJSON(seg + ".data")

    def exportDatabase(self):
        exportDB, _ = QFileDialog.getSaveFileName(
            self,
            "Choose Annotation Database",
            os.path.dirname(self.db.db_path),
            "Database files (*.db)",
            "",
            QFileDialog.Option.DontConfirmOverwrite,
        )
        # update database filepath and database
        if exportDB:
            exportDBHandler = Database.DatabaseHandler(exportDB)

        if self.currentSpecies == "Species":
            segments = self.db.get_dir_segments(self.SoundFileDir, self.confidenceRange)
        else:
            segments = self.db.get_dir_species_segments(
                self.SoundFileDir, self.currentSpecies, self.confidenceRange
            )

        to_export = self.prepareExportData(segments)

        for file in to_export:
            seglist = Segment.SegmentList()
            seglist.metadata["Duration"] = 0
            seglist.metadata["Operator"] = self.operator
            seglist.metadata["Reviewer"] = self.reviewer

            for s in to_export[file]:
                seglist.addSegment(s)

            exportDBHandler.insert_segments(seglist, file)

        msg = SupportClasses_GUI.MessagePopup(
            "d",
            "Success",
            "The data was exported to {}".format(os.path.basename(exportDB)),
        )

        msg.exec_()
        return

    def exportAudioSegments(self):
        """Listener for "Export segments" from actions menu;
        exports segments for selected species to specified folder"""

        if self.currentSpecies == "Species":
            fileDirList = self.db.get_grouped_dir_segments(
                self.SoundFileDir,
                self.confidenceRange,
            )
        else:
            fileDirList = self.db.get_grouped_dir_species_segments(
                self.SoundFileDir,
                self.currentSpecies,
                self.confidenceRange,
            )
        self.exportSegmentsDict = {}
        for entry in fileDirList:
            filename = os.path.join(entry[0], entry[1])
            segment = (entry[2], entry[3], entry[4], entry[5])
            if filename in self.exportSegmentsDict.keys():
                self.exportSegmentsDict[filename].append(segment)
            else:
                self.exportSegmentsDict[filename] = [segment]
        self.exportDialog = Dialogs.ExportSegmentsDialog(self, len(fileDirList))
        self.exportDialog.show()
        self.exportDialog.activateWindow()
        self.exportDialog.btnCopySegments.clicked.connect(self.copySegments)

    def copyFiles(self):
        for src in self.exportFilelist:
            dst = os.path.join(
                self.exportDialog.txtDst.text(),
                src[len(os.path.commonpath([self.SoundFileDir, src])) + 1 :],
            )
            pathlib.Path(os.path.dirname(dst)).mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dst)
        self.exportDialog.close()

    def copySegments(self):
        """Reads the audio files from self.exportSegmentsDict, extracts the segments and saves them as individual audio files"""
        dst = self.exportDialog.txtDst.text()
        for src in self.exportSegmentsDict:
            # read audio file
            audio = wavio.read(src)
            for seg in self.exportSegmentsDict[src]:
                # concatenate destination file path
                file_dst = os.path.join(
                    dst,
                    seg[2],
                    "{}_{:.1f}-{:.1f}.wav".format(
                        pathlib.Path(os.path.basename(src)).stem, seg[0], seg[1]
                    ),
                )
                if not os.path.exists(file_dst):
                    if not os.path.exists(os.path.dirname(file_dst)):
                        os.mkdir(os.path.dirname(file_dst))
                    wavio.write(
                        file_dst,
                        audio.data[int(seg[0] * audio.rate) : int(seg[1] * audio.rate)],
                        audio.rate,
                    )

        self.exportDialog.close()

    def importAvianzData(self):
        self.tempsl = Segment.SegmentList()
        filenames, _ = QFileDialog.getOpenFileNames(
            self,
            "Choose AviaNZ annotation files",
            self.SoundFileDir,
            "AviaNZ annotation files (*.data)",
        )
        for filename in filenames:
            if os.path.isfile(filename):
                print("Update entries for {}".format(filename[:-5]), end="")
                print("\r", end="")
                self.tempsl.parseJSON(filename, silent=True)
                if len(self.tempsl) > 0:
                    self.db.insert_segments(
                        self.tempsl,
                        filename[:-5],
                    )
        self.db.commit()
        print("Database successfully updated.")
        self.updateDirectory()

    def importRavenData(self):
        """Imports data from individual raven selection tables"""

        file_names, _ = QFileDialog.getOpenFileNames(
            self,
            "Choose Raven selection tables",
            self.SoundFileDir,
            "Raven selection tables (*.txt)",
        )
        seglistdir = {}
        for filename in file_names:
            # TODO: How to set filter -> manually? by name?
            if len(os.path.basename(filename).split(".")) == 5:
                filter = os.path.basename(filename).split(".")[-4]
            else:
                filter = "BirdNET-Analyzer"

            # create a directory of segment list. The keys are the filenames, the entries the lists of segments
            with open(filename, "r") as infile:
                next(infile)
                for line in infile:
                    l = line.split("\t")
                    audio = os.path.abspath(
                        os.path.join(os.path.dirname(l[10]), os.path.basename(l[10]))
                    )
                    print(audio)
                    s = Segment.Segment(
                        [
                            float(l[3]),  # start
                            float(l[4]),  # end
                            float(l[5]),  # low
                            float(l[6]),  # high
                            [
                                {
                                    "species": l[7],
                                    "certainty": float(l[9]) * 100,
                                    "filter": filter,
                                }
                            ],
                        ]
                    )
                    if audio not in seglistdir.keys():
                        seglist = Segment.SegmentList()
                        seglist.metadata = {
                            "Reviewer": self.reviewer,
                            "Operator": self.operator,
                        }
                        seglistdir[audio] = seglist
                    seglistdir[audio].addSegment(s)

        # insert the annotations into the database.
        for file in seglistdir:
            self.db.insert_segments(seglistdir[file], file)

        self.db.commit()
        print("Database successfully updated.")
        self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))

    def deleteDirAnnotations(self):
        self.db.delete_dir_segments(self.SoundFileDir)
        self.db.commit()
        self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))
        self.loadFile(self.filename)

    def updateDirectory(self):
        for root, dir, files in os.walk(self.SoundFileDir):
            for file in files:
                if file.endswith(".wav"):
                    print("root: {}, filename: {}".format(root, file))
                    self.db.add_file(file, root)
        self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))
        self.db.commit()
        self.loadFile(self.filename)

    def chooseDatabase(self):
        file, _ = QFileDialog.getSaveFileName(
            self,
            "Choose Annotation Database",
            os.path.dirname(self.db.db_path),
            "Database files (*.db)",
            "",
            QFileDialog.Option.DontConfirmOverwrite,
        )
        # update database filepath and database
        if file:
            self.dbPath = file
            self.db = None
            self.db = Database.DatabaseHandler(self.dbPath)
            self.listFiles.database = self.db

            # make sure the new database is saved to config
            self.saveConfig = True
            self.config["Database"] = self.dbPath
            self.statusDB.setText(os.path.basename(self.dbPath))

            # reload the list of files to update the file icons
            self.fillFileList(self.SoundFileDir, os.path.basename(self.filename))
            self.loadFile(self.filename)

    def changeSettings(self):
        """Create the parameter tree when the Interface settings menu is pressed."""
        fn1 = self.config["BirdListShort"]
        if "/" in fn1:
            fn1 = os.path.basename(fn1)
        fn2 = self.config["BirdListLong"]
        if fn2 is not None and "/" in fn2:
            fn2 = os.path.basename(fn2)
        fn3 = self.config["BatList"]
        if fn3 is not None and "/" in fn3:
            fn3 = os.path.basename(fn3)
        hasMultipleSegments = False
        for s in self.segments:
            if len(s[4]) > 1:
                hasMultipleSegments = True

        params = [
            {
                "name": "Mouse settings",
                "type": "group",
                "children": [
                    {
                        "name": "Use right button to make segments",
                        "type": "bool",
                        "tip": "If true, segments are drawn with right clicking.",
                        "value": self.config["drawingRightBtn"],
                    },
                    {
                        "name": "Spectrogram mouse action",
                        "type": "list",
                        "limits": {
                            "Mark segments by clicking": 1,
                            "Mark boxes by clicking": 2,
                            "Mark boxes by dragging": 3,
                        },
                        "value": self.config["specMouseAction"],
                    },
                ],
            },
            {
                "name": "Paging",
                "type": "group",
                "children": [
                    {
                        "name": "Page size",
                        "type": "float",
                        "value": self.config["maxFileShow"],
                        "limits": (5, 3600),
                        "step": 5,
                        "suffix": " sec",
                    },
                    {
                        "name": "Page overlap",
                        "type": "float",
                        "value": self.config["fileOverlap"],
                        "limits": (0, 20),
                        "step": 2,
                        "suffix": " sec",
                    },
                ],
            },
            {
                "name": "Annotation",
                "type": "group",
                "children": [
                    {
                        "name": "Annotation overview cell length",
                        "type": "float",
                        "value": self.config["widthOverviewSegment"],
                        "limits": (5, 300),
                        "step": 5,
                        "suffix": " sec",
                    },
                    {
                        "name": "Make boxes transparent",
                        "type": "bool",
                        "value": self.config["transparentBoxes"],
                    },
                    {
                        "name": "Auto save segments every",
                        "type": "float",
                        "value": self.config["secsSave"],
                        "step": 5,
                        "limits": (5, 900),
                        "suffix": " sec",
                    },
                    {
                        "name": "Segment colours",
                        "type": "group",
                        "children": [
                            {
                                "name": "Confirmed segments",
                                "type": "color",
                                "value": self.config["ColourNamed"],
                                "tip": "Correctly labeled segments",
                            },
                            {
                                "name": "Possible",
                                "type": "color",
                                "value": self.config["ColourPossible"],
                                "tip": "Segments that need further approval",
                            },
                            {
                                "name": "Don't know",
                                "type": "color",
                                "value": self.config["ColourNone"],
                                "tip": "Segments that are not labelled",
                            },
                            {
                                "name": "Currently selected",
                                "type": "color",
                                "value": self.config["ColourSelected"],
                                "tip": "Currently selected segment",
                            },
                        ],
                    },
                    {
                        "name": "Custom colour map",
                        "type": "colormap",
                        "value": pg.ColorMap(
                            self.config["ccmap"][0], self.config["ccmap"][1]
                        ),
                        "tip": "Custom colour map in View -> Choose colour map",
                    },
                    {
                        "name": "Guidelines",
                        "type": "group",
                        "children": [
                            {
                                "name": "Show frequency guides",
                                # "type": "list",
                                # "limits": {
                                #     "Always": "always",
                                #     "For bats only": "bat",
                                #     "Never": "never",
                                # },
                                "type": "bool",
                                "value": self.config["guidelinesOn"],
                            },
                            {
                                "name": "Guideline 1 frequency",
                                "type": "float",
                                "value": self.config["guidepos"][0] / 1000,
                                "limits": (0, 1000),
                                "suffix": " kHz",
                            },
                            {
                                "name": "Guideline 1 colour",
                                "type": "color",
                                "value": self.config["guidecol"][0],
                            },
                            {
                                "name": "Guideline 2 frequency",
                                "type": "float",
                                "value": self.config["guidepos"][1] / 1000,
                                "limits": (0, 1000),
                                "suffix": " kHz",
                            },
                            {
                                "name": "Guideline 2 colour",
                                "type": "color",
                                "value": self.config["guidecol"][1],
                            },
                            {
                                "name": "Guideline 3 frequency",
                                "type": "float",
                                "value": self.config["guidepos"][2] / 1000,
                                "limits": (0, 1000),
                                "suffix": " kHz",
                            },
                            {
                                "name": "Guideline 3 colour",
                                "type": "color",
                                "value": self.config["guidecol"][2],
                            },
                            {
                                "name": "Guideline 4 frequency",
                                "type": "float",
                                "value": self.config["guidepos"][3] / 1000,
                                "limits": (0, 1000),
                                "suffix": " kHz",
                            },
                            {
                                "name": "Guideline 4 colour",
                                "type": "color",
                                "value": self.config["guidecol"][3],
                            },
                        ],
                    },
                    {
                        "name": "Check-ignore protocol",
                        "type": "group",
                        "children": [
                            {
                                "name": "Show check-ignore marks",
                                "type": "bool",
                                "value": self.config["protocolOn"],
                            },
                            {
                                "name": "Length of checking zone",
                                "type": "float",
                                "value": self.config["protocolSize"],
                                "limits": (1, 300),
                                "step": 1,
                                "suffix": " sec",
                            },
                            {
                                "name": "Repeat zones every",
                                "type": "float",
                                "value": self.config["protocolInterval"],
                                "limits": (1, 600),
                                "step": 1,
                                "suffix": " sec",
                            },
                        ],
                    },
                ],
            },
            {
                "name": "Bird List",
                "type": "group",
                "children": [
                    {
                        "name": "Common Bird List",
                        "type": "group",
                        "children": [
                            # {'name': 'Filename', 'type': 'text', 'value': self.config['BirdListShort']},
                            {
                                "name": "Filename",
                                "type": "str",
                                "value": fn1,
                                "readonly": True,
                            },
                            {"name": "Choose File", "type": "action"},
                        ],
                    },
                    {
                        "name": "Full Bird List",
                        "type": "group",
                        "children": [
                            # {'name': 'Filename', 'type': 'str', 'value': fn2,'readonly':True, 'tip': "Can be None"},
                            {
                                "name": "Filename",
                                "type": "str",
                                "value": fn2,
                                "readonly": True,
                            },
                            # {'name': 'No long list', 'type': 'bool',
                            #'value': self.config['BirdListLong'] is None or self.config['BirdListLong'] == 'None',
                            #'tip': "If you don't have a long list of birds"},
                            {"name": "Choose File", "type": "action"},
                        ],
                    },
                    {
                        "name": "Bat List",
                        "type": "group",
                        "children": [
                            {
                                "name": "Filename",
                                "type": "str",
                                "value": fn3,
                                "readonly": True,
                            },
                            {"name": "Choose File", "type": "action"},
                        ],
                    },
                    {
                        "name": "Dynamically reorder bird list",
                        "type": "bool",
                        "value": self.config["ReorderList"],
                    },
                    {
                        "name": "Default to multiple species",
                        "type": "bool",
                        "value": self.config["MultipleSpecies"],
                        "readonly": hasMultipleSegments,
                    },
                ],
            },
            {
                "name": "User",
                "type": "group",
                "children": [
                    {
                        "name": "Operator",
                        "type": "str",
                        "value": self.config["operator"],
                        "tip": "Person name",
                    },
                    {
                        "name": "Reviewer",
                        "type": "str",
                        "value": self.config["reviewer"],
                        "tip": "Person name",
                    },
                ],
            },
            {
                "name": "Maximise window on startup",
                "type": "bool",
                "value": self.config["StartMaximized"],
            },
            {
                "name": "Require noise data",
                "type": "bool",
                "value": self.config["RequireNoiseData"],
            },
        ]
        self.settingsChanged = False
        self.updateFileIcons = False
        ## Create tree of Parameter objects
        self.p = Parameter.create(name="params", type="group", children=params)
        self.p.sigTreeStateChanged.connect(self.changeParams)
        ## Create ParameterTree widget
        self.t = SupportClasses_GUI.ParameterTreeWithClose()
        self.t.setParameters(self.p, showTop=False)
        self.t.treeclosed.connect(self.reloadFile)
        self.t.show()
        self.t.setWindowTitle("PAMalyzer - Interface Settings")
        self.t.setWindowIcon(QIcon("img/PAMalyzer.ico"))
        self.t.setFixedHeight(900)
        self.t.setMinimumWidth(520)

    def changeParams(self, param, changes):
        """Update the config and the interface if anything changes in the tree"""

        # some regexes to parse guideline settings
        rgx_guide_pos = re.compile(r"Annotation.Guidelines.Guideline ([0-9]) frequency")
        rgx_guide_col = re.compile(r"Annotation.Guidelines.Guideline ([0-9]) colour")

        for param, change, data in changes:
            path = self.p.childPath(param)
            if path is not None:
                childName = ".".join(path)
            else:
                childName = param.name()

            if childName == "Output parameters.Auto save segments every":
                self.config["secsSave"] = data
            elif childName == "Annotation.Annotation overview cell length":
                self.config["widthOverviewSegment"] = data

            elif childName == "Annotation.Make boxes transparent":
                self.config["transparentBoxes"] = data
                self.dragRectsTransparent()
            elif childName == "Mouse settings.Use right button to make segments":
                self.config["drawingRightBtn"] = data
                if self.config["drawingRightBtn"]:
                    self.MouseDrawingButton = Qt.RightButton
                    self.specPlot.unsetCursor()
                    self.p_ampl.unsetCursor()
                    self.bar.setCursor(Qt.OpenHandCursor)
                else:
                    self.MouseDrawingButton = Qt.LeftButton
                    self.bar.unsetCursor()
                    self.specPlot.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                    self.p_ampl.setCursor(
                        QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                    )
                self.bar.btn = self.MouseDrawingButton
            elif childName == "Mouse settings.Spectrogram mouse action":
                self.config["specMouseAction"] = data
                self.p_spec.enableDrag = data == 3 and not self.readonly.isChecked()
            elif childName == "Paging.Page size":
                self.config["maxFileShow"] = data
            elif childName == "Paging.Page overlap":
                self.config["fileOverlap"] = data
            elif childName == "Maximise window on startup":
                self.config["StartMaximized"] = data
                if data:
                    self.showMaximized()
            elif childName == "Bird List.Dynamically reorder bird list":
                self.config["ReorderList"] = data
            elif childName == "Bird List.Default to multiple species":
                self.config["MultipleSpecies"] = data
            elif childName == "Require noise data":
                self.config["RequireNoiseData"] = data
            elif childName == "Bird List.Common Bird List.Filename":
                self.config["BirdListShort"] = data
            elif childName == "Bird List.Full Bird List.Filename":
                self.config["BirdListLong"] = data
            elif childName == "Bird List.Bat List.Filename":
                self.config["BatList"] = data
            elif childName == "Annotation.Segment colours.Confirmed segments":
                rgbaNamed = list(data.getRgb())
                if rgbaNamed[3] > 100:
                    rgbaNamed[3] = 100
                self.config["ColourNamed"] = rgbaNamed
                self.ColourNamed = QtGui.QColor(
                    self.config["ColourNamed"][0],
                    self.config["ColourNamed"][1],
                    self.config["ColourNamed"][2],
                    self.config["ColourNamed"][3],
                )
                self.ColourNamedDark = QtGui.QColor(
                    self.config["ColourNamed"][0],
                    self.config["ColourNamed"][1],
                    self.config["ColourNamed"][2],
                    255,
                )
                self.listFiles.ColourNamed = self.ColourNamed
                self.updateFileIcons = True
            elif childName == "Annotation.Segment colours.Possible":
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config["ColourPossible"] = rgbaVal
                self.ColourPossible = QtGui.QColor(
                    self.config["ColourPossible"][0],
                    self.config["ColourPossible"][1],
                    self.config["ColourPossible"][2],
                    self.config["ColourPossible"][3],
                )
                self.ColourPossibleDark = QtGui.QColor(
                    self.config["ColourPossible"][0],
                    self.config["ColourPossible"][1],
                    self.config["ColourPossible"][2],
                    255,
                )
                self.listFiles.ColourPossibleDark = self.ColourPossibleDark
                self.updateFileIcons = True
            elif childName == "Annotation.Segment colours.Don't know":
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config["ColourNone"] = rgbaVal
                self.ColourNone = QtGui.QColor(
                    self.config["ColourNone"][0],
                    self.config["ColourNone"][1],
                    self.config["ColourNone"][2],
                    self.config["ColourNone"][3],
                )
                self.ColourNoneDark = QtGui.QColor(
                    self.config["ColourNone"][0],
                    self.config["ColourNone"][1],
                    self.config["ColourNone"][2],
                    255,
                )
                self.listFiles.ColourNone = self.ColourNone
                self.updateFileIcons = True
            elif childName == "Annotation.Segment colours.Currently selected":
                rgbaVal = list(data.getRgb())
                if rgbaVal[3] > 100:
                    rgbaVal[3] = 100
                self.config["ColourSelected"] = rgbaVal
                # update the interface
                self.ColourSelected = QtGui.QColor(
                    self.config["ColourSelected"][0],
                    self.config["ColourSelected"][1],
                    self.config["ColourSelected"][2],
                    self.config["ColourSelected"][3],
                )
                self.ColourSelectedDark = QtGui.QColor(
                    self.config["ColourSelected"][0],
                    self.config["ColourSelected"][1],
                    self.config["ColourSelected"][2],
                    255,
                )
            elif childName == "Annotation.Custom colour map":
                self.config["ccmap"] = [
                    data.getStops()[0].tolist(),
                    data.getStops()[1].tolist(),
                ]
                self.setColourMap("Custom")
            elif childName == "Annotation.Guidelines.Show frequency guides":
                self.config["guidelinesOn"] = data
                self.drawGuidelines()
            elif rgx_guide_pos.match(
                childName
            ):  # childName=='Annotation.Guidelines.Guideline 1 frequency':
                guideid = int(rgx_guide_pos.search(childName).group(1)) - 1
                self.config["guidepos"][guideid] = float(data) * 1000
                self.drawGuidelines()
            elif rgx_guide_col.match(
                childName
            ):  # childName=='Annotation.Guidelines.Guideline 1 colour':
                guideid = int(rgx_guide_col.search(childName).group(1)) - 1
                print(data)
                self.config["guidecol"][guideid] = data
                self.drawGuidelines()
            elif (
                childName == "Annotation.Check-ignore protocol.Show check-ignore marks"
            ):
                self.config["protocolOn"] = data
                self.drawProtocolMarks()
            elif (
                childName == "Annotation.Check-ignore protocol.Length of checking zone"
            ):
                self.config["protocolSize"] = data
                self.drawProtocolMarks()
            elif childName == "Annotation.Check-ignore protocol.Repeat zones every":
                self.config["protocolInterval"] = data
                self.drawProtocolMarks()
            elif childName == "User.Operator":
                self.config["operator"] = data
                self.operator = data
                self.statusRight.setText(
                    "Operator: "
                    + str(self.operator)
                    + ", Reviewer: "
                    + str(self.reviewer)
                )
            elif childName == "User.Reviewer":
                self.config["reviewer"] = data
                self.reviewer = data
                self.statusRight.setText(
                    "Operator: "
                    + str(self.operator)
                    + ", Reviewer: "
                    + str(self.reviewer)
                )
            elif childName == "Bird List.Common Bird List.Choose File":
                filename, drop = QFileDialog.getOpenFileName(
                    self,
                    "Choose Common Bird List",
                    self.SoundFileDir,
                    "Text files (*.txt)",
                )
                if filename == "":
                    print("no list file selected")
                    return
                else:
                    self.shortBirdList = self.ConfigLoader.shortbl(
                        filename, self.configdir
                    )
                    if self.shortBirdList is not None:
                        self.config["BirdListShort"] = filename
                        self.p["Bird List", "Common Bird List", "Filename"] = filename
                    else:
                        self.shortBirdList = self.ConfigLoader.shortbl(
                            self.config["BirdListShort"], self.configdir
                        )
            elif childName == "Bird List.Full Bird List.Choose File":
                filename, drop = QFileDialog.getOpenFileName(
                    self,
                    "Choose Full Bird List",
                    self.SoundFileDir,
                    "Text files (*.txt)",
                )
                if filename == "":
                    print("no list file selected")
                    return
                else:
                    self.longBirdList = self.ConfigLoader.longbl(
                        filename, self.configdir
                    )
                    if self.longBirdList is not None:
                        self.config["BirdListLong"] = filename
                        self.p["Bird List", "Full Bird List", "Filename"] = filename
                    else:
                        self.longBirdList = self.ConfigLoader.longbl(
                            self.config["BirdListLong"], self.configdir
                        )
            elif childName == "Bird List.Bat List.Choose File":
                filename, drop = QFileDialog.getOpenFileName(
                    self, "Choose Bat List", self.SoundFileDir, "Text files (*.txt)"
                )
                if filename == "":
                    print("no list file selected")
                    return
                else:
                    self.batList = self.ConfigLoader.batl(filename, self.configdir)
                    if self.batList is not None:
                        self.config["BatList"] = filename
                        self.p["Bird List", "Bat List", "Filename"] = filename
                    else:
                        self.batList = self.ConfigLoader.batl(
                            self.config["BatList"], self.configdir
                        )
                    # self.p['Bird List','Full Bird List','No long list'] = False
            # elif childName=='Bird List.Full Bird List.No long list':
            # if param.value():
            # self.config['BirdListLong'] = 'None'
            # self.p['Bird List','Full Bird List','Filename'] = 'None'
            # self.longBirdList = None
            # else:
            # if self.p['Bird List','Full Bird List','Filename'] is None or self.p['Bird List','Full Bird List','Filename'] == '' or self.p['Bird List','Full Bird List','Filename'] == 'None':
            # filename, drop = QFileDialog.getOpenFileName(self, 'Choose File', self.SoundFileDir, "Text files (*.txt)")
            # if filename == '':
            # print("no list file selected")
            # return
            # else:
            # self.p['Bird List','Full Bird List','Filename'] = filename
            # self.config['BirdListLong'] = filename
            # self.longBirdList = self.ConfigLoader.longbl(self.config['BirdListLong'], self.configdir)

        self.saveConfig = True
        self.settingsChanged = True

    def reloadFile(self):
        # pass the file name to reset interface properly
        if self.settingsChanged:
            self.saveSegments()
            self.loadFile(self.filename)
        # redraw file icons if colors changed
        if self.updateFileIcons:
            self.listFiles.restrict(
                self.currentSpecies, self.confidenceRange, self.timeRange
            )
            self.updateFileIcons = False

    # ============
    # Various actions: deleting segments, saving, quitting
    def confirmSegment(self):
        """Listener for the Confirm segment button.
        Ups the certainty to 100 on the current segment.
        DO NOT use for All Sp Review, as that one may also change species and
        needs to call refreshOverview with old species.
        """
        id = self.box1id
        print("confirming id:", id)

        if id > -1:
            # force wipe old overview to empty
            self.refreshOverviewWith(self.segments[id], delete=True)

            # raise certainty to 100 on all labels in this seg
            self.segments[id].confirmLabels()

            self.refreshOverviewWith(self.segments[id])
            self.refreshFileColor()
            self.updateText(id)
            self.updateColour(id)
            self.segInfo.setText(self.segments[id].infoString())
            self.db.update_segment_species(
                self.segments[id],
                self.SoundFileDir,
                self.listFiles.currentItem().text(),
                self.reviewer,
            )
            self.db.commit()

    def deleteSegment(self, id=-1, hr=False):
        """Listener for delete segment button, or backspace key. Also called when segments are deleted by the
        human classify dialogs.
        Stops playback immediately in all cases.
        Deletes the segment that is selected, otherwise does nothing.
        Updates the overview segments as well.
        """
        print("deleting id:", id)
        if self.media_obj.isPlaying() or self.media_slow.isPlaying():
            # includes resetting playback buttons
            self.stopPlayback()

        if not hr and id < 0:
            id = self.box1id

        if id > -1:
            self.refreshOverviewWith(self.segments[id], delete=True)

            if self.listRectanglesa1[id] is not None:
                try:
                    self.listRectanglesa1[id].sigRegionChangeFinished.disconnect()
                    self.listRectanglesa2[id].sigRegionChangeFinished.disconnect()
                except:
                    pass
                self.p_ampl.removeItem(self.listRectanglesa1[id])
                self.p_spec.removeItem(self.listRectanglesa2[id])
                self.p_spec.removeItem(self.listLabels[id])
            del self.listLabels[id]
            del self.listRectanglesa1[id]
            del self.listRectanglesa2[id]
            self.db.delete_segment(self.filename, self.SoundFileDir, self.segments[id])
            self.db.commit()
            del self.segments[id]

            self.refreshFileColor()

            self.box1id = -1
            self.segInfo.setText("")
            # reset segment playback buttons
            self.refreshSegmentControls()

    def deleteAll(self):
        """Listener for delete all button.
        Checks if the user meant to do it, then calls removeSegments()
        """
        if len(self.segments) == 0:
            msg = SupportClasses_GUI.MessagePopup(
                "w", "No segments", "No segments to delete"
            )
            msg.exec_()
            return
        else:
            msg = SupportClasses_GUI.MessagePopup(
                "t",
                "Delete All Segments?",
                "Are you sure you want to delete all segments?",
            )
            msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
            reply = msg.exec_()
            if reply == QMessageBox.Yes:
                self.removeSegments()
                self.db.delete_file_segments(self.filename)
                self.db.commit()

            # reset segment playback buttons
            self.refreshSegmentControls()
        self.refreshFileColor()

    def removeSegments(self, delete=True):
        """Remove all the segments in response to the menu selection, or when a new file is loaded."""
        for r in self.listLabels:
            if r is not None:
                self.p_spec.removeItem(r)
        for r in self.listRectanglesa1:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_ampl.removeItem(r)
                except:
                    pass
        for r in self.listRectanglesa2:
            if r is not None:
                try:
                    r.sigRegionChangeFinished.disconnect()
                    self.p_spec.removeItem(r)
                except:
                    pass

        # clear overview boxes and their count trackers
        for ovid in range(len(self.SegmentRects)):
            self.overviewSegments[ovid, :] = 0
            self.SegmentRects[ovid].setBrush(pg.mkBrush("w"))
            self.SegmentRects[ovid].update()

        self.segInfo.setText("")
        if delete:
            if hasattr(self, "segments"):
                self.segments.clear()
            self.listRectanglesa1 = []
            self.listRectanglesa2 = []
            self.listLabels = []
            self.box1id = -1

    def refreshFileColor(self):
        """Extracts the maximum confidence value of every species in the segments,
        sets it as item data of the current item and paints the item."""

        data = {}
        for segment in self.segments:
            for detection in segment[4]:
                sp = detection["species"]
                conf = detection["certainty"]
                data.setdefault(sp, []).append(conf)

        self.listFiles.setItemData(self.listFiles.currentItem(), list(data.items()))
        self.listFiles.currentItem().paint(
            self.confidenceRange, self.currentSpecies, self.timeRange
        )

        # update species list
        self.updateListSpecies()

    def saveSegments(self):
        """Save the segmentation data as a json file.
        Name of the file is the name of the wave file + .data"""

        # def checkSave():
        #     msg = QMessageBox()
        #     msg.setIcon(QMessageBox.Information)
        #     msg.setText("Do you want to save?")
        #     msg.setInformativeText("You didn't identify any segments, are you sure you want to save this annotation?")
        #     msg.setWindowTitle("No segments")
        #     msg.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        #     msg.buttonClicked.connect(msgbtn)
        #     retval = msg.exec_()
        #     print "value of pressed message box button:", retval
        #     return retval

        if self.segmentsToSave:
            self.segments.metadata["Operator"] = self.operator
            self.segments.metadata["Reviewer"] = self.reviewer

            self.segments.saveToDatabase(str(self.filename))

            # refresh this file's icon in file list dock
            self.segmentsToSave = False
            self.statusLeft.setText(
                "Segments saved at " + time.strftime("%X", time.localtime())
            )

    def closeFile(self):
        """Calls the appropriate functions when a file is gently closed (on quit or change of file)."""

        # save noise data if the user requires it
        if self.config["RequireNoiseData"]:
            if (
                "noiseLevel" not in self.segments.metadata
                or self.segments.metadata["noiseLevel"] is None
            ):
                self.addNoiseData()

        self.saveSegments()
        print("Closing", self.filename)

        # update recent files list
        if (
            self.filename is not None
            and self.filename not in self.config["RecentFiles"]
        ):
            self.config["RecentFiles"].append(self.filename)
            if len(self.config["RecentFiles"]) > 4:
                del self.config["RecentFiles"][0]
            # Note: we're making this flag useless as every new file open will update the config
            self.saveConfig = True

        # Add in the operator and reviewer at the top, and then save the segments and the config file.
        if self.saveConfig:
            self.ConfigLoader.configwrite(self.config, self.configfile)

        # Save the shortBirdList
        self.ConfigLoader.blwrite(
            self.shortBirdList, self.config["BirdListShort"], self.configdir
        )

    def closeEvent(self, event=None):
        """Catch the user closing the window by clicking the Close button or otherwise."""
        print("Quitting")
        self.closeFile()
        QApplication.exit(0)

    def backupDatafiles(self):
        # TODO: Can probably be removed
        print("Backing up files in ", self.SoundFileDir)
        listOfDataFiles = QDir(self.SoundFileDir).entryList(["*.data"])
        for file in listOfDataFiles:
            source = self.SoundFileDir + "/" + file
            destination = source[:-5] + ".backup"
            if os.path.isfile(destination):
                pass
                # print(destination," exists, not backing up")
            else:
                # print(source)
                # print(destination," doesn't exist")
                shutil.copyfile(source, destination)

    def eventFilter(self, obj, event):
        """Handles two types of events:
        1) Clicks for the context menu. It allows the user to select
        multiple birds by stopping the menu being closed on first click.
        2) Keyboard presses for spec/ampl plots:
          backspace to delete a segment
          escape to pause playback
          ctrl on Mac to detect right clicks
        """
        if isinstance(obj, QMenu) and event.type() in [
            QtCore.QEvent.MouseButtonRelease
        ]:
            if hasattr(self, "multipleBirds") and self.multipleBirds:
                if obj.activeAction():
                    if not obj.activeAction().menu():
                        # if the selected action does not have a submenu
                        # eat the event, but trigger the function
                        obj.activeAction().trigger()
                        return True
            return QMenu.eventFilter(self, obj, event)
        if isinstance(obj, pg.GraphicsLayoutWidget):
            if event.type() == QtCore.QEvent.KeyPress:
                key = event.key()
                if key == Qt.Key_Backspace or key == Qt.Key_Delete:
                    self.deleteSegment()
                    return True
                elif key == Qt.Key_Escape and (
                    self.media_obj.isPlaying() or self.media_slow.isPlaying()
                ):
                    self.stopPlayback()
                    return True
                elif key == Qt.Key_Meta and platform.system() == "Darwin":
                    # flip to rightMB cursors
                    if self.MouseDrawingButton == Qt.RightButton:
                        self.p_ampl.setCursor(
                            QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                        )
                        self.specPlot.setCursor(
                            QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                        )
                        self.bar.unsetCursor()
                    else:
                        self.p_ampl.unsetCursor()
                        self.specPlot.unsetCursor()
                        self.bar.setCursor(Qt.OpenHandCursor)
                    return True
            elif event.type() == QtCore.QEvent.KeyRelease:
                if event.key() == Qt.Key_Meta and platform.system() == "Darwin":
                    # revert to standard cursors (for leftMB)
                    if self.MouseDrawingButton == Qt.RightButton:
                        self.p_ampl.unsetCursor()
                        self.specPlot.unsetCursor()
                        self.bar.setCursor(Qt.OpenHandCursor)
                    else:
                        self.p_ampl.setCursor(
                            QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                        )
                        self.specPlot.setCursor(
                            QtGui.QCursor(QPixmap("img/cursor.bmp"), 0, 0)
                        )
                        self.bar.unsetCursor()
                    return True
        return False
